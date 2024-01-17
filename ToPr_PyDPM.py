# -*- coding: utf-8 -*

from matplotlib import scale
from numpy.core.numeric import _ones_like_dispatcher
from numpy.distutils.command.config import config
from numpy.lib.function_base import blackman

import pandas as pd

from os import listdir as ls
from os import chdir as cd
from os import getcwd as getcwd
from os import remove as remove
from math import sqrt as sqrt

import matplotlib
matplotlib.use("Agg")
from matplotlib import pyplot as plt
from matplotlib import cm
from matplotlib.colors import ListedColormap, LinearSegmentedColormap
from matplotlib.ticker import PercentFormatter as PF

from openpyxl import Workbook
from openpyxl import worksheet
from openpyxl.utils import get_column_letter as gcl

from numpy import ones as ones
from numpy import arange as arange
from numpy import linspace as linspace
from numpy import log10 as log

from datetime import datetime


from tkinter import *
from tkinter import filedialog


workdir = getcwd()
cTSE_blue = "#00bce7"
cgrey = "#e6e6e6"
clg = "#fafafa"
cwh = "#ffffff"

class tabulka:
    def __init__(self, name, location, file, positionunit, velocityunit, diameterunit, temperatureunit, mfrunit, massunit, frequencyunit, timeunit):
    # region header
        self.name           = name
        self.location       = location
        self.file           = file
        self.source = pd.read_table(self.location + "/" +self.file, sep = "\t")
        self.position       = []
        self.x              = []
        self.y              = []
        self.z              = []
        self.velocity       = []
        self.vx             = []
        self.vy             = []
        self.vz             = []
        self.vmag           = []
        self.diameter       = []
        self.temperature    = []
        self.mfr            = []
        self.mass           = []
        self.frequency      = []
        self.time           = []
        self.cmap_list = ["gist_rainbow", "turbo", "viridis", "jet"]
        self.cmap = self.cmap_list[3]
        self.size = 25
        self.workdir = ""
        self.range_count = 1
        self.hist_labels    = []
        self.unit_multihisto_label = ""
        self.dataset_length  =  len(self.x)
    #endregion 

    #region units                            
        self.positionunit = positionunit
        self.velocityunit = velocityunit
        self.diameterunit = diameterunit
        self.temperatureunit = temperatureunit
        self.mfrunit = mfrunit
        self.massunit = massunit
        self.frequencyunit = frequencyunit
        self.timeunit = timeunit
        self.distributionunit = "%"
    #endregion

    #region plotter-standards
        self.language = "CZ"

        self.val_dpi = 400
        self.val_bin_single = 20
        self.val_bin_multi  = 5
        self.range_count = 10

        self.xmin_histo = 0.0
        self.xmax_histo = 1000.0
        self.xlim_histo_bool = False

        self.xmin_multihisto = 0
        self.xmax_multihisto = 1000
        self.histo_xlim_bool = False


        self.scatter_ylim_bool = False
        self.scatter_xlim_bool = False       
        self.xmin_scatter = 0.0
        self.xmax_scatter = 1000.0
        self.ymax_scatter = -1000.0
        self.ymin_scatter = 1000.0
    #endregion

    #region booleans
        self.bool_diameter = True
        self.bool_diameter_ascending = False
        
        self.bool_temperature = True
        self.bool_temperature_ascending = False

        self.bool_mfr = True
        self.bool_mfr_ascending = False

        self.bool_mass = True
        self.bool_mass_ascending = False

        self.bool_frequency = True
        self.bool_frequency_ascending = False


        self.bool_time = True
        self.bool_time_ascending = False

        self.bool_distribution_unit = True
    #endregion        

    #region variable-lists
        for i in range (0, len(self.source)):
            self.position.append({"ID": self.source["name"][i].split(":")[1]        , "position": (self.source["x"][i], self.source["y"][i], self.source["z"][i])                                                                                   , "unit": "m"})
            self.velocity.append({"ID": self.source["name"][i].split(":")[1]        , "velocity": (self.source["u"][i], self.source["v"][i], self.source["w"][i], sqrt(self.source["u"][i]**2 + self.source["v"][i]**2 + self.source["w"][i]**2))   , "unit": "m/s"})
            self.diameter.append({"ID": self.source["name"][i].split(":")[1]        , "diameter": self.source["diameter"][i]                                                                                                                        , "unit": "m"})
            self.temperature.append({"ID": self.source["name"][i].split(":")[1]     , "temperature":  self.source["t"][i]                                                                                                                           , "unit": "K"})
            self.mfr.append({"ID": self.source["name"][i].split(":")[1]             , "mass-flow-rate": self.source["mass-flow"][i]                                                                                                                 , "unit": "kg/s"})
            self.mass.append({"ID": self.source["name"][i].split(":")[1]            , "mass":  self.source["mass"][i]                                                                                                                               , "unit": "kg"})
            self.frequency.append({"ID": self.source["name"][i].split(":")[1]       , "frequency":  self.source["frequency"][i]                                                                                                                     , "unit": "1/s"})
            self.time.append({"ID": self.source["name"][i].split(":")[1]            , "time":  self.source["time"][i]                                                                                                                               , "unit": "s"})

            self.x.append({"ID": self.source["name"][i].split(":")[1], "x": self.source["x"][i], "unit": "m"})
            self.y.append({"ID": self.source["name"][i].split(":")[1], "y": self.source["y"][i], "unit": "m"})
            self.z.append({"ID": self.source["name"][i].split(":")[1], "z": self.source["z"][i], "unit": "m"})
            self.vx.append({"ID": self.source["name"][i].split(":")[1], "vx": self.source["u"][i], "unit": "m/s"})
            self.vy.append({"ID": self.source["name"][i].split(":")[1], "vy": self.source["v"][i], "unit": "m/s"})
            self.vz.append({"ID": self.source["name"][i].split(":")[1], "vz": self.source["w"][i], "unit": "m/s"})
            self.vmag.append({"ID": self.source["name"][i].split(":")[1], "vmag": sqrt(self.source["u"][i]**2 + self.source["v"][i]**2 + self.source["w"][i]**2), "unit": "m/s"})

    #endregion

    #region variable-unit-keys
        self.key0pos = list(self.position[0])[0]
        self.key1pos = list(self.position[0])[1]
        self.key2pos = list(self.position[0])[2]

        self.key0posx = list(self.position[0])[0]
        self.key0posy = list(self.position[0])[0]
        self.key0posz = list(self.position[0])[0]

        self.key1posx = list(self.position[0])[1]
        self.key1posy = list(self.position[0])[1]
        self.key1posz = list(self.position[0])[1]

        self.key2posx = list(self.position[0])[2]
        self.key2posy = list(self.position[0])[2]
        self.key2posz = list(self.position[0])[2]


        self.key0velo = list(self.velocity[0])[0]
        self.key1velo = list(self.velocity[0])[1]
        self.key2velo = list(self.velocity[0])[2]      

        self.key0velox = list(self.velocity[0])[0]
        self.key0veloy = list(self.velocity[0])[0]
        self.key0veloz =  list(self.velocity[0])[0]
        self.key0velomag =  list(self.velocity[0])[0]

        self.key1velox = list(self.velocity[0])[1]
        self.key1veloy = list(self.velocity[0])[1]
        self.key1veloz =  list(self.velocity[0])[1]
        self.key1velomag =  list(self.velocity[0])[1]

        self.key2velox = list(self.velocity[0])[2]
        self.key2veloy = list(self.velocity[0])[2]
        self.key2veloz =  list(self.velocity[0])[2]
        self.key2velomag =  list(self.velocity[0])[2]  

        self.key0dia = list(self.diameter[0])[0]
        self.key1dia = list(self.diameter[0])[1]
        self.key2dia = list(self.diameter[0])[2]

        self.key0temp = list(self.temperature[0])[0]
        self.key1temp = list(self.temperature[0])[1]
        self.key2temp = list(self.temperature[0])[2]

        self.key0mfr = list(self.mfr[0])[0]
        self.key1mfr = list(self.mfr[0])[1]
        self.key2mfr = list(self.mfr[0])[2]

        self.key0mass = list(self.mass[0])[0]
        self.key1mass = list(self.mass[0])[1]
        self.key2mass = list(self.mass[0])[2]

        self.key0freq = list(self.frequency[0])[0]
        self.key1freq = list(self.frequency[0])[1]
        self.key2freq = list(self.frequency[0])[2]

        self.key0time = list(self.time[0])[0]
        self.key1time = list(self.time[0])[1]
        self.key2time = list(self.time[0])[2]    
    #endregion
        print("Tabulka s názvem {} vytvořena".format(self.name))
    #region SORTING
        #region   DIAMETER 
    def sort_diameter(self):
        if self.bool_diameter == True:
            if self.bool_diameter_ascending == True:
                self.diameter = sorted(self.diameter, key = lambda k:k["diameter"], reverse = False)
            else:
                self.diameter = sorted(self.diameter, key = lambda k:k["diameter"], reverse = True)
        else:
            if self.bool_diameter_ascending == True:
                self.diameter = sorted(self.diameter, key = lambda k:k["ID"], reverse = False)
            else:
                self.diameter = sorted(self.diameter, key = lambda k:k["ID"], reverse = True)
                
    def diameter_ascending(self):
        if self.bool_diameter_ascending == True:
            self.bool_diameter_ascending = False
        else:
            self.bool_diameter_ascending = True
        #endregion
            
        #region   TEMPERATURE 
    def sort_temperature(self):
        if self.bool_temperature == True:
            if self.bool_temperature_ascending == True:
                self.temperature = sorted(self.temperature, key = lambda k:k["temperature"], reverse = False)
            else:
                self.temperature = sorted(self.temperature, key = lambda k:k["temperature"], reverse = True)
        else:
            if self.bool_temperature_ascending == True:
                self.temperature = sorted(self.temperature, key = lambda k:k["ID"], reverse = False)
            else:
                self.temperature = sorted(self.temperature, key = lambda k:k["ID"], reverse = True)
                
    def temperature_ascending(self):
        if self.bool_temperature_ascending == True:
            self.bool_temperature_ascending = False
        else:
            self.bool_temperature_ascending = True
        #endregion

        #region   MFR            
    def sort_mfr(self):
        if self.bool_mfr == True:
            if self.bool_mfr_ascending == True:
                self.mfr = sorted(self.mfrtable, key = lambda k:k["mass-flow-rate"], reverse = False)
            else:
                self.mfr = sorted(self.mfrtable, key = lambda k:k["mass-flow-rate"], reverse = True)
        else:
            if self.bool_mfr_ascending == True:
                self.mfr = sorted(self.mfrtable, key = lambda k:k["mass-flow-rate"], reverse = False)
            else:
                self.mfr = sorted(self.mfrtable, key = lambda k:k["mass-flow-rate"], reverse = True)
                
    def mfr_ascending(self):
        if self.bool_mfr_ascending == True:
            self.bool_mfr_ascending = False
        else:
            self.bool_mfr_ascending = True
        #endregion

        #region   MASS 
    def sort_mass(self):
        if self.bool_mass == True:
            if self.bool_mass_ascending == True:
                self.mass = sorted(self.mass, key = lambda k:k["mfr"], reverse = False)
            else:
                self.mass = sorted(self.mass, key = lambda k:k["mfr"], reverse = True)
        else:
            if self.bool_mass_ascending == True:
                self.mass = sorted(self.mass, key = lambda k:k["ID"], reverse = False)
            else:
                self.mass = sorted(self.mass, key = lambda k:k["ID"], reverse = True)
                
    def mass_ascending(self):
        if self.bool_mass_ascending == True:
            self.bool_mass_ascending = False
        else:
            self.bool_mass_ascending = True
        #endregion

        #region   FREQUENCY 
    def sort_frequency(self):
        if self.bool_frequency == True:
            if self.bool_frequency_ascending == True:
                self.frequency = sorted(self.frequency, key = lambda k:k["mfr"], reverse = False)
            else:
                self.frequency = sorted(self.frequency, key = lambda k:k["mfr"], reverse = True)
        else:
            if self.bool_frequency_ascending == True:
                self.frequency = sorted(self.frequency, key = lambda k:k["ID"], reverse = False)
            else:
                self.frequency = sorted(self.frequency, key = lambda k:k["ID"], reverse = True)
                
    def mfr_ascending(self):
        if self.bool_frequency_ascending == True:
            self.bool_frequency_ascending = False
        else:
            self.bool_frequency_ascending = True
        #endregion
       
        #region   TIME 
    def sort_time(self):
        if self.bool_time == True:
            if self.bool_time_ascending == True:
                self.time = sorted(self.time, key = lambda k:k["time"], reverse = False)
            else:
                self.time = sorted(self.time, key = lambda k:k["time"], reverse = True)
        else:
            if self.bool_time_ascending == True:
                self.time = sorted(self.time, key = lambda k:k["ID"], reverse = False)
            else:
                self.time = sorted(self.time, key = lambda k:k["ID"], reverse = True)
                
    def time_ascending(self):
        if self.bool_time_ascending == True:
            self.bool_time_ascending = False
        else:
            self.bool_time_ascending = True
        #endregion
    #endregion
    
    #region UNIT OPERATIONS
    def update_units(self):
        self.positionunit = self.position[0].get(self.key2pos)
        self.velocityunit = self.velocity[0].get(self.key2velo)
        self.diameterunit = self.diameter[0].get(self.key2dia)
        self.temperatureunit =  self.temperature[0].get(self.key2temp)
        self.mfrunit =  self.mfr[0].get(self.key2mfr)
        self.massunit = self.mass[0].get(self.key2mass)
        self.frequencyunit =  self.frequency[0].get(self.key2freq)
        self.timeunit = self.time[0].get(self.key2time)
        print("Update units: ")
        print(self.position[0],self.velocity[0],self.diameter[0],self.temperature[0],self.mfr[0],self.mass[0],self.frequency[0],self.time[0])

    def unitswitch_pos(self,targetunit):
        self.tmp_storage_pos = []
        self.tmp_storage_posx = []
        self.tmp_storage_posy = []
        self.tmp_storage_posz = []

        self.key0pos = list(self.position[0])[0]
        self.key1pos = list(self.position[0])[1]
        self.key2pos = list(self.position[0])[2]
        
        self.key0posx = list(self.x[0])[0]
        self.key0posy = list(self.y[0])[0]
        self.key0posz = list(self.z[0])[0]

        self.key1posx = list(self.x[0])[1]
        self.key1posy = list(self.y[0])[1]
        self.key1posz = list(self.z[0])[1]

        self.key2posx = list(self.x[0])[2]
        self.key2posy = list(self.y[0])[2]
        self.key2posz = list(self.z[0])[2]

        

        self.old_pos_unit = self.position[0].get(self.key2pos)
        print(2*"\n")
        print("***************UNITSWITCH: POSITION - BEGIN***************")
        print(self.old_pos_unit)

        #region COMPOUND
        if targetunit == "mm":
            if self.old_pos_unit == "mm":
                nasobic = 1
                for part in self.position:
                    self.va0pos = part.get(self.key0pos)
                    self.va1pos = part.get(self.key1pos)
                    self.va2pos = part.get(self.key2pos)
                    line = {self.key0pos:    self.va0pos, self.key1pos: (self.va1pos[0]*nasobic, self.va1pos[1]*nasobic, self.va1pos[2]*nasobic), self.key2pos:  targetunit}
                    self.tmp_storage_pos.append(line)

            elif self.old_pos_unit == "cm":
                nasobic = 10
                for part in self.position:
                    self.va0pos = part.get(self.key0pos)
                    self.va1pos = part.get(self.key1pos)
                    self.va2pos = part.get(self.key2pos)
                    line = {self.key0pos:    self.va0pos, self.key1pos: (self.va1pos[0]*nasobic, self.va1pos[1]*nasobic, self.va1pos[2]*nasobic), self.key2pos:  targetunit}
                    self.tmp_storage_pos.append(line)
            
            elif self.old_pos_unit == "dm":
                nasobic = 100
                for part in self.position:
                    self.va0pos = part.get(self.key0pos)
                    self.va1pos = part.get(self.key1pos)
                    self.va2pos = part.get(self.key2pos)
                    line = {self.key0pos:    self.va0pos, self.key1pos: (self.va1pos[0]*nasobic, self.va1pos[1]*nasobic, self.va1pos[2]*nasobic), self.key2pos:  targetunit}
                    self.tmp_storage_pos.append(line)

            elif self.old_pos_unit == "m":
                nasobic = 1000
                for part in self.position:
                    self.va0pos = part.get(self.key0pos)
                    self.va1pos = part.get(self.key1pos)
                    self.va2pos = part.get(self.key2pos)
                    line = {self.key0pos:    self.va0pos, self.key1pos: (self.va1pos[0]*nasobic, self.va1pos[1]*nasobic, self.va1pos[2]*nasobic), self.key2pos:  targetunit}
                    self.tmp_storage_pos.append(line)
        
        elif targetunit == "cm":
            if self.old_pos_unit == "mm":
                nasobic = 0.1
                for part in self.position:
                    self.va0pos = part.get(self.key0pos)
                    self.va1pos = part.get(self.key1pos)
                    self.va2pos = part.get(self.key2pos)
                    line = {self.key0pos:    self.va0pos, self.key1pos: (self.va1pos[0]*nasobic, self.va1pos[1]*nasobic, self.va1pos[2]*nasobic), self.key2pos:  targetunit}
                    self.tmp_storage_pos.append(line)

            elif self.old_pos_unit == "cm":
                nasobic = 1
                for part in self.position:
                    self.va0pos = part.get(self.key0pos)
                    self.va1pos = part.get(self.key1pos)
                    self.va2pos = part.get(self.key2pos)
                    line = {self.key0pos:    self.va0pos, self.key1pos: (self.va1pos[0]*nasobic, self.va1pos[1]*nasobic, self.va1pos[2]*nasobic), self.key2pos:  targetunit}
                    self.tmp_storage_pos.append(line)
            
            elif self.old_pos_unit == "dm":
                nasobic = 10
                for part in self.position:
                    self.va0pos = part.get(self.key0pos)
                    self.va1pos = part.get(self.key1pos)
                    self.va2pos = part.get(self.key2pos)
                    line = {self.key0pos:    self.va0pos, self.key1pos: (self.va1pos[0]*nasobic, self.va1pos[1]*nasobic, self.va1pos[2]*nasobic), self.key2pos:  targetunit}
                    self.tmp_storage_pos.append(line)

            elif self.old_pos_unit == "m":
                nasobic = 100
                for part in self.position:
                    self.va0pos = part.get(self.key0pos)
                    self.va1pos = part.get(self.key1pos)
                    self.va2pos = part.get(self.key2pos)
                    line = {self.key0pos:    self.va0pos, self.key1pos: (self.va1pos[0]*nasobic, self.va1pos[1]*nasobic, self.va1pos[2]*nasobic), self.key2pos:  targetunit}
                    self.tmp_storage_pos.append(line)

        elif targetunit == "dm":
            if self.old_pos_unit == "mm":
                nasobic = 0.01
                for part in self.position:
                    self.va0pos = part.get(self.key0pos)
                    self.va1pos = part.get(self.key1pos)
                    self.va2pos = part.get(self.key2pos)
                    line = {self.key0pos:    self.va0pos, self.key1pos: (self.va1pos[0]*nasobic, self.va1pos[1]*nasobic, self.va1pos[2]*nasobic), self.key2pos:  targetunit}
                    self.tmp_storage_pos.append(line)

            elif self.old_pos_unit == "cm":
                nasobic = 0.1
                for part in self.position:
                    self.va0pos = part.get(self.key0pos)
                    self.va1pos = part.get(self.key1pos)
                    self.va2pos = part.get(self.key2pos)
                    line = {self.key0pos:    self.va0pos, self.key1pos: (self.va1pos[0]*nasobic, self.va1pos[1]*nasobic, self.va1pos[2]*nasobic), self.key2pos:  targetunit}
                    self.tmp_storage_pos.append(line)
            
            elif self.old_pos_unit == "dm":
                nasobic = 1
                for part in self.position:
                    self.va0pos = part.get(self.key0pos)
                    self.va1pos = part.get(self.key1pos)
                    self.va2pos = part.get(self.key2pos)
                    line = {self.key0pos:    self.va0pos, self.key1pos: (self.va1pos[0]*nasobic, self.va1pos[1]*nasobic, self.va1pos[2]*nasobic), self.key2pos:  targetunit}
                    self.tmp_storage_pos.append(line)

            elif self.old_pos_unit == "m":
                nasobic = 10
                for part in self.position:
                    self.va0pos = part.get(self.key0pos)
                    self.va1pos = part.get(self.key1pos)
                    self.va2pos = part.get(self.key2pos)
                    line = {self.key0pos:    self.va0pos, self.key1pos: (self.va1pos[0]*nasobic, self.va1pos[1]*nasobic, self.va1pos[2]*nasobic), self.key2pos:  targetunit}
                    self.tmp_storage_pos.append(line)

        elif targetunit == "m":
            if self.old_pos_unit == "mm":
                nasobic = 0.001
                for part in self.position:
                    self.va0pos = part.get(self.key0pos)
                    self.va1pos = part.get(self.key1pos)
                    self.va2pos = part.get(self.key2pos)
                    line = {self.key0pos:    self.va0pos, self.key1pos: (self.va1pos[0]*nasobic, self.va1pos[1]*nasobic, self.va1pos[2]*nasobic), self.key2pos:  targetunit}
                    self.tmp_storage_pos.append(line)

            elif self.old_pos_unit == "cm":
                nasobic = 0.01
                for part in self.position:
                    self.va0pos = part.get(self.key0pos)
                    self.va1pos = part.get(self.key1pos)
                    self.va2pos = part.get(self.key2pos)
                    line = {self.key0pos:    self.va0pos, self.key1pos: (self.va1pos[0]*nasobic, self.va1pos[1]*nasobic, self.va1pos[2]*nasobic), self.key2pos:  targetunit}
                    self.tmp_storage_pos.append(line)
            
            elif self.old_pos_unit == "dm":
                nasobic = 0.1
                for part in self.position:
                    self.va0pos = part.get(self.key0pos)
                    self.va1pos = part.get(self.key1pos)
                    self.va2pos = part.get(self.key2pos)
                    line = {self.key0pos:    self.va0pos, self.key1pos: (self.va1pos[0]*nasobic, self.va1pos[1]*nasobic, self.va1pos[2]*nasobic), self.key2pos:  targetunit}
                    self.tmp_storage_pos.append(line)

            elif self.old_pos_unit == "m":
                nasobic = 1
                for part in self.position:
                    self.va0pos = part.get(self.key0pos)
                    self.va1pos = part.get(self.key1pos)
                    self.va2pos = part.get(self.key2pos)
                    line = {self.key0pos:    self.va0pos, self.key1pos: (self.va1pos[0]*nasobic, self.va1pos[1]*nasobic, self.va1pos[2]*nasobic), self.key2pos:  targetunit}
                    self.tmp_storage_pos.append(line)
        #endregion

        #region X-pos
        if targetunit == "mm":
            if self.old_pos_unit == "mm":
                nasobic = 1
                for part in self.x:
                    self.va0posx = part.get("ID")
                    self.va1posx = part.get("x")
                    self.va2posx = part.get("unit")
                    linex = {self.key0posx:    self.va0posx, self.key1posx: self.va1posx*nasobic, self.key2posx:  targetunit}
                    self.tmp_storage_posx.append(linex)

            elif self.old_pos_unit == "cm":
                nasobic = 10
                for part in self.x:
                    self.va0posx = part.get("ID")
                    self.va1posx = part.get("x")
                    self.va2posx = part.get("unit")
                    linex = {self.key0posx:    self.va0posx, self.key1posx: self.va1posx*nasobic, self.key2posx:  targetunit}
                    self.tmp_storage_posx.append(linex)
            
            elif self.old_pos_unit == "dm":
                nasobic = 100
                for part in self.x:
                    self.va0posx = part.get("ID")
                    self.va1posx = part.get("x")
                    self.va2posx = part.get("unit")
                    linex = {self.key0posx:    self.va0posx, self.key1posx: self.va1posx*nasobic, self.key2posx:  targetunit}
                    self.tmp_storage_posx.append(linex)

            elif self.old_pos_unit == "m":
                nasobic = 1000
                for part in self.x:
                    self.va0posx = part.get("ID")
                    self.va1posx = part.get("x")
                    self.va2posx = part.get("unit")
                    linex = {self.key0posx:    self.va0posx, self.key1posx: self.va1posx*nasobic, self.key2posx:  targetunit}
                    self.tmp_storage_posx.append(linex)
        
        elif targetunit == "cm":
            if self.old_pos_unit == "mm":
                nasobic = 0.1
                for part in self.x:
                    self.va0posx = part.get("ID")
                    self.va1posx = part.get("x")
                    self.va2posx = part.get("unit")
                    linex = {self.key0posx:    self.va0posx, self.key1posx: self.va1posx*nasobic, self.key2posx:  targetunit}
                    self.tmp_storage_posx.append(linex)

            elif self.old_pos_unit == "cm":
                nasobic = 1
                for part in self.x:
                    self.va0posx = part.get("ID")
                    self.va1posx = part.get("x")
                    self.va2posx = part.get("unit")
                    linex = {self.key0posx:    self.va0posx, self.key1posx: self.va1posx*nasobic, self.key2posx:  targetunit}
                    self.tmp_storage_posx.append(linex)
            
            elif self.old_pos_unit == "dm":
                nasobic = 10
                for part in self.x:
                    self.va0posx = part.get("ID")
                    self.va1posx = part.get("x")
                    self.va2posx = part.get("unit")
                    linex = {self.key0posx:    self.va0posx, self.key1posx: self.va1posx*nasobic, self.key2posx:  targetunit}
                    self.tmp_storage_posx.append(linex)

            elif self.old_pos_unit == "m":
                nasobic = 100
                for part in self.x:
                    self.va0posx = part.get("ID")
                    self.va1posx = part.get("x")
                    self.va2posx = part.get("unit")
                    linex = {self.key0posx:    self.va0posx, self.key1posx: self.va1posx*nasobic, self.key2posx:  targetunit}
                    self.tmp_storage_posx.append(linex)

        elif targetunit == "dm":
            if self.old_pos_unit == "mm":
                nasobic = 0.01
                for part in self.x:
                    self.va0posx = part.get("ID")
                    self.va1posx = part.get("x")
                    self.va2posx = part.get("unit")
                    linex = {self.key0posx:    self.va0posx, self.key1posx: self.va1posx*nasobic, self.key2posx:  targetunit}
                    self.tmp_storage_posx.append(linex)

            elif self.old_pos_unit == "cm":
                nasobic = 0.1
                for part in self.x:
                    self.va0posx = part.get("ID")
                    self.va1posx = part.get("x")
                    self.va2posx = part.get("unit")
                    linex = {self.key0posx:    self.va0posx, self.key1posx: self.va1posx*nasobic, self.key2posx:  targetunit}
                    self.tmp_storage_posx.append(linex)
            
            elif self.old_pos_unit == "dm":
                nasobic = 1
                for part in self.x:
                    self.va0posx = part.get("ID")
                    self.va1posx = part.get("x")
                    self.va2posx = part.get("unit")
                    linex = {self.key0posx:    self.va0posx, self.key1posx: self.va1posx*nasobic, self.key2posx:  targetunit}
                    self.tmp_storage_posx.append(linex)

            elif self.old_pos_unit == "m":
                nasobic = 10
                for part in self.x:
                    self.va0posx = part.get("ID")
                    self.va1posx = part.get("x")
                    self.va2posx = part.get("unit")
                    linex = {self.key0posx:    self.va0posx, self.key1posx: self.va1posx*nasobic, self.key2posx:  targetunit}
                    self.tmp_storage_posx.append(linex)

        elif targetunit == "m":
            if self.old_pos_unit == "mm":
                nasobic = 0.001
                for part in self.x:
                    self.va0posx = part.get("ID")
                    self.va1posx = part.get("x")
                    self.va2posx = part.get("unit")
                    linex = {self.key0posx:    self.va0posx, self.key1posx: self.va1posx*nasobic, self.key2posx:  targetunit}
                    self.tmp_storage_posx.append(linex)

            elif self.old_pos_unit == "cm":
                nasobic = 0.01
                for part in self.x:
                    self.va0posx = part.get("ID")
                    self.va1posx = part.get("x")
                    self.va2posx = part.get("unit")
                    linex = {self.key0posx:    self.va0posx, self.key1posx: self.va1posx*nasobic, self.key2posx:  targetunit}
                    self.tmp_storage_posx.append(linex)
            
            elif self.old_pos_unit == "dm":
                nasobic = 0.1
                for part in self.x:
                    self.va0posx = part.get("ID")
                    self.va1posx = part.get("x")
                    self.va2posx = part.get("unit")
                    linex = {self.key0posx:    self.va0posx, self.key1posx: self.va1posx*nasobic, self.key2posx:  targetunit}
                    self.tmp_storage_posx.append(linex)

            elif self.old_pos_unit == "m":
                nasobic = 1
                for part in self.x:
                    self.va0posx = part.get("ID")
                    self.va1posx = part.get("x")
                    self.va2posx = part.get("unit")
                    linex = {self.key0posx:    self.va0posx, self.key1posx: self.va1posx*nasobic, self.key2posx:  targetunit}
                    self.tmp_storage_posx.append(linex)
        #endregion

        #region Y-pos
        if targetunit == "mm":
            if self.old_pos_unit == "mm":
                nasobic = 1
                for part in self.y:
                    self.va0posy = part.get("ID")
                    self.va1posy = part.get("y")
                    self.va2posy = part.get("unit")
                    liney = {self.key0posy:    self.va0posy, self.key1posy: self.va1posy*nasobic, self.key2posy:  targetunit}
                    self.tmp_storage_posy.append(liney)

            elif self.old_pos_unit == "cm":
                nasobic = 10
                for part in self.y:
                    self.va0posy = part.get("ID")
                    self.va1posy = part.get("y")
                    self.va2posy = part.get("unit")
                    liney = {self.key0posy:    self.va0posy, self.key1posy: self.va1posy*nasobic, self.key2posy:  targetunit}
                    self.tmp_storage_posy.append(liney)
            
            elif self.old_pos_unit == "dm":
                nasobic = 100
                for part in self.y:
                    self.va0posy = part.get("ID")
                    self.va1posy = part.get("y")
                    self.va2posy = part.get("unit")
                    liney = {self.key0posy:    self.va0posy, self.key1posy: self.va1posy*nasobic, self.key2posy:  targetunit}
                    self.tmp_storage_posy.append(liney)

            elif self.old_pos_unit == "m":
                nasobic = 1000
                for part in self.y:
                    self.va0posy = part.get("ID")
                    self.va1posy = part.get("y")
                    self.va2posy = part.get("unit")
                    liney = {self.key0posy:    self.va0posy, self.key1posy: self.va1posy*nasobic, self.key2posy:  targetunit}
                    self.tmp_storage_posy.append(liney)
        
        elif targetunit == "cm":
            if self.old_pos_unit == "mm":
                nasobic = 0.1
                for part in self.y:
                    self.va0posy = part.get("ID")
                    self.va1posy = part.get("y")
                    self.va2posy = part.get("unit")
                    liney = {self.key0posy:    self.va0posy, self.key1posy: self.va1posy*nasobic, self.key2posy:  targetunit}
                    self.tmp_storage_posy.append(liney)

            elif self.old_pos_unit == "cm":
                nasobic = 1
                for part in self.y:
                    self.va0posy = part.get("ID")
                    self.va1posy = part.get("y")
                    self.va2posy = part.get("unit")
                    liney = {self.key0posy:    self.va0posy, self.key1posy: self.va1posy*nasobic, self.key2posy:  targetunit}
                    self.tmp_storage_posy.append(liney)
            
            elif self.old_pos_unit == "dm":
                nasobic = 10
                for part in self.y:
                    self.va0posy = part.get("ID")
                    self.va1posy = part.get("y")
                    self.va2posy = part.get("unit")
                    liney = {self.key0posy:    self.va0posy, self.key1posy: self.va1posy*nasobic, self.key2posy:  targetunit}
                    self.tmp_storage_posy.append(liney)

            elif self.old_pos_unit == "m":
                nasobic = 100
                for part in self.y:
                    self.va0posy = part.get("ID")
                    self.va1posy = part.get("y")
                    self.va2posy = part.get("unit")
                    liney = {self.key0posy:    self.va0posy, self.key1posy: self.va1posy*nasobic, self.key2posy:  targetunit}
                    self.tmp_storage_posy.append(liney)

        elif targetunit == "dm":
            if self.old_pos_unit == "mm":
                nasobic = 0.01
                for part in self.y:
                    self.va0posy = part.get("ID")
                    self.va1posy = part.get("y")
                    self.va2posy = part.get("unit")
                    liney = {self.key0posy:    self.va0posy, self.key1posy: self.va1posy*nasobic, self.key2posy:  targetunit}
                    self.tmp_storage_posy.append(liney)

            elif self.old_pos_unit == "cm":
                nasobic = 0.1
                for part in self.y:
                    self.va0posy = part.get("ID")
                    self.va1posy = part.get("y")
                    self.va2posy = part.get("unit")
                    liney = {self.key0posy:    self.va0posy, self.key1posy: self.va1posy*nasobic, self.key2posy:  targetunit}
                    self.tmp_storage_posy.append(liney)
            
            elif self.old_pos_unit == "dm":
                nasobic = 1
                for part in self.y:
                    self.va0posy = part.get("ID")
                    self.va1posy = part.get("y")
                    self.va2posy = part.get("unit")
                    liney = {self.key0posy:    self.va0posy, self.key1posy: self.va1posy*nasobic, self.key2posy:  targetunit}
                    self.tmp_storage_posy.append(liney)

            elif self.old_pos_unit == "m":
                nasobic = 10
                for part in self.y:
                    self.va0posy = part.get("ID")
                    self.va1posy = part.get("y")
                    self.va2posy = part.get("unit")
                    liney = {self.key0posy:    self.va0posy, self.key1posy: self.va1posy*nasobic, self.key2posy:  targetunit}
                    self.tmp_storage_posy.append(liney)

        elif targetunit == "m":
            if self.old_pos_unit == "mm":
                nasobic = 0.001
                for part in self.y:
                    self.va0posy = part.get("ID")
                    self.va1posy = part.get("y")
                    self.va2posy = part.get("unit")
                    liney = {self.key0posy:    self.va0posy, self.key1posy: self.va1posy*nasobic, self.key2posy:  targetunit}
                    self.tmp_storage_posy.append(liney)

            elif self.old_pos_unit == "cm":
                nasobic = 0.01
                for part in self.y:
                    self.va0posy = part.get("ID")
                    self.va1posy = part.get("y")
                    self.va2posy = part.get("unit")
                    liney = {self.key0posy:    self.va0posy, self.key1posy: self.va1posy*nasobic, self.key2posy:  targetunit}
                    self.tmp_storage_posy.append(liney)
            
            elif self.old_pos_unit == "dm":
                nasobic = 0.1
                for part in self.y:
                    self.va0posy = part.get("ID")
                    self.va1posy = part.get("y")
                    self.va2posy = part.get("unit")
                    liney = {self.key0posy:    self.va0posy, self.key1posy: self.va1posy*nasobic, self.key2posy:  targetunit}
                    self.tmp_storage_posy.append(liney)

            elif self.old_pos_unit == "m":
                nasobic = 1
                for part in self.y:
                    self.va0posy = part.get("ID")
                    self.va1posy = part.get("y")
                    self.va2posy = part.get("unit")
                    liney = {self.key0posy:    self.va0posy, self.key1posy: self.va1posy*nasobic, self.key2posy:  targetunit}
                    self.tmp_storage_posy.append(liney)
        #endregion     
                
        #region Z-pos
        if targetunit == "mm":
            if self.old_pos_unit == "mm":
                nasobic = 1
                for part in self.z:
                    self.va0posz = part.get("ID")
                    self.va1posz = part.get("z")
                    self.va2posz = part.get("unit")
                    linez = {self.key0posz:    self.va0posz, self.key1posz: self.va1posz*nasobic, self.key2posz:  targetunit}
                    self.tmp_storage_posz.append(linez)

            elif self.old_pos_unit == "cm":
                nasobic = 10
                for part in self.z:
                    self.va0posz = part.get("ID")
                    self.va1posz = part.get("z")
                    self.va2posz = part.get("unit")
                    linez = {self.key0posz:    self.va0posz, self.key1posz: self.va1posz*nasobic, self.key2posz:  targetunit}
                    self.tmp_storage_posz.append(linez)
            
            elif self.old_pos_unit == "dm":
                nasobic = 100
                for part in self.z:
                    self.va0posz = part.get("ID")
                    self.va1posz = part.get("z")
                    self.va2posz = part.get("unit")
                    linez = {self.key0posz:    self.va0posz, self.key1posz: self.va1posz*nasobic, self.key2posz:  targetunit}
                    self.tmp_storage_posz.append(linez)

            elif self.old_pos_unit == "m":
                nasobic = 1000
                for part in self.z:
                    self.va0posz = part.get("ID")
                    self.va1posz = part.get("z")
                    self.va2posz = part.get("unit")
                    linez = {self.key0posz:    self.va0posz, self.key1posz: self.va1posz*nasobic, self.key2posz:  targetunit}
                    self.tmp_storage_posz.append(linez)
        
        elif targetunit == "cm":
            if self.old_pos_unit == "mm":
                nasobic = 0.1
                for part in self.z:
                    self.va0posz = part.get("ID")
                    self.va1posz = part.get("z")
                    self.va2posz = part.get("unit")
                    linez = {self.key0posz:    self.va0posz, self.key1posz: self.va1posz*nasobic, self.key2posz:  targetunit}
                    self.tmp_storage_posz.append(linez)

            elif self.old_pos_unit == "cm":
                nasobic = 1
                for part in self.z:
                    self.va0posz = part.get("ID")
                    self.va1posz = part.get("z")
                    self.va2posz = part.get("unit")
                    linez = {self.key0posz:    self.va0posz, self.key1posz: self.va1posz*nasobic, self.key2posz:  targetunit}
                    self.tmp_storage_posz.append(linez)
            
            elif self.old_pos_unit == "dm":
                nasobic = 10
                for part in self.z:
                    self.va0posz = part.get("ID")
                    self.va1posz = part.get("z")
                    self.va2posz = part.get("unit")
                    linez = {self.key0posz:    self.va0posz, self.key1posz: self.va1posz*nasobic, self.key2posz:  targetunit}
                    self.tmp_storage_posz.append(linez)

            elif self.old_pos_unit == "m":
                nasobic = 100
                for part in self.z:
                    self.va0posz = part.get("ID")
                    self.va1posz = part.get("z")
                    self.va2posz = part.get("unit")
                    linez = {self.key0posz:    self.va0posz, self.key1posz: self.va1posz*nasobic, self.key2posz:  targetunit}
                    self.tmp_storage_posz.append(linez)

        elif targetunit == "dm":
            if self.old_pos_unit == "mm":
                nasobic = 0.01
                for part in self.z:
                    self.va0posz = part.get("ID")
                    self.va1posz = part.get("z")
                    self.va2posz = part.get("unit")
                    linez = {self.key0posz:    self.va0posz, self.key1posz: self.va1posz*nasobic, self.key2posz:  targetunit}
                    self.tmp_storage_posz.append(linez)

            elif self.old_pos_unit == "cm":
                nasobic = 0.1
                for part in self.z:
                    self.va0posz = part.get("ID")
                    self.va1posz = part.get("z")
                    self.va2posz = part.get("unit")
                    linez = {self.key0posz:    self.va0posz, self.key1posz: self.va1posz*nasobic, self.key2posz:  targetunit}
                    self.tmp_storage_posz.append(linez)
            
            elif self.old_pos_unit == "dm":
                nasobic = 1
                for part in self.z:
                    self.va0posz = part.get("ID")
                    self.va1posz = part.get("z")
                    self.va2posz = part.get("unit")
                    linez = {self.key0posz:    self.va0posz, self.key1posz: self.va1posz*nasobic, self.key2posz:  targetunit}
                    self.tmp_storage_posz.append(linez)

            elif self.old_pos_unit == "m":
                nasobic = 10
                for part in self.z:
                    self.va0posz = part.get("ID")
                    self.va1posz = part.get("z")
                    self.va2posz = part.get("unit")
                    linez = {self.key0posz:    self.va0posz, self.key1posz: self.va1posz*nasobic, self.key2posz:  targetunit}
                    self.tmp_storage_posz.append(linez)

        elif targetunit == "m":
            if self.old_pos_unit == "mm":
                nasobic = 0.001
                for part in self.z:
                    self.va0posz = part.get("ID")
                    self.va1posz = part.get("z")
                    self.va2posz = part.get("unit")
                    linez = {self.key0posz:    self.va0posz, self.key1posz: self.va1posz*nasobic, self.key2posz:  targetunit}
                    self.tmp_storage_posz.append(linez)

            elif self.old_pos_unit == "cm":
                nasobic = 0.01
                for part in self.z:
                    self.va0posz = part.get("ID")
                    self.va1posz = part.get("z")
                    self.va2posz = part.get("unit")
                    linez = {self.key0posz:    self.va0posz, self.key1posz: self.va1posz*nasobic, self.key2posz:  targetunit}
                    self.tmp_storage_posz.append(linez)
            
            elif self.old_pos_unit == "dm":
                nasobic = 0.1
                for part in self.z:
                    self.va0posz = part.get("ID")
                    self.va1posz = part.get("z")
                    self.va2posz = part.get("unit")
                    linez = {self.key0posz:    self.va0posz, self.key1posz: self.va1posz*nasobic, self.key2posz:  targetunit}
                    self.tmp_storage_posz.append(linez)

            elif self.old_pos_unit == "m":
                nasobic = 1
                for part in self.z:
                    self.va0posz = part.get("ID")
                    self.va1posz = part.get("z")
                    self.va2posz = part.get("unit")
                    linez = {self.key0posz:    self.va0posz, self.key1posz: self.va1posz*nasobic, self.key2posz:  targetunit}
                    self.tmp_storage_posz.append(linez)
        #endregion
        
        self.position = self.tmp_storage_pos[:]
        self.positionunit = self.position[0].get(self.key2pos)
        self.tmp_storage_pos.clear()

        self.x = self.tmp_storage_posx[:]
        self.y = self.tmp_storage_posy[:]
        self.z = self.tmp_storage_posz[:]
        self.tmp_storage_posx.clear()
        self.tmp_storage_posy.clear()
        self.tmp_storage_posx.clear()
        print(self.positionunit)
        print("***************UNITSWITCH: POSITION - END***************")
        print(2*"\n")

    def unitswitch_velo(self,targetunit):
        self.tmp_storage_velocity = []
        self.tmp_storage_vx = []
        self.tmp_storage_vy = []
        self.tmp_storage_vz = []
        self.tmp_storage_vmag = []


        self.key0velo = list(self.velocity[0])[0]
        self.key1velo = list(self.velocity[0])[1]
        self.key2velo = list(self.velocity[0])[2]

        self.key0velox = list(self.vx[0])[0]
        self.key0veloy = list(self.vy[0])[0]
        self.key0veloz =  list(self.vz[0])[0]
        self.key0velomag =  list(self.vmag[0])[0]

        self.key1velox = list(self.vx[0])[1]
        self.key1veloy = list(self.vy[0])[1]
        self.key1veloz =  list(self.vz[0])[1]
        self.key1velomag =  list(self.vmag[0])[1]

        self.key2velox = list(self.vx[0])[2]
        self.key2veloy = list(self.vy[0])[2]
        self.key2veloz =  list(self.vz[0])[2]
        self.key2velomag =  list(self.vmag[0])[2]


        self.old_velo_unit = self.velocity[0].get(self.key2velo)
        print(2*"\n")
        print("***************UNITSWITCH: VELOCITY - BEGIN***************")
        print(self.old_velo_unit)
            
        #region COMPOUND
        if targetunit == "mm/s":
            if self.old_velo_unit == "mm/s":
                nasobic = 1
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)

            elif self.old_velo_unit == "cm/s":
                nasobic = 10
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = 100
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)

            elif self.old_velo_unit == "m/s":
                nasobic = 1000
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)

            elif self.old_velo_unit == "km/h":
                nasobic = 3600
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)
        
        elif targetunit == "cm/s":
            if self.old_velo_unit == "mm/s":
                nasobic = .1
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)

            elif self.old_velo_unit == "cm/s":
                nasobic = 1.0
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = 10.0
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)

            elif self.old_velo_unit == "m/s":
                nasobic = 100.0
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)

            elif self.old_velo_unit == "km/h":
                nasobic = 360.0
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)

        elif targetunit == "dm/s":
            if self.old_velo_unit == "mm/s":
                nasobic = .01
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)

            elif self.old_velo_unit == "cm/s":
                nasobic = .10
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = 1.00
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)

            elif self.old_velo_unit == "m/s":
                nasobic = 10.00
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)

            elif self.old_velo_unit == "km/h":
                nasobic = 36.00
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)

        elif targetunit == "m/s":
            if self.old_velo_unit == "mm/s":
                nasobic = .001
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)

            elif self.old_velo_unit == "cm/s":
                nasobic = .010
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = .100
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)

            elif self.old_velo_unit == "m/s":
                nasobic = 1.000
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)

            elif self.old_velo_unit == "km/h":
                nasobic = 1/3.600
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)
    
        elif targetunit == "km/h":
            if self.old_velo_unit == "mm/s":
                nasobic = 3.6e-3
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)

            elif self.old_velo_unit == "cm/s":
                nasobic = 3.6e-2
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = 3.6e-1
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)

            elif self.old_velo_unit == "m/s":
                nasobic = 3.6e0
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)

            elif self.old_velo_unit == "km/h":
                nasobic = 1
                for part in self.velocity:
                    self.va0velo = part.get(self.key0velo)
                    self.va1velo = part.get(self.key1velo)
                    self.va2velo = part.get(self.key2velo)
                    line = {self.key0velo:    self.va0velo, self.key1velo: (self.va1velo[0]*nasobic, self.va1velo[1]*nasobic, self.va1velo[2]*nasobic, self.va1velo[3]*nasobic), self.key2velo:  targetunit}
                    self.tmp_storage_velocity.append(line)
        #endregion

        #region x-velo
        if targetunit == "mm/s":
            if self.old_velo_unit == "mm/s":
                nasobic = 1
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)

            elif self.old_velo_unit == "cm/s":
                nasobic = 10
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = 100
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)

            elif self.old_velo_unit == "m/s":
                nasobic = 1000
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)
                    print(linevx)

            elif self.old_velo_unit == "km/h":
                nasobic = 3600
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)
        
        elif targetunit == "cm/s":
            if self.old_velo_unit == "mm/s":
                nasobic = .1
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)

            elif self.old_velo_unit == "cm/s":
                nasobic = 1.0
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = 10.0
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)

            elif self.old_velo_unit == "m/s":
                nasobic = 100.0
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)

            elif self.old_velo_unit == "km/h":
                nasobic = 360.0
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)

        elif targetunit == "dm/s":
            if self.old_velo_unit == "mm/s":
                nasobic = .01
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)

            elif self.old_velo_unit == "cm/s":
                nasobic = .10
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = 1.00
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)

            elif self.old_velo_unit == "m/s":
                nasobic = 10.00
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)

            elif self.old_velo_unit == "km/h":
                nasobic = 36.00
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)

        elif targetunit == "m/s":
            if self.old_velo_unit == "mm/s":
                nasobic = .001
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)

            elif self.old_velo_unit == "cm/s":
                nasobic = .010
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = .100
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)

            elif self.old_velo_unit == "m/s":
                nasobic = 1.000
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)

            elif self.old_velo_unit == "km/h":
                nasobic = 1/3.600
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)
    
        elif targetunit == "km/h":
            if self.old_velo_unit == "mm/s":
                nasobic = 3.6e-3
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)

            elif self.old_velo_unit == "cm/s":
                nasobic = 3.6e-2
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = 3.6e-1
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)

            elif self.old_velo_unit == "m/s":
                nasobic = 3.6e0
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)

            elif self.old_velo_unit == "km/h":
                nasobic = 1
                for part in self.vx:
                    self.va0velox = part.get("ID")
                    self.va1velox = part.get("vx")
                    self.va2velox = part.get("unit")
                    linevx = {self.key0velox:    self.va0velox, self.key1velox: self.va1velox*nasobic, self.key2velox:  targetunit}
                    self.tmp_storage_vx.append(linevx)
        
        #endregion

        #region y-velo
        if targetunit == "mm/s":
            if self.old_velo_unit == "mm/s":
                nasobic = 1
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)

            elif self.old_velo_unit == "cm/s":
                nasobic = 10
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = 100
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)

            elif self.old_velo_unit == "m/s":
                nasobic = 1000
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)

            elif self.old_velo_unit == "km/h":
                nasobic = 3600
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)
        
        elif targetunit == "cm/s":
            if self.old_velo_unit == "mm/s":
                nasobic = .1
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)

            elif self.old_velo_unit == "cm/s":
                nasobic = 1.0
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = 10.0
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)

            elif self.old_velo_unit == "m/s":
                nasobic = 100.0
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)

            elif self.old_velo_unit == "km/h":
                nasobic = 360.0
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)

        elif targetunit == "dm/s":
            if self.old_velo_unit == "mm/s":
                nasobic = .01
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)

            elif self.old_velo_unit == "cm/s":
                nasobic = .10
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = 1.00
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)

            elif self.old_velo_unit == "m/s":
                nasobic = 10.00
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)

            elif self.old_velo_unit == "km/h":
                nasobic = 36.00
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)

        elif targetunit == "m/s":
            if self.old_velo_unit == "mm/s":
                nasobic = .001
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)

            elif self.old_velo_unit == "cm/s":
                nasobic = .010
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = .100
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)

            elif self.old_velo_unit == "m/s":
                nasobic = 1.000
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)

            elif self.old_velo_unit == "km/h":
                nasobic = 1/3.600
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)
    
        elif targetunit == "km/h":
            if self.old_velo_unit == "mm/s":
                nasobic = 3.6e-3
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)

            elif self.old_velo_unit == "cm/s":
                nasobic = 3.6e-2
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = 3.6e-1
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)

            elif self.old_velo_unit == "m/s":
                nasobic = 3.6e0
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)

            elif self.old_velo_unit == "km/h":
                nasobic = 1
                for part in self.vy:
                    self.va0veloy = part.get("ID")
                    self.va1veloy = part.get("vy")
                    self.va2veloy = part.get("unit")
                    linevy = {self.key0veloy:    self.va0veloy, self.key1veloy: self.va1veloy*nasobic, self.key2veloy:  targetunit}
                    self.tmp_storage_vy.append(linevy)            
        #endregion

        #region z-velo

        if targetunit == "mm/s":
            if self.old_velo_unit == "mm/s":
                nasobic = 1
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)

            elif self.old_velo_unit == "cm/s":
                nasobic = 10
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = 100
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)

            elif self.old_velo_unit == "m/s":
                nasobic = 1000
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)

            elif self.old_velo_unit == "km/h":
                nasobic = 3600
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)
        
        elif targetunit == "cm/s":
            if self.old_velo_unit == "mm/s":
                nasobic = .1
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)

            elif self.old_velo_unit == "cm/s":
                nasobic = 1.0
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = 10.0
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)

            elif self.old_velo_unit == "m/s":
                nasobic = 100.0
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)

            elif self.old_velo_unit == "km/h":
                nasobic = 360.0
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)

        elif targetunit == "dm/s":
            if self.old_velo_unit == "mm/s":
                nasobic = .01
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)

            elif self.old_velo_unit == "cm/s":
                nasobic = .10
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = 1.00
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)

            elif self.old_velo_unit == "m/s":
                nasobic = 10.00
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)

            elif self.old_velo_unit == "km/h":
                nasobic = 36.00
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)

        elif targetunit == "m/s":
            if self.old_velo_unit == "mm/s":
                nasobic = .001
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)

            elif self.old_velo_unit == "cm/s":
                nasobic = .010
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = .100
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)

            elif self.old_velo_unit == "m/s":
                nasobic = 1.000
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)

            elif self.old_velo_unit == "km/h":
                nasobic = 1/3.600
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)
    
        elif targetunit == "km/h":
            if self.old_velo_unit == "mm/s":
                nasobic = 3.6e-3
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)

            elif self.old_velo_unit == "cm/s":
                nasobic = 3.6e-2
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = 3.6e-1
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)

            elif self.old_velo_unit == "m/s":
                nasobic = 3.6e0
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)

            elif self.old_velo_unit == "km/h":
                nasobic = 1
                for part in self.vz:
                    self.va0veloz= part.get("ID")
                    self.va1veloz = part.get("vz")
                    self.va2veloz= part.get("unit")
                    linevz = {self.key0veloz:    self.va0veloz, self.key1veloz: self.va1veloz*nasobic, self.key2veloz:  targetunit}
                    self.tmp_storage_vz.append(linevz)
        
        #endregion

        #region velo-mag

        if targetunit == "mm/s":
            if self.old_velo_unit == "mm/s":
                nasobic = 1
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)

            elif self.old_velo_unit == "cm/s":
                nasobic = 10
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = 100
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)

            elif self.old_velo_unit == "m/s":
                nasobic = 1000
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)

            elif self.old_velo_unit == "km/h":
                nasobic = 3600
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)
        
        elif targetunit == "cm/s":
            if self.old_velo_unit == "mm/s":
                nasobic = .1
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)

            elif self.old_velo_unit == "cm/s":
                nasobic = 1.0
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = 10.0
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)

            elif self.old_velo_unit == "m/s":
                nasobic = 100.0
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)

            elif self.old_velo_unit == "km/h":
                nasobic = 360.0
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)

        elif targetunit == "dm/s":
            if self.old_velo_unit == "mm/s":
                nasobic = .01
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)

            elif self.old_velo_unit == "cm/s":
                nasobic = .10
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = 1.00
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)

            elif self.old_velo_unit == "m/s":
                nasobic = 10.00
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)

            elif self.old_velo_unit == "km/h":
                nasobic = 36.00
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)

        elif targetunit == "m/s":
            if self.old_velo_unit == "mm/s":
                nasobic = .001
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)

            elif self.old_velo_unit == "cm/s":
                nasobic = .010
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = .100
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)

            elif self.old_velo_unit == "m/s":
                nasobic = 1.000
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)

            elif self.old_velo_unit == "km/h":
                nasobic = 1/3.600
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)
    
        elif targetunit == "km/h":
            if self.old_velo_unit == "mm/s":
                nasobic = 3.6e-3
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)

            elif self.old_velo_unit == "cm/s":
                nasobic = 3.6e-2
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)
            
            elif self.old_velo_unit == "dm/s":
                nasobic = 3.6e-1
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)

            elif self.old_velo_unit == "m/s":
                nasobic = 3.6e0
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)

            elif self.old_velo_unit == "km/h":
                nasobic = 1
                for part in self.vmag:
                    self.va0velomag = part.get("ID")
                    self.va1velomag = part.get("vmag")
                    self.va2velomag = part.get("unit")
                    linevmag = {self.key0velomag:    self.va0velomag, self.key1velomag: self.va1velomag*nasobic, self.key2velomag:  targetunit}
                    self.tmp_storage_vmag.append(linevmag)
        
        #endregion
        
        


        self.velocity = self.tmp_storage_velocity[:]
        self.velocityunit = self.velocity[0].get(self.key2velo)
        self.tmp_storage_velocity.clear()

        self.vx = self.tmp_storage_vx[:]
        self.vy = self.tmp_storage_vy[:]
        self.vz = self.tmp_storage_vz[:]
        self.vmag = self.tmp_storage_vmag[:]
        self.tmp_storage_vx.clear()
        self.tmp_storage_vy.clear()
        self.tmp_storage_vz.clear()
        self.tmp_storage_vmag.clear()
        print(self.velocityunit)
        print("***************UNITSWITCH: VELOCITY - END ***************")
        print(2*"\n")

    def unitswitch_diameter(self, targetunit):
        self.tmp_storage_diameter = []
        self.key0dia = list(self.diameter[0])[0]
        self.key1dia = list(self.diameter[0])[1]
        self.key2dia = list(self.diameter[0])[2]
        self.old_dia_unit = self.diameter[0].get(self.key2dia)
        print(2*"\n")
        print("***************UNITSWITCH: DIAMETER - BEGIN***************")
        print(self.old_dia_unit)
        print(self.key0dia, self.key1dia, self.key2dia)

        if targetunit == "nm":
            if self.old_dia_unit == "nm":
                nasobic = 1
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "um":
                nasobic = 1e3
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)
            
            elif self.old_dia_unit == "mm":
                nasobic = 1e6
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "cm":
                nasobic = 1e7
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "dm":
                nasobic = 1e8
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "m":
                nasobic = 1e9
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

        elif targetunit == "um":
            if self.old_dia_unit == "nm":
                nasobic = 1e-3
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "um":
                nasobic = 1
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)
            
            elif self.old_dia_unit == "mm":
                nasobic = 1e3
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "cm":
                nasobic = 1e4
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "dm":
                nasobic = 1e5
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "m":
                nasobic = 1e6
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

        elif targetunit == "mm":
            if self.old_dia_unit == "nm":
                nasobic = 1e-6
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "um":
                nasobic = 1e-3
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)
            
            elif self.old_dia_unit == "mm":
                nasobic = 1
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "cm":
                nasobic = 10
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "dm":
                nasobic = 100
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "m":
                nasobic = 1000
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

        elif targetunit == "cm":
            if self.old_dia_unit == "nm":
                nasobic = 1e-7
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "um":
                nasobic = 1e-4
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)
            
            elif self.old_dia_unit == "mm":
                nasobic = 1e-1
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "cm":
                nasobic = 1
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "dm":
                nasobic = 10
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "m":
                nasobic = 100
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

        elif targetunit == "dm":
            if self.old_dia_unit == "nm":
                nasobic = 1e-8
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "um":
                nasobic = 1e-5
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)
            
            elif self.old_dia_unit == "mm":
                nasobic = 1e-2
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "cm":
                nasobic = 1e-1
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "dm":
                nasobic = 1
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "m":
                nasobic = 10
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

        elif targetunit == "m":
            if self.old_dia_unit == "nm":
                nasobic = 1e-9
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "um":
                nasobic = 1e-6
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)
            
            elif self.old_dia_unit == "mm":
                nasobic = 1e-3
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "cm":
                nasobic = 1e-2
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "dm":
                nasobic = 1e-1
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

            elif self.old_dia_unit == "m":
                nasobic = 1
                for part in self.diameter:
                    self.va0dia = part.get(self.key0dia)
                    self.va1dia = part.get(self.key1dia)
                    self.va2dia = part.get(self.key2dia)
                    line = {self.key0dia:    self.va0dia, self.key1dia: self.va1dia*nasobic, self.key2dia:  targetunit}
                    self.tmp_storage_diameter.append(line)

        else:
            print("Invalid unit")

        self.diameter = self.tmp_storage_diameter[:]
        self.diameterunit = self.diameter[0].get(self.key2dia)
        self.tmp_storage_diameter.clear()
        print(self.diameterunit)
        print("***************UNITSWITCH: DIAMETER - END***************")
        print(2*"\n")

    def unitswitch_temperature(self, targetunit):
        self.tmp_storage_temperature = []
        self.key0temp = list(self.temperature[0])[0]
        self.key1temp = list(self.temperature[0])[1]
        self.key2temp = list(self.temperature[0])[2]
        self.old_temp_unit = self.temperature[0].get(self.key2temp)
        print(2*"\n")
        print("***************UNITSWITCH: TEMP - BEGIN***************")
        print(self.old_temp_unit) 

        if targetunit == "K":
            posun = 0
            if self.old_temp_unit == "K":
                for part in self.temperature:
                    self.va0temp = part.get(self.key0temp)
                    self.va1temp = part.get(self.key1temp)
                    self.va2temp = part.get(self.key2temp)
                    line = {self.key0temp:    self.va0temp, self.key1temp: self.va1temp + posun, self.key2temp:  targetunit}
                    self.tmp_storage_temperature.append(line)
            
            elif self.old_temp_unit == "°C":
                posun = +273.15
                for part in self.temperature:
                    self.va0temp = part.get(self.key0temp)
                    self.va1temp = part.get(self.key1temp)
                    self.va2temp = part.get(self.key2temp)
                    line = {self.key0temp:    self.va0temp, self.key1temp: self.va1temp + posun, self.key2temp:  targetunit}
                    self.tmp_storage_temperature.append(line)

        elif targetunit == "°C":
            posun = -273.15
            if self.old_temp_unit == "K":
                for part in self.temperature:
                    self.va0temp = part.get(self.key0temp)
                    self.va1temp = part.get(self.key1temp)
                    self.va2temp = part.get(self.key2temp)
                    line = {self.key0temp:    self.va0temp, self.key1temp: self.va1temp + posun, self.key2temp:  targetunit}
                    self.tmp_storage_temperature.append(line)
            
            elif self.old_temp_unit == "°C":
                posun = 0
                for part in self.temperature:
                    self.va0temp = part.get(self.key0temp)
                    self.va1temp = part.get(self.key1temp)
                    self.va2temp = part.get(self.key2temp)
                    line = {self.key0temp:    self.va0temp, self.key1temp: self.va1temp + posun, self.key2temp:  targetunit}
                    self.tmp_storage_temperature.append(line)

        self.temperature = self.tmp_storage_temperature[:]
        self.temperatureunit = self.temperature[0].get(self.key2temp)
        self.tmp_storage_temperature.clear()      
        print(self.temperatureunit)
        print("***************UNITSWITCH: TEMPERATURE - END ***************")
        print(2*"\n")

    def unitswitch_mfr(self,targetunit):
        self.tmp_storage_mfr = []
        self.key0mfr = list(self.mfr[0])[0]
        self.key1mfr = list(self.mfr[0])[1]
        self.key2mfr = list(self.mfr[0])[2]
        self.old_mfr_unit = self.mfr[0].get(self.key2mfr)
        print(2*"\n")
        print("***************UNITSWITCH: MFR - BEGIN***************")
        print(self.old_mfr_unit)         

        if (targetunit == "ug/s"):
            if self.old_mfr_unit == "ug/s":
                nasobic = 1
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)

            elif self.old_mfr_unit == "mg/s":
                nasobic = 1e3
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)

            elif self.old_mfr_unit == "g/s":
                nasobic = 1e6
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)  

            elif self.old_mfr_unit == "kg/s":
                nasobic = 1e9
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)

            elif self.old_mfr_unit == "kg/h":
                nasobic = 1e9/3600
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)                                           

        elif (targetunit == "mg/s"):
            if self.old_mfr_unit == "ug/s":
                nasobic = 1e-3
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)

            elif self.old_mfr_unit == "mg/s":
                nasobic = 1
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)

            elif self.old_mfr_unit == "g/s":
                nasobic = 1e3
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)  

            elif self.old_mfr_unit == "kg/s":
                nasobic = 1e6
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)

            elif self.old_mfr_unit == "kg/h":
                nasobic = 1e6/3600
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)      

        elif (targetunit == "g/s"):
            if self.old_mfr_unit == "ug/s":
                nasobic = 1e-6
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)

            elif self.old_mfr_unit == "mg/s":
                nasobic = 1e-3
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)

            elif self.old_mfr_unit == "g/s":
                nasobic = 1
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)  

            elif self.old_mfr_unit == "kg/s":
                nasobic = 1e3
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)

            elif self.old_mfr_unit == "kg/h":
                nasobic = 1e3/3600
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)     

        elif (targetunit == "kg/s"):
            if self.old_mfr_unit == "ug/s":
                nasobic = 1e-9
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)

            elif self.old_mfr_unit == "mg/s":
                nasobic = 1e-6
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)

            elif self.old_mfr_unit == "g/s":
                nasobic = 1e-3
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)  

            elif self.old_mfr_unit == "kg/s":
                nasobic = 1
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)

            elif self.old_mfr_unit == "kg/h":
                nasobic = 1/3600
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)     

        elif (targetunit == "kg/h"):
            if self.old_mfr_unit == "ug/s":
                nasobic = 1e-9*3600
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)

            elif self.old_mfr_unit == "mg/s":
                nasobic = 1e-6*3600
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)

            elif self.old_mfr_unit == "g/s":
                nasobic = 1e-3*3600
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)  

            elif self.old_mfr_unit == "kg/s":
                nasobic = 3600
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)

            elif self.old_mfr_unit == "kg/h":
                nasobic = 1
                for part in self.mfr:
                    self.va0mfr = part.get(self.key0mfr)
                    self.va1mfr = part.get(self.key1mfr)
                    self.va2mfr = part.get(self.key2mfr)
                    line = {self.key0mfr:    self.va0mfr, self.key1mfr: self.va1mfr*nasobic, self.key2mfr:  targetunit}
                    self.tmp_storage_mfr.append(line)   

        self.mfr = self.tmp_storage_mfr[:]
        self.mfrunit = self.mfr[0].get(self.key2mfr)
        self.tmp_storage_mfr.clear()     
        print(self.mfrunit)
        print("***************UNITSWITCH: MFR - END***************")
        print(2*"\n")

    def unitswitch_mass(self,targetunit):
        self.tmp_storage_mass = []
        self.key0mass = list(self.mass[0])[0]
        self.key1mass = list(self.mass[0])[1]
        self.key2mass = list(self.mass[0])[2]
        self.old_mass_unit = self.mass[0].get(self.key2mass)
        print(2*"\n")
        print("***************UNITSWITCH: MASS - BEGIN***************")
        print(self.old_mass_unit)         

        if (targetunit == "ng"):
            if self.old_mass_unit == "ng":
                nasobic = 1
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)

            elif self.old_mass_unit == "ug":
                nasobic = 1e3
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)

            elif self.old_mass_unit == "mg":
                nasobic = 1e6
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)

            elif self.old_mass_unit == "g":
                nasobic = 1e9
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)  

            elif self.old_mass_unit == "kg":
                nasobic = 1e12
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)
        
        elif (targetunit == "ug"):
            if self.old_mass_unit == "ng":
                nasobic = 1e-3
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)

            elif self.old_mass_unit == "ug":
                nasobic = 1
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)

            elif self.old_mass_unit == "mg":
                nasobic = 1e3
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)

            elif self.old_mass_unit == "g":
                nasobic = 1e6
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)  

            elif self.old_mass_unit == "kg":
                nasobic = 1e9
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)
                                       

        elif (targetunit == "mg"):
            if self.old_mass_unit == "ng":
                nasobic = 1e-6
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)

            elif self.old_mass_unit == "ug":
                nasobic = 1e-3
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)

            elif self.old_mass_unit == "mg":
                nasobic = 1
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)

            elif self.old_mass_unit == "g":
                nasobic = 1e3
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)  

            elif self.old_mass_unit == "kg":
                nasobic = 1e6
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)
     

        elif (targetunit == "g"):
            if self.old_mass_unit == "ng":
                nasobic = 1e-9
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)

            elif self.old_mass_unit == "ug":
                nasobic = 1e-6
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)

            elif self.old_mass_unit == "mg":
                nasobic = 1e-3
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)

            elif self.old_mass_unit == "g":
                nasobic = 1
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)  

            elif self.old_mass_unit == "kg":
                nasobic = 1e3
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)
   

        elif (targetunit == "kg"):
            if self.old_mass_unit == "ng":
                nasobic = 1e-12
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)

            elif self.old_mass_unit == "ug":
                nasobic = 1e-9
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)

            elif self.old_mass_unit == "mg":
                nasobic = 1e-6
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)

            elif self.old_mass_unit == "g":
                nasobic = 1e-3
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)  

            elif self.old_mass_unit == "kg":
                nasobic = 1
                for part in self.mass:
                    self.va0mass = part.get(self.key0mass)
                    self.va1mass = part.get(self.key1mass)
                    self.va2mass = part.get(self.key2mass)
                    line = {self.key0mass:    self.va0mass, self.key1mass: self.va1mass*nasobic, self.key2mass:  targetunit}
                    self.tmp_storage_mass.append(line)


        self.mass = self.tmp_storage_mass[:]
        self.massunit = self.mass[0].get(self.key2mass)
        self.tmp_storage_mass.clear()     
        print(self.massunit)
        print("***************UNITSWITCH: MASS - END***************")      
        print(2*"\n")

    def unitswitch_frequency(self, targetunit):
        self.tmp_storage_frequency = []
        self.key0freq = list(self.frequency[0])[0]
        self.key1freq = list(self.frequency[0])[1]
        self.key2freq = list(self.frequency[0])[2]
        self.old_freq_unit = self.frequency[0].get(self.key2freq)
        print(2*"\n")
        print("***************UNITSWITCH: FREQUENCY - BEGIN ***************")      
        print(self.old_freq_unit)   


        if (targetunit == "1/ms"):
            if self.old_freq_unit == "1/ms":
                nasobic = 1
                for part in self.frequency:
                    self.va0freq = part.get(self.key0freq)
                    self.va1freq = part.get(self.key1freq)
                    self.va2freq = part.get(self.key2freq)
                    line = {self.key0freq:    self.va0freq, self.key1freq: self.va1freq*nasobic, self.key2freq:  targetunit}
                    self.tmp_storage_frequency.append(line)       

            elif self.old_freq_unit == "1/s":
                nasobic = 1000
                for part in self.frequency:
                    self.va0freq = part.get(self.key0freq)
                    self.va1freq = part.get(self.key1freq)
                    self.va2freq = part.get(self.key2freq)
                    line = {self.key0freq:    self.va0freq, self.key1freq: self.va1freq*nasobic, self.key2freq:  targetunit}
                    self.tmp_storage_frequency.append(line)   

            elif self.old_freq_unit == "1/min":
                nasobic = 60*1000
                for part in self.frequency:
                    self.va0freq = part.get(self.key0freq)
                    self.va1freq = part.get(self.key1freq)
                    self.va2freq = part.get(self.key2freq)
                    line = {self.key0freq:    self.va0freq, self.key1freq: self.va1freq*nasobic, self.key2freq:  targetunit}
                    self.tmp_storage_frequency.append(line)   

            elif self.old_freq_unit == "1/h":
                nasobic = 60*60*1000
                for part in self.frequency:
                    self.va0freq = part.get(self.key0freq)
                    self.va1freq = part.get(self.key1freq)
                    self.va2freq = part.get(self.key2freq)
                    line = {self.key0freq:    self.va0freq, self.key1freq: self.va1freq*nasobic, self.key2freq:  targetunit}
                    self.tmp_storage_frequency.append(line)   
        
        elif (targetunit == "1/s"):
            if self.old_freq_unit == "1/ms":
                nasobic = 1e-3
                for part in self.frequency:
                    self.va0freq = part.get(self.key0freq)
                    self.va1freq = part.get(self.key1freq)
                    self.va2freq = part.get(self.key2freq)
                    line = {self.key0freq:    self.va0freq, self.key1freq: self.va1freq*nasobic, self.key2freq:  targetunit}
                    self.tmp_storage_frequency.append(line)       

            elif self.old_freq_unit == "1/s":
                nasobic = 1
                for part in self.frequency:
                    self.va0freq = part.get(self.key0freq)
                    self.va1freq = part.get(self.key1freq)
                    self.va2freq = part.get(self.key2freq)
                    line = {self.key0freq:    self.va0freq, self.key1freq: self.va1freq*nasobic, self.key2freq:  targetunit}
                    self.tmp_storage_frequency.append(line)   

            elif self.old_freq_unit == "1/min":
                nasobic = 60
                for part in self.frequency:
                    self.va0freq = part.get(self.key0freq)
                    self.va1freq = part.get(self.key1freq)
                    self.va2freq = part.get(self.key2freq)
                    line = {self.key0freq:    self.va0freq, self.key1freq: self.va1freq*nasobic, self.key2freq:  targetunit}
                    self.tmp_storage_frequency.append(line)   

            elif self.old_freq_unit == "1/h":
                nasobic = 60*60
                for part in self.frequency:
                    self.va0freq = part.get(self.key0freq)
                    self.va1freq = part.get(self.key1freq)
                    self.va2freq = part.get(self.key2freq)
                    line = {self.key0freq:    self.va0freq, self.key1freq: self.va1freq*nasobic, self.key2freq:  targetunit}
                    self.tmp_storage_frequency.append(line) 

        elif (targetunit == "1/min"):
            if self.old_freq_unit == "1/ms":
                nasobic = 60*1000
                for part in self.frequency:
                    self.va0freq = part.get(self.key0freq)
                    self.va1freq = part.get(self.key1freq)
                    self.va2freq = part.get(self.key2freq)
                    line = {self.key0freq:    self.va0freq, self.key1freq: self.va1freq*nasobic, self.key2freq:  targetunit}
                    self.tmp_storage_frequency.append(line)       

            elif self.old_freq_unit == "1/s":
                nasobic = 60
                for part in self.frequency:
                    self.va0freq = part.get(self.key0freq)
                    self.va1freq = part.get(self.key1freq)
                    self.va2freq = part.get(self.key2freq)
                    line = {self.key0freq:    self.va0freq, self.key1freq: self.va1freq*nasobic, self.key2freq:  targetunit}
                    self.tmp_storage_frequency.append(line)   

            elif self.old_freq_unit == "1/min":
                nasobic = 1
                for part in self.frequency:
                    self.va0freq = part.get(self.key0freq)
                    self.va1freq = part.get(self.key1freq)
                    self.va2freq = part.get(self.key2freq)
                    line = {self.key0freq:    self.va0freq, self.key1freq: self.va1freq*nasobic, self.key2freq:  targetunit}
                    self.tmp_storage_frequency.append(line)   

            elif self.old_freq_unit == "1/h":
                nasobic = 1/60
                for part in self.frequency:
                    self.va0freq = part.get(self.key0freq)
                    self.va1freq = part.get(self.key1freq)
                    self.va2freq = part.get(self.key2freq)
                    line = {self.key0freq:    self.va0freq, self.key1freq: self.va1freq*nasobic, self.key2freq:  targetunit}
                    self.tmp_storage_frequency.append(line) 
 
        elif (targetunit == "1/h"):
            if self.old_freq_unit == "1/ms":
                nasobic = 3600000
                for part in self.frequency:
                    self.va0freq = part.get(self.key0freq)
                    self.va1freq = part.get(self.key1freq)
                    self.va2freq = part.get(self.key2freq)
                    line = {self.key0freq:    self.va0freq, self.key1freq: self.va1freq*nasobic, self.key2freq:  targetunit}
                    self.tmp_storage_frequency.append(line)       

            elif self.old_freq_unit == "1/s":
                nasobic = 3600
                for part in self.frequency:
                    self.va0freq = part.get(self.key0freq)
                    self.va1freq = part.get(self.key1freq)
                    self.va2freq = part.get(self.key2freq)
                    line = {self.key0freq:    self.va0freq, self.key1freq: self.va1freq*nasobic, self.key2freq:  targetunit}
                    self.tmp_storage_frequency.append(line)   

            elif self.old_freq_unit == "1/min":
                nasobic = 60
                for part in self.frequency:
                    self.va0freq = part.get(self.key0freq)
                    self.va1freq = part.get(self.key1freq)
                    self.va2freq = part.get(self.key2freq)
                    line = {self.key0freq:    self.va0freq, self.key1freq: self.va1freq*nasobic, self.key2freq:  targetunit}
                    self.tmp_storage_frequency.append(line)   

            elif self.old_freq_unit == "1/h":
                nasobic = 1
                for part in self.frequency:
                    self.va0freq = part.get(self.key0freq)
                    self.va1freq = part.get(self.key1freq)
                    self.va2freq = part.get(self.key2freq)
                    line = {self.key0freq:    self.va0freq, self.key1freq: self.va1freq*nasobic, self.key2freq:  targetunit}
                    self.tmp_storage_frequency.append(line) 
        self.frequency = self.tmp_storage_frequency[:]
        self.frequencyunit = self.frequency[0].get(self.key2freq)
        self.tmp_storage_frequency.clear()        
        print(self.frequencyunit)
        print("***************UNITSWITCH: FREQUENCY - END***************")      
        print(2*"\n")

    def unitswitch_time(self,targetunit):
        self.tmp_storage_time = []
        self.key0time = list(self.time[0])[0]
        self.key1time = list(self.time[0])[1]
        self.key2time = list(self.time[0])[2]
        self.old_time_unit = self.time[0].get(self.key2time)
        print(2*"\n")
        print("***************UNITSWITCH: TIME - BEGIN***************")
        print(self.old_time_unit)

        if (targetunit == "ms"):
            if self.old_time_unit == "ms":
                nasobic = 1
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line)  

            elif self.old_time_unit == "s":
                nasobic = 1000
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line)  

            elif self.old_time_unit == "min":
                nasobic = 60*1000
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line)  

            elif self.old_time_unit == "h":
                nasobic = 60*60*1000
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line) 

            elif self.old_time_unit == "den":
                nasobic = 24*60*60*1000
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line) 

        elif (targetunit == "s"):
            if self.old_time_unit == "ms":
                nasobic = 1/60
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line)  

            elif self.old_time_unit == "s":
                nasobic = 1
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line)  

            elif self.old_time_unit == "min":
                nasobic = 60
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line)  

            elif self.old_time_unit == "h":
                nasobic = 60*60
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line) 

            elif self.old_time_unit == "den":
                nasobic = 24*60*60
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line) 

        elif (targetunit == "min"):
            if self.old_time_unit == "ms":
                nasobic = 1/(60*1000)
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line)  

            elif self.old_time_unit == "s":
                nasobic = 1/60
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line)  

            elif self.old_time_unit == "min":
                nasobic = 1
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line)  

            elif self.old_time_unit == "h":
                nasobic = 60
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line) 

            elif self.old_time_unit == "den":
                nasobic = 24*60
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line) 

        elif (targetunit == "h"):
            if self.old_time_unit == "ms":
                nasobic = 1/(60*60*1000)
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line)  

            elif self.old_time_unit == "s":
                nasobic = 1/(60*60)
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line)  

            elif self.old_time_unit == "min":
                nasobic = 1/60
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line)  

            elif self.old_time_unit == "h":
                nasobic = 1
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line) 

            elif self.old_time_unit == "den":
                nasobic = 24
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line) 

        elif (targetunit == "den"):
            if self.old_time_unit == "ms":
                nasobic = 1/(24*60*60*1000)
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line)  

            elif self.old_time_unit == "s":
                nasobic = 1/(24*60*60)
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line)  

            elif self.old_time_unit == "min":
                nasobic = 1/(24*60)
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line)  

            elif self.old_time_unit == "h":
                nasobic = 1/24
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line) 

            elif self.old_time_unit == "den":
                nasobic = 1
                for part in self.time:
                    self.va0time = part.get(self.key0time)
                    self.va1time = part.get(self.key1time)
                    self.va2time = part.get(self.key2time)
                    line = {self.key0time:    self.va0time, self.key1time: self.va1time*nasobic, self.key2time:  targetunit}
                    self.tmp_storage_time.append(line) 
        self.time = self.tmp_storage_time[:]
        self.timeunit = self.time[0].get(self.key2time)
        self.tmp_storage_time.clear()
        print(self.timeunit)
        print("***************UNITSWITCH: TIME - END ***************")
        print(2*"\n")

    #endregion


    def dpmhisto_single(self, dataframe):
        print("***************************")
        print("DPM HISTO SINGLE... BEGIN")
        if self.workdir == "":
            self.workdir = getcwd()
        
        self.tmpstorage_single = []  
        self.update_units()
        nazev_a:str()
        nazev_b:str()
        print("DPM HISTO SINGLE... Dataframe: {}".format(str(dataframe[0])))
        self.key = list(dataframe[0])[1]

        for i in range (0, len(dataframe)):
            self.tmpstorage_single.append(dataframe[i].get(self.key))
        print(self.tmpstorage_single[0:10])
        if (self.bool_distribution_unit == True):
            plt.hist(self.tmpstorage_single, weights = ones(len(self.tmpstorage_single)) / len(self.tmpstorage_single), bins= self.val_bin_single, rwidth= 0.5)
        else:
            plt.hist(self.tmpstorage_single, bins= self.val_bin_single,  rwidth= 0.5)

        self.key = list(dataframe[0])[1]
        print("DPM HISTO SINGLE... key: {}, bins: {}, rangecount: {}, rwidth: {}".format(self.key, self.val_bin_single, self.range_count, (1/self.range_count)-self.range_count/8))
        print("DPM HISTO SINGLE... plot data line 1: {}".format(self.tmpstorage_single[0]))
        print("DPM HISTO SINGLE... language: {}".format(self.language))
        #region popisek osy x
        if self.language == "CZ":
            if (self.key == "time"):
                plt.xlabel("Čas [{}]".format(self.timeunit))
                plt.title("Záchyt částic v čase")
                nazev = "_TIME"
            elif (self.key == "diameter"):
                plt.xlabel("Průměr [{}]".format(self.diameterunit))
                plt.title("Distribuce průměrů částic")
                nazev = "_PRUMER"
            elif (self.key == "temperature"):
                plt.xlabel("Teplota [{}]".format(self.temperatureunit))
                plt.title("Distribuce teplot částic")
                nazev = "_TEPLOTA"
            elif (self.key == "position"):
                plt.xlabel("Pozice [{}]".format(self.positionunit))
                plt.title("Distribuce pozice částic")
                nazev = "_POZICE"
            elif (self.key == "velocity"):
                plt.xlabel("Rychlost [{}]".format(self.velocityunit))
                plt.title("Distribuce rychlosti částic")
                nazev = "_RYCHLOST"
            elif (self.key == "mass-flow-rate"):
                plt.xlabel("Hmotnostní průtok [{}]".format(self.mfrunit))
                plt.title("Distribuce hm. průtoku částic")
                nazev = "_MFR"
            elif (self.key == "mass"):
                plt.xlabel("Hmotnost [{}]".format(self.massunit))
                plt.title("Distribuce hmotnosti částic")
                nazev = "_HMOTNOST"
            elif (self.key == "frequency"):
                plt.xlabel("Frekvence [{}]".format(self.frequencyunit))
                plt.title("Distribuce frekvence emisí částic")
                nazev = "_FREKVENCE"
            elif (self.key == "x"):
                plt.xlabel("Pozice X [{}]".format(self.positionunit))
                plt.title("Distribuce pozic X částic")
                nazev = "_X"
            elif (self.key == "y"):
                plt.xlabel("Pozice Y [{}]".format(self.positionunit))
                plt.title("Distribuce pozic Y částic")
                nazev = "_Y"
            elif (self.key == "z"):
                plt.xlabel("Pozice Z [{}]".format(self.positionunit))
                plt.title("Distribuce pozic Z částic")
                nazev = "_Z"
            elif (self.key == "vx"):
                plt.xlabel("Rychlost X [{}]".format(self.velocityunit))
                plt.title("Distribuce rychlostí X částic")
                nazev = "_Rychlost_X"
            elif (self.key == "vy"):
                plt.xlabel("Rychlost Y [{}]".format(self.velocityunit))
                plt.title("Distribuce rychlostí Y částic")
                nazev = "_Rychlost_Y"
            elif (self.key == "vz"):
                plt.xlabel("Rychlost Z [{}]".format(self.velocityunit))
                plt.title("Distribuce rychlostí Z částic")
                nazev = "_Rychlost_Z"
            elif (self.key == "vmag"):
                plt.xlabel("Rychlost (výslednice) [{}]".format(self.velocityunit))
                plt.title("Distribuce rychlostí částic")
                nazev = "_RYCHLOST_MAG"
            else:
                pass
        
        elif self.language == "EN":
            if (self.key == "time"):
                plt.xlabel("Time [{}]".format(self.timeunit))
                plt.title("Particle Passage Time Distribution")
                nazev = "_TIME"
            elif (self.key == "diameter"):
                plt.xlabel("Diameter [{}]".format(self.diameterunit))
                plt.title("Particle Diameter Distribution")
                nazev = "_PRUMER"
            elif (self.key == "temperature"):
                plt.xlabel("Temperature [{}]".format(self.temperatureunit))
                plt.title("Particle Temperature Distribution")
                nazev = "_TEPLOTA"
            elif (self.key == "position"):
                plt.xlabel("Position [{}]".format(self.positionunit))
                plt.title("Particle Position Distribution")
                nazev = "_POZICE"
            elif (self.key == "velocity"):
                plt.xlabel("Velocity [{}]".format(self.velocityunit))
                plt.title("Particle Velocity distribution")
                nazev = "_RYCHLOST"
            elif (self.key == "mass-flow-rate"):
                plt.xlabel("Mass Flow Rate [{}]".format(self.mfrunit))
                plt.title("Particle Mass Flow Rate Distribution")
                nazev = "_MFR"
            elif (self.key == "mass"):
                plt.xlabel("Mass [{}]".format(self.massunit))
                plt.title("Particle Mass Distribution")
                nazev = "_HMOTNOST"
            elif (self.key == "frequency"):
                plt.xlabel("Emission Frequency [{}]".format(self.frequencyunit))
                plt.title("Particle Emission Frequency Distribution")
                nazev = "_FREKVENCE"
            elif (self.key == "x"):
                plt.xlabel("X position[{}]".format(self.positionunit))
                plt.title("Particle X Position Distribution")
                nazev = "_X"
            elif (self.key == "y"):
                plt.xlabel("Y Position [{}]".format(self.positionunit))
                plt.title("Particle Y Position Distribution")
                nazev = "_Y"
            elif (self.key == "z"):
                plt.xlabel("Z Position [{}]".format(self.positionunit))
                plt.title("Particle Z Position Distribution")
                nazev = "_Z"
            elif (self.key == "vx"):
                plt.xlabel("X Velocity [{}]".format(self.velocityunit))
                plt.title("Particle X Velocity Distribution")
                nazev = "_Rychlost_X"
            elif (self.key == "vy"):
                plt.xlabel("Y Velocity [{}]".format(self.velocityunit))
                plt.title("Particle Y Velocity Distribution")
                nazev = "_Rychlost_Y"
            elif (self.key == "vz"):
                plt.xlabel("Z Velocity [{}]".format(self.velocityunit))
                plt.title("Particle Z Velocity Distribution")
                nazev = "_Rychlost_Z"
            elif (self.key == "vmag"):
                plt.xlabel("Velocity Magnitude [{}]".format(self.velocityunit))
                plt.title("Particle Velocity Distribution")
                nazev = "_RYCHLOST_MAG"
            else:
                pass
        #endregion

        #region popisek osy y
        if self.bool_distribution_unit == True:
            self.distributionunit = "%"
            plt.gca().yaxis.set_major_formatter(PF(1))
        else:
            self.distributionunit = "-"
        if self.language == "CZ":
            plt.ylabel("Zastoupení [{}]".format(self.distributionunit))
        elif self.language == "EN":
            plt.ylabel("Ammount [{}]".format(self.distributionunit))
        
    #endregion
        print("DPM HISTO SINGLE: Histobool {}".format(self.histo_xlim_bool))
        
        if self.histo_xlim_bool == True:
            plt.xlim(self.xmin_histo, self.xmax_histo)
            print("DPM HISTO: XMIN:    {}; XMAX:   {}".format(self.xmin_histo, self.xmax_histo))
        ax = plt.gca()
        ax.patch.set_facecolor('xkcd:pale blue')
        ax.patch.set_alpha(0.5)
        namefile_p1 = "histo_{}__".format(len(ls(self.workdir)))
        namefile_p2 = datetime.now().strftime("%d-%m-%Y-%H-%M-%S")
        namefile_p3 = "__img.png"
        namefile = namefile_p1 + namefile_p2 + nazev + namefile_p3
        print("DPM HISTO SINGLE... file name: {}".format(namefile))
        plt.savefig(self.workdir +"/"+ namefile, dpi = self.val_dpi, bbox_inches = "tight")
        plt.clf()
        plt.close()

        self.tmpstorage_single.clear()
        print("DPM HISTO SINGLE... END")       
        print("***************************")

    def dpmhisto_multiple(self,dataframe):
        print("***************************")
        print("DPM HISTO MULTI... BEGIN")
        if self.workdir == "":
            self.workdir = getcwd()
        self.tmpstorage_multi = []  
        self.update_units()
        nazev_a:str()
        nazev_b:str()
        print("DPM HISTO MULTIPLE... DATAFRAME \n{}".format(str(dataframe)[0:150]))
        self.key = list(dataframe[0][0])[1]
        print(10*"\n", "DPM HISTO MULTIPLE: key = {}".format(self.key))
        tmp_storage_ones = []
        self.dataset_length = 0
        for i in range (0, len(dataframe)):
            length = 0

            self.tmpstorage_multi.append([])
            tmp_storage_ones.append([])
            print("DPM HISTO MULTIPLE... současný dataframe part {}".format(str(dataframe[i])[0:100]))
            for j in range (0, len(dataframe[i])):
                self.tmpstorage_multi[i].append(dataframe[i][j].get(self.key))
                length += 1
                self.dataset_length += 1
            tmp_storage_ones[i].append(     ones(length)     / length)
            print("DPM HISTO MULTIPLE... nyní je počet položek seznamu {} a poslední položka podseznamu má délku {}".format(len(self.tmpstorage_multi), len(self.tmpstorage_multi[i])))
        #weights = ones(length) / length
        #print (self.tmpstorage_multi)
        #weights_generation
        print("DPM HISTO MULTI: velikost balíčku dat: {}, to je o {} víc než na začátku!".format(self.dataset_length, (len(self.x) - self.dataset_length)/len(self.x)*100))

        for object in self.tmpstorage_multi:
            print(len(object))
        if (self.bool_distribution_unit == True):
            plt.hist(self.tmpstorage_multi, bins= self.val_bin_multi,  rwidth= 1/len(self.tmpstorage_multi)+0.5, label =self.hist_labels)
            #plt.hist(self.tmpstorage_multi, weights = tmp_storage_ones, bins= self.val_bin_single, rwidth= (1/self.range_count)-self.range_count/8)
            
        else:
            plt.hist(self.tmpstorage_multi, bins= self.val_bin_multi,  rwidth= 0.5)

        #region popisek osy x
        if self.language == "CZ":
            if (self.key == "time"):
                plt.xlabel("Čas [{}]".format(self.timeunit))
                plt.title("Záchyt částic v čase")
                nazev = "_TIME"
            elif (self.key == "diameter"):
                plt.xlabel("Průměr [{}]".format(self.diameterunit))
                plt.title("Distribuce průměrů částic")
                nazev = "_PRUMER"
            elif (self.key == "temperature"):
                plt.xlabel("Teplota [{}]".format(self.temperatureunit))
                plt.title("Distribuce teplot částic")
                nazev = "_TEPLOTA"
            elif (self.key == "position"):
                plt.xlabel("Pozice [{}]".format(self.positionunit))
                plt.title("Distribuce pozice částic")
                nazev = "_POZICE"
            elif (self.key == "velocity"):
                plt.xlabel("Rychlost [{}]".format(self.velocityunit))
                plt.title("Distribuce rychlosti částic")
                nazev = "_RYCHLOST"
            elif (self.key == "mass-flow-rate"):
                plt.xlabel("Hmotnostní průtok [{}]".format(self.mfrunit))
                plt.title("Distribuce hm. průtoku částic")
                nazev = "_MFR"
            elif (self.key == "mass"):
                plt.xlabel("Hmotnost [{}]".format(self.massunit))
                plt.title("Distribuce hmotnosti částic")
                nazev = "_HMOTNOST"
            elif (self.key == "frequency"):
                plt.xlabel("Frekvence [{}]".format(self.frequencyunit))
                plt.title("Distribuce frekvence emisí částic")
                nazev = "_FREKVENCE"
            elif (self.key == "x"):
                plt.xlabel("Pozice X [{}]".format(self.positionunit))
                plt.title("Distribuce pozic X částic")
                nazev = "_X"
            elif (self.key == "y"):
                plt.xlabel("Pozice Y [{}]".format(self.positionunit))
                plt.title("Distribuce pozic Y částic")
                nazev = "_Y"
            elif (self.key == "z"):
                plt.xlabel("Pozice Z [{}]".format(self.positionunit))
                plt.title("Distribuce pozic Z částic")
                nazev = "_Z"
            elif (self.key == "vx"):
                plt.xlabel("Rychlost X [{}]".format(self.velocityunit))
                plt.title("Distribuce rychlostí X částic")
                nazev = "_Rychlost_X"
            elif (self.key == "vy"):
                plt.xlabel("Rychlost Y [{}]".format(self.velocityunit))
                plt.title("Distribuce rychlostí Y částic")
                nazev = "_Rychlost_Y"
            elif (self.key == "vz"):
                plt.xlabel("Rychlost Z [{}]".format(self.velocityunit))
                plt.title("Distribuce rychlostí Z částic")
                nazev = "_Rychlost_Z"
            elif (self.key == "vmag"):
                plt.xlabel("Rychlost (výslednice) [{}]".format(self.velocityunit))
                plt.title("Distribuce rychlostí částic")
                nazev = "_RYCHLOST_MAG"
            else:
                pass
        
        elif self.language == "EN":
            if (self.key == "time"):
                plt.xlabel("Time [{}]".format(self.timeunit))
                plt.title("Particle Passage Time Distribution")
                nazev = "_TIME"
            elif (self.key == "diameter"):
                plt.xlabel("Průměr [{}]".format(self.diameterunit))
                plt.title("Diameter Diameter Distribution")
                nazev = "_PRUMER"
            elif (self.key == "temperature"):
                plt.xlabel("Temperature [{}]".format(self.temperatureunit))
                plt.title("Particle Temperature Distribution")
                nazev = "_TEPLOTA"
            elif (self.key == "position"):
                plt.xlabel("Position [{}]".format(self.positionunit))
                plt.title("Particle Position Distribution")
                nazev = "_POZICE"
            elif (self.key == "velocity"):
                plt.xlabel("Velocity [{}]".format(self.velocityunit))
                plt.title("Particle Velocity Distribution")
                nazev = "_RYCHLOST"
            elif (self.key == "mass-flow-rate"):
                plt.xlabel("Mass Flow Rate [{}]".format(self.mfrunit))
                plt.title("Particle Mass Flow Rate Distribution")
                nazev = "_MFR"
            elif (self.key == "mass"):
                plt.xlabel("Mass [{}]".format(self.massunit))
                plt.title("Particle Mass Distribution")
                nazev = "_HMOTNOST"
            elif (self.key == "frequency"):
                plt.xlabel("Emission Frequency [{}]".format(self.frequencyunit))
                plt.title("Particle Emission Frequency Distribution")
                nazev = "_FREKVENCE"
            elif (self.key == "x"):
                plt.xlabel("X Position [{}]".format(self.positionunit))
                plt.title("Particle X Position Distribution")
                nazev = "_X"
            elif (self.key == "y"):
                plt.xlabel("Y Position [{}]".format(self.positionunit))
                plt.title("Particle Y Position Distribution")
                nazev = "_Y"
            elif (self.key == "z"):
                plt.xlabel("Z Position [{}]".format(self.positionunit))
                plt.title("Particle Z Position Distribution")
                nazev = "_Z"
            elif (self.key == "vx"):
                plt.xlabel("X Velocity [{}]".format(self.velocityunit))
                plt.title("Particle X Velocity Distribution")
                nazev = "_Rychlost_X"
            elif (self.key == "vy"):
                plt.xlabel("Y Velocity [{}]".format(self.velocityunit))
                plt.title("Particle Y Velocity Distribution")
                nazev = "_Rychlost_Y"
            elif (self.key == "vz"):
                plt.xlabel("Z Velocity [{}]".format(self.velocityunit))
                plt.title("Particle Z Velocity Distribution")
                nazev = "_Rychlost_Z"
            elif (self.key == "vmag"):
                plt.xlabel("Velocity Magnitude [{}]".format(self.velocityunit))
                plt.title("Particle Velocity Distribution")
                nazev = "_RYCHLOST_MAG"
            else:
                pass
        #endregion

        #region popisek osy y 
        self.distributionunit = "-"
        if self.language == "CZ":
            plt.ylabel("Zastoupení [{}]".format(self.distributionunit))
        elif self.language == "EN":
            plt.ylabel("Ammount [{}]".format(self.distributionunit))
        #endregion
        
        #region legenda
        leg = plt.legend(bbox_to_anchor = (0.5,-0.35), loc='center', ncol=2)
        leg.get_frame().set_alpha(0.75)
        #endregion
        if self.xlim_bool == True:
            plt.xlim(self.xmin_multihisto, self.xmax_multihisto)
        ax = plt.gca()
        ax.patch.set_facecolor('xkcd:pale blue')
        ax.patch.set_alpha(0.5)
        namefile_p1 = "histo_multiple_{}__".format(len(ls(self.workdir)))
        namefile_p2 = datetime.now().strftime("%d-%m-%Y-%H-%M-%S")
        namefile_p3 = "__img.png"
        namefile = namefile_p1 + namefile_p2 + nazev + namefile_p3
        print(namefile)
        plt.savefig(self.workdir +"/"+ namefile, dpi = self.val_dpi, bbox_inches = "tight")
        plt.clf()
        plt.close()
        tmp_storage_ones.clear()
        self.tmpstorage_multi.clear()
        print("DPM HISTO MULTI... CLEAR STORAGE: {}".format(self.tmpstorage_multi))  
        self.hist_labels.clear()
        self.range_vals.clear()
        print("DPM HISTO MULTI... END")  
        print("***************************")
    
    def dpmscatter(self,dataframe_x, dataframe_y, *args):
        if self.workdir == "":
            self.workdir = getcwd()
        self.tmpstorage_x = []
        self.tmpstorage_y = []
        self.tmpstorage_color = []
        self.update_units()
        nazev_x:str()
        nazev_y:str()
        nazev_color:str()
        nazev_tot:str()
        self.key_x = list(dataframe_x[0])[1]
        self.key_y = list(dataframe_y[0])[1]


        fig = plt.figure()
        for i in range (0, len(dataframe_x)):
            self.tmpstorage_x.append(dataframe_x[i].get(self.key_x))
            self.tmpstorage_y.append(dataframe_y[i].get(self.key_y))

        if len(args) == 0:
            self.tmpstorage_color = ""
            sc = fig.add_subplot(111)
            sc.scatter(self.tmpstorage_x, self.tmpstorage_y, s = self.size,  edgecolors='black', linewidth=0.1)
        
        else:
            self.tmp_storage_color_tmp = args[0][:]
            self.key_color = list(self.tmp_storage_color_tmp[0])[1]
            for i in range (0, len(self.tmp_storage_color_tmp)):
                self.tmpstorage_color.append(self.tmp_storage_color_tmp[i].get(self.key_color))

    
            tmpmin = min(self.tmpstorage_color)
            tmpmax = max(self.tmpstorage_color)
            if (abs(tmpmax)-abs(tmpmin)) == 0:
                tmpmin = tmpmin*0.99
                tmpmax = tmpmax*1.01
            tmpstep = (tmpmax - tmpmin)/10
            #colorbar = arange(tmpmin, tmpmax, tmpstep)
            sc = plt.scatter(self.tmpstorage_x, self.tmpstorage_y, s = self.size, c = self.tmpstorage_color, cmap = self.cmap, edgecolors='black', linewidth=0.1)
            colorbar = plt.colorbar(sc)

        #region X-label
        if self.language == "CZ":
            if (self.key_x == "time"):
                plt.xlabel("Čas [{}]".format(self.timeunit))
                nazev_x = "_xTIME"
            elif (self.key_x == "diameter"):
                plt.xlabel("Průměr [{}]".format(self.diameterunit))
                nazev_x = "_xPRUMER"
            elif (self.key_x == "temperature"):
                plt.xlabel("Teplota [{}]".format(self.temperatureunit))
                nazev_x = "_xTEPLOTA"
            elif (self.key_x == "position"):
                plt.xlabel("Pozice [{}]".format(self.positionunit))
                nazev_x = "_xPOZICE"
            elif (self.key_x == "velocity"):
                plt.xlabel("Rychlost [{}]".format(self.velocityunit))
                nazev_x = "_xRYCHLOST"
            elif (self.key_x == "mass-flow-rate"):
                plt.xlabel("Hmotnostní průtok [{}]".format(self.mfrunit))
                nazev_x = "_xMFR"
            elif (self.key_x == "mass"):
                plt.xlabel("Hmotnost [{}]".format(self.massunit))
                nazev_x = "_xHMOTNOST"
            elif (self.key_x == "frequency"):
                plt.xlabel("Frekvence [{}]".format(self.frequencyunit))
                nazev_x = "_xFREKVENCE"
            elif (self.key_x == "x"):
                plt.xlabel("Pozice X [{}]".format(self.positionunit))
                nazev_x = "_xX"
            elif (self.key_x == "y"):
                plt.xlabel("Pozice Y [{}]".format(self.positionunit))
                nazev_x = "_xY"
            elif (self.key_x == "z"):
                plt.xlabel("Pozice Z [{}]".format(self.positionunit))
                nazev_x = "_xZ"
            elif (self.key_x == "vx"):
                plt.xlabel("Rychlost X [{}]".format(self.velocityunit))
                nazev_x = "_xRychlost_X"
            elif (self.key_x == "vy"):
                plt.xlabel("Rychlost Y [{}]".format(self.velocityunit))
                nazev_x = "_xRychlost_Y"
            elif (self.key_x == "vz"):
                plt.xlabel("Rychlost Z [{}]".format(self.velocityunit))
                nazev_x = "_xRychlost_Z"
            elif (self.key_x == "vmag"):
                plt.xlabel("Rychlost (výslednice) [{}]".format(self.velocityunit))
                nazev_x = "_xRYCHLOST_MAG"
            else:
                pass
        
        elif self.language == "EN":
            if (self.key_x == "time"):
                plt.xlabel("Time [{}]".format(self.timeunit))
                nazev_x = "_xTIME"
            elif (self.key_x == "diameter"):
                plt.xlabel("Diameter [{}]".format(self.diameterunit))
                nazev_x = "_xPRUMER"
            elif (self.key_x == "temperature"):
                plt.xlabel("Temperature [{}]".format(self.temperatureunit))
                nazev_x = "_xTEPLOTA"
            elif (self.key_x == "position"):
                plt.xlabel("Position [{}]".format(self.positionunit))
                nazev_x = "_xPOZICE"
            elif (self.key_x == "velocity"):
                plt.xlabel("Velocity [{}]".format(self.velocityunit))
                nazev_x = "_xRYCHLOST"
            elif (self.key_x == "mass-flow-rate"):
                plt.xlabel("Mass Flow Rate [{}]".format(self.mfrunit))
                nazev_x = "_xMFR"
            elif (self.key_x == "mass"):
                plt.xlabel("Hmotnost [{}]".format(self.massunit))
                nazev_x = "_xHMOTNOST"
            elif (self.key_x == "frequency"):
                plt.xlabel("Emission Frequency [{}]".format(self.frequencyunit))
                nazev_x = "_xFREKVENCE"
            elif (self.key_x == "x"):
                plt.xlabel("X Position [{}]".format(self.positionunit))
                nazev_x = "_xX"
            elif (self.key_x == "y"):
                plt.xlabel("Y Position [{}]".format(self.positionunit))
                nazev_x = "_xY"
            elif (self.key_x == "z"):
                plt.xlabel("Z Position [{}]".format(self.positionunit))
                nazev_x = "_xZ"
            elif (self.key_x == "vx"):
                plt.xlabel("X Velocity [{}]".format(self.velocityunit))
                nazev_x = "_xRychlost_X"
            elif (self.key_x == "vy"):
                plt.xlabel("Y Velocity [{}]".format(self.velocityunit))
                nazev_x = "_xRychlost_Y"
            elif (self.key_x == "vz"):
                plt.xlabel("Z Velocity [{}]".format(self.velocityunit))
                nazev_x = "_xRychlost_Z"
            elif (self.key_x == "vmag"):
                plt.xlabel("Velocity magnitude [{}]".format(self.velocityunit))
                nazev_x = "_xRYCHLOST_MAG"
            else:
                pass        
        #endregion

        #region Y-label
        if self.language == "CZ":
            if (self.key_y == "time"):
                plt.ylabel("Čas [{}]".format(self.timeunit))
                nazev_y = "_yTIME"
            elif (self.key_y == "diameter"):
                plt.ylabel("Průměr [{}]".format(self.diameterunit))
                nazev_y = "_yPRUMER"
            elif (self.key_y == "temperature"):
                plt.ylabel("Teplota [{}]".format(self.temperatureunit))
                nazev_y = "_yTEPLOTA"
            elif (self.key_y == "position"):
                plt.ylabel("Pozice [{}]".format(self.positionunit))
                nazev_y = "_yPOZICE"
            elif (self.key_y == "velocity"):
                plt.ylabel("Rychlost [{}]".format(self.velocityunit))
                nazev_y = "_yRYCHLOST"
            elif (self.key_y == "mass-flow-rate"):
                plt.ylabel("Hmotnostní průtok [{}]".format(self.mfrunit))
                nazev_y = "_yMFR"
            elif (self.key_y == "mass"):
                plt.ylabel("Hmotnost [{}]".format(self.massunit))
                nazev_y = "_yHMOTNOST"
            elif (self.key_y == "frequency"):
                plt.ylabel("Frekvence [{}]".format(self.frequencyunit))
                nazev_y = "_yFREKVENCE"
            elif (self.key_y == "x"):
                plt.ylabel("Pozice X [{}]".format(self.positionunit))
                nazev_y = "_yX"
            elif (self.key_y == "y"):
                plt.ylabel("Pozice Y [{}]".format(self.positionunit))
                nazev_y = "_yY"
            elif (self.key_y == "z"):
                plt.ylabel("Pozice Z [{}]".format(self.positionunit))
                nazev_y = "_yZ"
            elif (self.key_y == "vx"):
                plt.ylabel("Rychlost X [{}]".format(self.velocityunit))
                nazev_y = "_yRychlost_X"
            elif (self.key_y == "vy"):
                plt.ylabel("Rychlost Y [{}]".format(self.velocityunit))
                nazev_y = "_yRychlost_Y"
            elif (self.key_y == "vz"):
                plt.ylabel("Rychlost Z [{}]".format(self.velocityunit))
                nazev_y = "_yRychlost_Z"
            elif (self.key_y == "vmag"):
                plt.ylabel("Rychlost (výslednice) [{}]".format(self.velocityunit))
                nazev_y = "_yRYCHLOST_MAG"
            else:
                pass
        
        elif self.language == "EN":
            if (self.key_y == "time"):
                plt.ylabel("Time [{}]".format(self.timeunit))
                nazev_y = "_yTIME"
            elif (self.key_y == "diameter"):
                plt.ylabel("Diameter [{}]".format(self.diameterunit))
                nazev_y = "_yPRUMER"
            elif (self.key_y == "temperature"):
                plt.ylabel("Temperature [{}]".format(self.temperatureunit))
                nazev_y = "_yTEPLOTA"
            elif (self.key_y == "position"):
                plt.ylabel("Position [{}]".format(self.positionunit))
                nazev_y = "_yPOZICE"
            elif (self.key_y == "velocity"):
                plt.ylabel("Velocity [{}]".format(self.velocityunit))
                nazev_y = "_yRYCHLOST"
            elif (self.key_y == "mass-flow-rate"):
                plt.ylabel("Mass Flow Rate [{}]".format(self.mfrunit))
                nazev_y = "_yMFR"
            elif (self.key_y == "mass"):
                plt.ylabel("Mass [{}]".format(self.massunit))
                nazev_y = "_yHMOTNOST"
            elif (self.key_y == "frequency"):
                plt.ylabel("Emission Frequency [{}]".format(self.frequencyunit))
                nazev_y = "_yFREKVENCE"
            elif (self.key_y == "x"):
                plt.ylabel("X Position [{}]".format(self.positionunit))
                nazev_y = "_yX"
            elif (self.key_y == "y"):
                plt.ylabel("Y Position [{}]".format(self.positionunit))
                nazev_y = "_yY"
            elif (self.key_y == "z"):
                plt.ylabel("Z Position [{}]".format(self.positionunit))
                nazev_y = "_yZ"
            elif (self.key_y == "vx"):
                plt.ylabel("X Velocity [{}]".format(self.velocityunit))
                nazev_y = "_yRychlost_X"
            elif (self.key_y == "vy"):
                plt.ylabel("Y Velocity [{}]".format(self.velocityunit))
                nazev_y = "_yRychlost_Y"
            elif (self.key_y == "vz"):
                plt.ylabel("Z Velocity [{}]".format(self.velocityunit))
                nazev_y = "_yRychlost_Z"
            elif (self.key_y == "vmag"):
                plt.ylabel("Velocity Magnitude [{}]".format(self.velocityunit))
                nazev_y = "_yRYCHLOST_MAG"
            else:
                pass
        #endregion
        
        #region Color-label
        if len(args) != 0:
            if self.language == "CZ":
                if (self.key_color == "time"):
                    nazev_color = "_cTIME"
                    colorbar.set_label("Čas [{}]".format(self.timeunit))
                elif (self.key_color == "diameter"):
                    nazev_color = "_cPRUMER"
                    colorbar.set_label("Průměr [{}]".format(self.diameterunit))
                elif (self.key_color == "temperature"):
                    nazev_color = "_cTEPLOTA"
                    colorbar.set_label("Teplota [{}]".format(self.temperatureunit))
                elif (self.key_color == "position"):
                    nazev_color = "_cPOZICE"
                    colorbar.set_label("Pozice [{}]".format(self.positionunit))
                elif (self.key_color == "velocity"):
                    nazev_color = "_cRYCHLOST"
                    colorbar.set_label("Rychlost [{}]".format(self.velocityunit))
                elif (self.key_color == "mass-flow-rate"):
                    nazev_color = "_cMFR"
                    colorbar.set_label("Hm. průtok [{}]".format(self.mfrunit))
                elif (self.key_color == "mass"):
                    nazev_color = "_cHMOTNOST"
                    colorbar.set_label("Hmotnost [{}]".format(self.massunit))
                elif (self.key_color == "frequency"):
                    nazev_color = "_cFREKVENCE"
                    colorbar.set_label("Frekvence emisí [{}]".format(self.frequencyunit))
                elif (self.key_color == "x"):
                    nazev_color = "_cX"
                    colorbar.set_label("Pozice X [{}]".format(self.positionunit))
                elif (self.key_color == "y"):
                    nazev_color = "_cY"
                    colorbar.set_label("Pozice Y [{}]".format(self.positionunit))
                elif (self.key_color == "z"):
                    nazev_color = "_cZ"
                    colorbar.set_label("Pozice Z [{}]".format(self.positionunit))
                elif (self.key_color == "vx"):
                    nazev_color = "_cRychlost_X"
                    colorbar.set_label("Rychlost X [{}]".format(self.velocityunit))
                elif (self.key_color == "vy"):
                    nazev_color = "_cRychlost_Y"
                    colorbar.set_label("Rychlost Y [{}]".format(self.velocityunit))
                elif (self.key_color == "vz"):
                    nazev_color = "_cRychlost_Z"
                    colorbar.set_label("Rychlost Z [{}]".format(self.velocityunit))
                elif (self.key_color == "vmag"):
                    nazev_color = "_cRYCHLOST_MAG"
                    colorbar.set_label("Rychlost (výslednice) [{}]".format(self.velocityunit))
                else:
                    pass

            elif self.language == "EN":
                if (self.key_color == "time"):
                    nazev_color = "_cTIME"
                    colorbar.set_label("Time [{}]".format(self.timeunit))
                elif (self.key_color == "diameter"):
                    nazev_color = "_cPRUMER"
                    colorbar.set_label("Diameter [{}]".format(self.diameterunit))
                elif (self.key_color == "temperature"):
                    nazev_color = "_cTEPLOTA"
                    colorbar.set_label("Temperature [{}]".format(self.temperatureunit))
                elif (self.key_color == "position"):
                    nazev_color = "_cPOZICE"
                    colorbar.set_label("Position [{}]".format(self.positionunit))
                elif (self.key_color == "velocity"):
                    nazev_color = "_cRYCHLOST"
                    colorbar.set_label("Velocity [{}]".format(self.velocityunit))
                elif (self.key_color == "mass-flow-rate"):
                    nazev_color = "_cMFR"
                    colorbar.set_label("Mass flow rate [{}]".format(self.mfrunit))
                elif (self.key_color == "mass"):
                    nazev_color = "_cHMOTNOST"
                    colorbar.set_label("Mass [{}]".format(self.massunit))
                elif (self.key_color == "frequency"):
                    nazev_color = "_cFREKVENCE"
                    colorbar.set_label("Emission Frequency [{}]".format(self.frequencyunit))
                elif (self.key_color == "x"):
                    nazev_color = "_cX"
                    colorbar.set_label("X Position [{}]".format(self.positionunit))
                elif (self.key_color == "y"):
                    nazev_color = "_cY"
                    colorbar.set_label("Y Position [{}]".format(self.positionunit))
                elif (self.key_color == "z"):
                    nazev_color = "_cZ"
                    colorbar.set_label("Z Position [{}]".format(self.positionunit))
                elif (self.key_color == "vx"):
                    nazev_color = "_cRychlost_X"
                    colorbar.set_label("X Velocity [{}]".format(self.velocityunit))
                elif (self.key_color == "vy"):
                    nazev_color = "_cRychlost_Y"
                    colorbar.set_label("Y Velocity [{}]".format(self.velocityunit))
                elif (self.key_color == "vz"):
                    nazev_color = "_cRychlost_Z"
                    colorbar.set_label("Z Velocity [{}]".format(self.velocityunit))
                elif (self.key_color == "vmag"):
                    nazev_color = "_cRYCHLOST_MAG"
                    colorbar.set_label("Velocity Magnitude [{}]".format(self.velocityunit))
                else:
                    pass
        #endregion Color-label

        namefile_p2 = datetime.now().strftime("%d-%m-%Y-%H-%M-%S")
        if len(args) == 0:
            nazev_tot = "SCATTER_" + nazev_x[2:].lower() + "_vs_" + nazev_y[2:].lower() +"__{}".format(namefile_p2) + "__img.png"
        elif len(args) == 1:
            nazev_tot = "SCATTER_" + nazev_x[2:].lower() + "_vs_" + nazev_y[2:].lower() + "_vs_" + nazev_color[2:].lower() + "__{}".format(namefile_p2) + "__img.png"
        else:
            print("DPM SCATTER: Něco se pohnojilo, máš více argumentů než je povoleno")



        ax = plt.gca()
        ax.patch.set_facecolor('xkcd:pale blue')
        ax.patch.set_alpha(0.5)
        plt.title(self.return_name(self.key_y, self.key_x))

        if self.scatter_xlim_bool == True:
            print("DPM SCATTER: XMIN:    {}; XMAX:   {}".format(self.xmin_scatter, self.xmax_scatter))
            plt.xlim(self.xmin_scatter, self.xmax_scatter)

        if self.scatter_ylim_bool == True:
            if len(args) > 0:
                plt.clim(self.ymin_scatter, self.ymax_scatter)
        plt.savefig(self.workdir +"/"+ nazev_tot, dpi = self.val_dpi, bbox_inches = "tight")
        plt.clf()
        plt.close()
        print(self.workdir +"/"+ nazev_tot)

    def return_name(self, key_x, key_y):

        if self.language == "CZ":
            word_1 = "Závislost "
            word_2 = ""
            word_3 = " na "
            word_4 = ""
            #region vel_1
            if key_x == "x":
                word_2 = "pozice X "

            elif key_x == "y":
                word_2 = "pozice Y "

            elif key_x == "z":
                word_2 = "pozice Z "

            elif key_x == "vx":
                word_2 = "rychlosti X"

            elif key_x == "vy":
                word_2 = "rychlosti Y"

            elif key_x == "vz":
                word_2 = "rychlosti Z"

            elif key_x == "vmag":
                word_2 = "rychlosti "

            elif key_x == "diameter":
                word_2 = "průměru "

            elif key_x == "temperature":
                word_2 = "teploty "

            elif key_x == "mass-flow-rate":
                word_2 = "hm. průtoku "

            elif key_x == "mass":
                word_2 = "hmotnosti "

            elif key_x == "time":
                word_2 = "času "
            #endregion vel_1

            #region vel_2
            if key_y == "x":
                word_4 = "pozici X "

            elif key_y == "y":
                word_4 = "pozici Y "

            elif key_y == "z":
                word_4 = "pozici Z "

            elif key_y == "vx":
                word_4 = "rychlosti X"

            elif key_y == "vy":
                word_4 = "rychlosti Y"

            elif key_y == "vz":
                word_4 = "rychlosti Z"

            elif key_y == "vmag":
                word_4 = "rychlosti "

            elif key_y == "diameter":
                word_4 = "průměru "

            elif key_y == "temperature":
                word_4 = "teplotě "

            elif key_y == "mass-flow-rate":
                word_4 = "hm. průtoku "

            elif key_y == "mass":
                word_4 = "hmotnosti "

            elif key_y == "time":
                word_4 = "času "
            #endregion vel_2
            label = word_1 + word_2 + word_3 + word_4
            print(label)
            del word_1, word_2, word_3, word_4
            return str(label)

        elif self.language == "EN":
            word_1 = "Dependency of "
            word_2 = ""
            word_3 = " on "
            word_4 = ""
            #region vel_1
            if key_x == "x":
                word_2 = "X position "

            elif key_x == "y":
                word_2 = "Y position "

            elif key_x == "z":
                word_2 = "Z position "

            elif key_x == "vx":
                word_2 = "X velocity"

            elif key_x == "vy":
                word_2 = "Y velocity"

            elif key_x == "vz":
                word_2 = "Z velocity"

            elif key_x == "vmag":
                word_2 = "Velocity "

            elif key_x == "diameter":
                word_2 = "Diameter"

            elif key_x == "temperature":
                word_2 = "Temperature "

            elif key_x == "mass-flow-rate":
                word_2 = "Mass Flow Rate "

            elif key_x == "mass":
                word_2 = "Mass "

            elif key_x == "time":
                word_2 = "Time "
            #endregion vel_1

            #region vel_2
            if key_y == "x":
                word_4 = "X position "

            elif key_y == "y":
                word_4 = "Y position "

            elif key_y == "z":
                word_4 = "Z position "

            elif key_y == "vx":
                word_4 = "X velocity"

            elif key_y == "vy":
                word_4 = "Y velocity"

            elif key_y == "vz":
                word_4 = "Z velocity"

            elif key_y == "vmag":
                word_4 = "Velocity "

            elif key_y == "diameter":
                word_4 = "Diameter"

            elif key_y == "temperature":
                word_4 = "Temperature "

            elif key_y == "mass-flow-rate":
                word_4 = "Mass Flow Rate "

            elif key_y == "mass":
                word_4 = "Mass "

            elif key_y == "time":
                word_4 = "Time "
            #endregion vel_2
            label = word_1 + word_2 + word_3 + word_4
            print(label)
            del word_1, word_2, word_3, word_4
            return str(label)

    def table_spawn_csv(self):
        print("TABLE_SPAWN_CSV... begin")
        self.csv_x           = pd.DataFrame(data = self.x)
        self.csv_y           = pd.DataFrame(data = self.y)
        self.csv_z           = pd.DataFrame(data = self.z)
        self.csv_vx          = pd.DataFrame(data = self.vx)
        self.csv_vy          = pd.DataFrame(data = self.vy)
        self.csv_vz          = pd.DataFrame(data = self.vz)
        self.csv_vmag        = pd.DataFrame(data = self.vmag)
        self.csv_diameter    = pd.DataFrame(data = self.diameter)
        self.csv_temperature = pd.DataFrame(data = self.temperature)
        self.csv_mfr         = pd.DataFrame(data = self.mfr)
        self.csv_mass        = pd.DataFrame(data = self.mass)
        self.csv_time        = pd.DataFrame(data = self.time)
        self.csv_summary     = pd.DataFrame()

        self.csv_summary.insert(len(self.csv_summary.columns), "ID částice [-]", self.csv_x[self.csv_x.columns[0]])
        self.csv_summary.insert(len(self.csv_summary.columns), "Pozice x [{}]".format(self.csv_x["unit"].loc[0]), self.csv_x[self.csv_x.columns[1]])
        self.csv_summary.insert(len(self.csv_summary.columns), "Pozice y [{}]".format(self.csv_z["unit"].loc[0]), self.csv_y[self.csv_y.columns[1]])
        self.csv_summary.insert(len(self.csv_summary.columns), "Pozice z [{}]".format(self.csv_z["unit"].loc[0]), self.csv_z[self.csv_z.columns[1]])
        self.csv_summary.insert(len(self.csv_summary.columns), "Rychlost x [{}]".format(self.csv_vx["unit"].loc[0]), self.csv_vx[self.csv_vx.columns[1]])
        self.csv_summary.insert(len(self.csv_summary.columns), "Rychlost y [{}]".format(self.csv_vy["unit"].loc[0]), self.csv_vy[self.csv_vy.columns[1]])
        self.csv_summary.insert(len(self.csv_summary.columns), "Rychlost z [{}]".format(self.csv_vz["unit"].loc[0]), self.csv_vz[self.csv_vz.columns[1]])
        self.csv_summary.insert(len(self.csv_summary.columns), "Rychlost [{}]".format(self.csv_vmag["unit"].loc[0]), self.csv_vmag[self.csv_vmag.columns[1]])
        self.csv_summary.insert(len(self.csv_summary.columns), "Průměr [{}]".format(self.csv_diameter["unit"].loc[0]), self.csv_diameter[self.csv_diameter.columns[1]])
        self.csv_summary.insert(len(self.csv_summary.columns), "Teplota [{}]".format(self.csv_temperature["unit"].loc[0]), self.csv_temperature[self.csv_temperature.columns[1]])
        self.csv_summary.insert(len(self.csv_summary.columns), "Hmotnostní průtok [{}]".format(self.csv_mfr["unit"].loc[0]), self.csv_mfr[self.csv_mfr.columns[1]])       
        self.csv_summary.insert(len(self.csv_summary.columns), "Hmotnost [{}]".format(self.csv_mass["unit"].loc[0]), self.csv_mass[self.csv_mass.columns[1]])
        self.csv_summary.insert(len(self.csv_summary.columns), "Čas letu [{}]".format(self.csv_time["unit"].loc[0]), self.csv_time[self.csv_time.columns[1]])
        print(self.csv_summary.head(3))

        self.csv_statistics   = pd.DataFrame()
        self.csv_statistics["Veličina"] = ""
        self.csv_statistics["Minimum"] = ""
        self.csv_statistics["Maximum"] = ""
        self.csv_statistics["Ar. průměr"] = ""
        self.csv_statistics["Medián"] = ""

       #region redundant
        """        
        self.csv_statistics["Medián uspořádaný"] = ""
        self.csv_statistics = self.csv_statistics.append({"Veličina": "Pozice x [{}]".format(self.csv_x["unit"].loc[0]), "Minimum":  self.csv_x["x"].min(), "Maximum": self.csv_x["x"].max(), "Ar. průměr":  self.csv_x["x"].mean(), "Medián":  self.csv_x["x"].median(), "Medián uspořádaný": self.csv_x.sort_values(by = ["x"], ascending = True)["x"].median()}, ignore_index=True)
        self.csv_statistics = self.csv_statistics.append({"Veličina": "Pozice y [{}]".format(self.csv_y["unit"].loc[0]), "Minimum":  self.csv_y["y"].min(), "Maximum": self.csv_y["y"].max(), "Ar. průměr":  self.csv_y["y"].mean(), "Medián":  self.csv_y["y"].median(), "Medián uspořádaný": self.csv_y.sort_values(by = ["y"], ascending = True)["y"].median()}, ignore_index=True)
        self.csv_statistics = self.csv_statistics.append({"Veličina": "Pozice z [{}]".format(self.csv_z["unit"].loc[0]), "Minimum":  self.csv_z["z"].min(), "Maximum": self.csv_z["z"].max(), "Ar. průměr":  self.csv_z["z"].mean(), "Medián":  self.csv_z["z"].median(), "Medián uspořádaný": self.csv_z.sort_values(by = ["z"], ascending = True)["z"].median()}, ignore_index=True)
        self.csv_statistics = self.csv_statistics.append({"Veličina": "Rychlost x [{}]".format(self.csv_vx["unit"].loc[0]), "Minimum":  self.csv_vx["vx"].min(), "Maximum": self.csv_vx["vx"].max(), "Ar. průměr":  self.csv_vx["vx"].mean(), "Medián":  self.csv_vx["vx"].median(), "Medián uspořádaný": self.csv_vx.sort_values(by = ["vx"], ascending = True)["vx"].median()}, ignore_index=True)
        self.csv_statistics = self.csv_statistics.append({"Veličina": "Rychlost y [{}]".format(self.csv_vy["unit"].loc[0]), "Minimum":  self.csv_vy["vy"].min(), "Maximum": self.csv_vy["vy"].max(), "Ar. průměr":  self.csv_vy["vy"].mean(), "Medián":  self.csv_vy["vy"].median(), "Medián uspořádaný": self.csv_vy.sort_values(by = ["vy"], ascending = True)["vy"].median()}, ignore_index=True)
        self.csv_statistics = self.csv_statistics.append({"Veličina": "Rychlost z [{}]".format(self.csv_vz["unit"].loc[0]), "Minimum":  self.csv_vz["vz"].min(), "Maximum": self.csv_vz["vz"].max(), "Ar. průměr":  self.csv_vz["vz"].mean(), "Medián":  self.csv_vz["vz"].median(), "Medián uspořádaný": self.csv_vz.sort_values(by = ["vz"], ascending = True)["vz"].median()}, ignore_index=True)
        self.csv_statistics = self.csv_statistics.append({"Veličina": "Rychlost [{}]".format(self.csv_vmag["unit"].loc[0]), "Minimum":  self.csv_vmag["vmag"].min(), "Maximum": self.csv_vmag["vmag"].max(), "Ar. průměr":  self.csv_vmag["vmag"].mean(), "Medián":  self.csv_vmag["vmag"].median(), "Medián uspořádaný": self.csv_vmag.sort_values(by = ["vmag"], ascending = True)["vmag"].median()}, ignore_index=True)
        self.csv_statistics = self.csv_statistics.append({"Veličina": "Průměr [{}]".format(self.csv_diameter["unit"].loc[0]), "Minimum":  self.csv_diameter["diameter"].min(), "Maximum": self.csv_diameter["diameter"].max(), "Ar. průměr":  self.csv_diameter["diameter"].mean(), "Medián":  self.csv_diameter["diameter"].median(), "Medián uspořádaný": self.csv_diameter.sort_values(by = ["diameter"], ascending = True)["diameter"].median()}, ignore_index=True)
        self.csv_statistics = self.csv_statistics.append({"Veličina": "Teplota [{}]".format(self.csv_temperature["unit"].loc[0]), "Minimum":  self.csv_temperature["temperature"].min(), "Maximum": self.csv_temperature["temperature"].max(), "Ar. průměr":  self.csv_temperature["temperature"].mean(), "Medián":  self.csv_temperature["temperature"].median(), "Medián uspořádaný": self.csv_temperature.sort_values(by = ["temperature"], ascending = True)["temperature"].median()}, ignore_index=True)
        self.csv_statistics = self.csv_statistics.append({"Veličina": "Hmotnostní průtok [{}]".format(self.csv_mfr["unit"].loc[0]), "Minimum":  self.csv_mfr["mass-flow-rate"].min(), "Maximum": self.csv_mfr["mass-flow-rate"].max(), "Ar. průměr":  self.csv_mfr["mass-flow-rate"].mean(), "Medián":  self.csv_mfr["mass-flow-rate"].median(), "Medián uspořádaný": self.csv_mfr.sort_values(by = ["mass-flow-rate"], ascending = True)["mass-flow-rate"].median()}, ignore_index=True)
        self.csv_statistics = self.csv_statistics.append({"Veličina": "Hmotnost [{}]".format(self.csv_mass["unit"].loc[0]), "Minimum":  self.csv_mass["mass"].min(), "Maximum": self.csv_mass["mass"].max(), "Ar. průměr":  self.csv_mass["mass"].mean(), "Medián":  self.csv_mass["mass"].median(), "Medián uspořádaný": self.csv_mass.sort_values(by = ["mass"], ascending = True)["mass"].median()}, ignore_index=True)
        self.csv_statistics = self.csv_statistics.append({"Veličina": "Čas letu [{}]".format(self.csv_time["unit"].loc[0]), "Minimum":  self.csv_time["time"].min(), "Maximum": self.csv_time["time"].max(), "Ar. průměr":  self.csv_time["time"].mean(), "Medián":  self.csv_time["time"].median(), "Medián uspořádaný": self.csv_time.sort_values(by = ["time"], ascending = True)["time"].median()}, ignore_index=True)
        print(self.csv_statistics)
        """
        #endregion redundant

        self.csv_statistics = self.csv_statistics.append({"Veličina": "Pozice x [{}]".format(self.csv_x["unit"].loc[0]), "Minimum":  self.csv_x["x"].min(), "Maximum": self.csv_x["x"].max(), "Ar. průměr":  self.csv_x["x"].mean(), "Medián":  self.csv_x["x"].median()}, ignore_index=True)
        self.csv_statistics = self.csv_statistics.append({"Veličina": "Pozice y [{}]".format(self.csv_y["unit"].loc[0]), "Minimum":  self.csv_y["y"].min(), "Maximum": self.csv_y["y"].max(), "Ar. průměr":  self.csv_y["y"].mean(), "Medián":  self.csv_y["y"].median()}, ignore_index=True)
        self.csv_statistics = self.csv_statistics.append({"Veličina": "Pozice z [{}]".format(self.csv_z["unit"].loc[0]), "Minimum":  self.csv_z["z"].min(), "Maximum": self.csv_z["z"].max(), "Ar. průměr":  self.csv_z["z"].mean(), "Medián":  self.csv_z["z"].median()}, ignore_index=True)
        self.csv_statistics = self.csv_statistics.append({"Veličina": "Rychlost x [{}]".format(self.csv_vx["unit"].loc[0]), "Minimum":  self.csv_vx["vx"].min(), "Maximum": self.csv_vx["vx"].max(), "Ar. průměr":  self.csv_vx["vx"].mean(), "Medián":  self.csv_vx["vx"].median()}, ignore_index=True)
        self.csv_statistics = self.csv_statistics.append({"Veličina": "Rychlost y [{}]".format(self.csv_vy["unit"].loc[0]), "Minimum":  self.csv_vy["vy"].min(), "Maximum": self.csv_vy["vy"].max(), "Ar. průměr":  self.csv_vy["vy"].mean(), "Medián":  self.csv_vy["vy"].median()}, ignore_index=True)
        self.csv_statistics = self.csv_statistics.append({"Veličina": "Rychlost z [{}]".format(self.csv_vz["unit"].loc[0]), "Minimum":  self.csv_vz["vz"].min(), "Maximum": self.csv_vz["vz"].max(), "Ar. průměr":  self.csv_vz["vz"].mean(), "Medián":  self.csv_vz["vz"].median()}, ignore_index=True)
        self.csv_statistics = self.csv_statistics.append({"Veličina": "Rychlost [{}]".format(self.csv_vmag["unit"].loc[0]), "Minimum":  self.csv_vmag["vmag"].min(), "Maximum": self.csv_vmag["vmag"].max(), "Ar. průměr":  self.csv_vmag["vmag"].mean(), "Medián":  self.csv_vmag["vmag"].median()}, ignore_index=True)
        self.csv_statistics = self.csv_statistics.append({"Veličina": "Průměr [{}]".format(self.csv_diameter["unit"].loc[0]), "Minimum":  self.csv_diameter["diameter"].min(), "Maximum": self.csv_diameter["diameter"].max(), "Ar. průměr":  self.csv_diameter["diameter"].mean(), "Medián":  self.csv_diameter["diameter"].median()}, ignore_index=True)
        self.csv_statistics = self.csv_statistics.append({"Veličina": "Teplota [{}]".format(self.csv_temperature["unit"].loc[0]), "Minimum":  self.csv_temperature["temperature"].min(), "Maximum": self.csv_temperature["temperature"].max(), "Ar. průměr":  self.csv_temperature["temperature"].mean(), "Medián":  self.csv_temperature["temperature"].median()}, ignore_index=True)
        self.csv_statistics = self.csv_statistics.append({"Veličina": "Hmotnostní průtok [{}]".format(self.csv_mfr["unit"].loc[0]), "Minimum":  self.csv_mfr["mass-flow-rate"].min(), "Maximum": self.csv_mfr["mass-flow-rate"].max(), "Ar. průměr":  self.csv_mfr["mass-flow-rate"].mean(), "Medián":  self.csv_mfr["mass-flow-rate"].median()}, ignore_index=True)
        self.csv_statistics = self.csv_statistics.append({"Veličina": "Hmotnost [{}]".format(self.csv_mass["unit"].loc[0]), "Minimum":  self.csv_mass["mass"].min(), "Maximum": self.csv_mass["mass"].max(), "Ar. průměr":  self.csv_mass["mass"].mean(), "Medián":  self.csv_mass["mass"].median()}, ignore_index=True)
        self.csv_statistics = self.csv_statistics.append({"Veličina": "Čas letu [{}]".format(self.csv_time["unit"].loc[0]), "Minimum":  self.csv_time["time"].min(), "Maximum": self.csv_time["time"].max(), "Ar. průměr":  self.csv_time["time"].mean(), "Medián":  self.csv_time["time"].median()}, ignore_index=True)
       
       
        if self.language == "CZ":
            self.csv_summary.columns = ["ID částice [-]", "Pozice x [{}]".format(self.csv_x["unit"].loc[0]), "Pozice Y [{}]".format(self.csv_y["unit"].loc[0]), "Pozice Z [{}]".format(self.csv_z["unit"].loc[0]), "Rychlost X [{}]".format(self.csv_vx["unit"].loc[0]), "Rychlost Y [{}]".format(self.csv_vy["unit"].loc[0]), "Rychlost Z [{}]".format(self.csv_vz["unit"].loc[0]), "Rychlost [{}]".format(self.csv_vmag["unit"].loc[0]), "Průměr [{}]".format(self.csv_diameter["unit"].loc[0]), "Teplota [{}]".format(self.csv_temperature["unit"].loc[0]), "Hmotnostní průtok [{}]".format(self.csv_mfr["unit"].loc[0]), "Hmotnost [{}]".format(self.csv_mass["unit"].loc[0]), "Čas letu [{}]".format(self.csv_time["unit"].loc[0])]
            self.csv_statistics.columns = ["Veličina", "Minimum", "Maximum", "Ar. průměr", "Medián"]
        elif self.language == "EN":
            self.csv_summary.columns = ["Name [-]", "X Position [{}]".format(self.csv_x["unit"].loc[0]), "Y Position [{}]".format(self.csv_y["unit"].loc[0]), "Z Position [{}]".format(self.csv_z["unit"].loc[0]), "X Velocity [{}]".format(self.csv_vx["unit"].loc[0]), "Y Velocity [{}]".format(self.csv_vy["unit"].loc[0]), "Z Velocity [{}]".format(self.csv_vz["unit"].loc[0]), "Velocity [{}]".format(self.csv_vmag["unit"].loc[0]), "Diameter [{}]".format(self.csv_diameter["unit"].loc[0]), "Temperature [{}]".format(self.csv_temperature["unit"].loc[0]), "Mass Flow Rate [{}]".format(self.csv_mfr["unit"].loc[0]), "Mass [{}]".format(self.csv_mass["unit"].loc[0]), "Time of Flight [{}]".format(self.csv_time["unit"].loc[0])]
            self.csv_statistics.columns = ["Variable", "Minimum", "Maximum", "Ar. Mean", "Median"]
        
        print(self.csv_statistics)
        workdir = self.workdir
        if self.language == "CZ":
            self.excel_file = Workbook()
            self.excel_file.worksheets[0].title = "souhrn"
            self.excel_file.create_sheet("statistika")
            self.excel_file.save(workdir + "/{}_table_{}.xlsx".format(self.name, self.language))
            self.excel_file.close()

            with pd.ExcelWriter(workdir + "/{}_table_{}.xlsx".format(self.name, self.language), engine="openpyxl") as writer:
                    self.csv_summary.to_excel(writer, index = False, sheet_name= "souhrn")
                    self.csv_statistics.to_excel(writer, index = False, sheet_name= "statistika")
                    self.columns_best_fit(self.excel_file.worksheets[0])
                    self.columns_best_fit(self.excel_file.worksheets[1])

        elif self.language == "EN":
            self.excel_file = Workbook()
            self.excel_file.worksheets[0].title = "summary"
            self.excel_file.create_sheet("statistics")
            self.excel_file.save(workdir + "/{}_table_{}.xlsx".format(self.name, self.language))
            self.excel_file.close()  

            self.csv_summary.columns = ["Name [-]", "X Position [{}]".format(self.csv_x["unit"].loc[0]), "Y Position [{}]".format(self.csv_z["unit"].loc[0]), "Z Position [{}]".format(self.csv_z["unit"].loc[0]), "X Velocity [{}]".format(self.csv_vx["unit"].loc[0]), "Y Velocity [{}]".format(self.csv_vy["unit"].loc[0]), "Z Velocity [{}]".format(self.csv_vz["unit"].loc[0]), "Velocity [{}]".format(self.csv_vmag["unit"].loc[0]), "Diameter [{}]".format(self.csv_diameter["unit"].loc[0]), "Temperature [{}]".format(self.csv_temperature["unit"].loc[0]), "Mass Flow Rate [{}]".format(self.csv_mfr["unit"].loc[0]), "Mass [{}]".format(self.csv_mass["unit"].loc[0]), "Time of Flight [{}]".format(self.csv_time["unit"].loc[0])]
            self.csv_statistics["Variable"] = ["X Position [{}]".format(self.csv_x["unit"].loc[0]), "Y Position [{}]".format(self.csv_z["unit"].loc[0]), "Z Position [{}]".format(self.csv_z["unit"].loc[0]), "X Velocity [{}]".format(self.csv_vx["unit"].loc[0]), "Y Velocity [{}]".format(self.csv_vy["unit"].loc[0]), "Z Velocity [{}]".format(self.csv_vz["unit"].loc[0]), "Velocity [{}]".format(self.csv_vmag["unit"].loc[0]), "Diameter [{}]".format(self.csv_diameter["unit"].loc[0]), "Temperature [{}]".format(self.csv_temperature["unit"].loc[0]), "Mass Flow Rate [{}]".format(self.csv_mfr["unit"].loc[0]), "Mass [{}]".format(self.csv_mass["unit"].loc[0]), "Time of Flight [{}]".format(self.csv_time["unit"].loc[0])]

            with pd.ExcelWriter(workdir + "/{}_table_{}.xlsx".format(self.name, self.language), engine="openpyxl") as writer:
                    self.csv_summary.to_excel(writer, index = False, sheet_name= "summary")
                    self.csv_statistics.to_excel(writer, index = False, sheet_name= "statistics")
                    self.columns_best_fit(self.excel_file.worksheets[0])
                    self.columns_best_fit(self.excel_file.worksheets[1])                  
                #self.csv_summary.to_excel(workdir + "/{}_table.xlsx".format(self.name), index = False, sheet_name= "souhrn")
                #self.csv_statistics.to_excel(workdir + "/{}_table.xlsx".format(self.name), index = False, sheet_name= "statistika")
        
        if self.language == "CZ":
            self.csv_summary.columns = ["ID castice [-]", "Pozice X [{}]".format(self.csv_x["unit"].loc[0]), "Pozice Y [{}]".format(self.csv_z["unit"].loc[0]), "Pozice Z [{}]".format(self.csv_z["unit"].loc[0]), "Rychlost X [{}]".format(self.csv_vx["unit"].loc[0]), "Rychlost Y [{}]".format(self.csv_vy["unit"].loc[0]), "Rychlost Z [{}]".format(self.csv_vz["unit"].loc[0]), "Rychlost".format(self.csv_vmag["unit"].loc[0]), "Prumer [{}]".format(self.csv_diameter["unit"].loc[0]), "Teplota [{}]".format(self.csv_temperature["unit"].loc[0]), "Hmotnostni prutok [{}]".format(self.csv_mfr["unit"].loc[0]), "Hmotnost [{}]".format(self.csv_mass["unit"].loc[0]), "Cas letu [{}]".format(self.csv_time["unit"].loc[0])]
            self.csv_statistics["Veličina"] = ["Pozice X [{}]".format(self.csv_x["unit"].loc[0]), "Pozice Y [{}]".format(self.csv_z["unit"].loc[0]), "Pozice Z [{}]".format(self.csv_z["unit"].loc[0]), "Rychlost X [{}]".format(self.csv_vx["unit"].loc[0]), "Rychlost Y [{}]".format(self.csv_vy["unit"].loc[0]), "Rychlost Z [{}]".format(self.csv_vz["unit"].loc[0]), "Rychlost [{}]".format(self.csv_vmag["unit"].loc[0]), "Prumer [{}]".format(self.csv_diameter["unit"].loc[0]), "Teplota [{}]".format(self.csv_temperature["unit"].loc[0]), "Hmotnostni prutok[{}]".format(self.csv_mfr["unit"].loc[0]), "Hmotnost [{}]".format(self.csv_mass["unit"].loc[0]), "Cas letu [{}]".format(self.csv_time["unit"].loc[0])]
            self.csv_summary.to_csv(workdir + "/{}_table_summary_CZ.csv".format(self.name), index = False)
            self.csv_statistics.columns = ["Promenna", "Minimum", "Maximum", "Ar. prumer", "Median"]
            self.csv_statistics.to_csv(workdir + "/{}_table_statistics_CZ.csv".format(self.name), index = False)
        
        elif self.language == "EN":
            self.csv_summary.columns = ["Name [-]", "X Position [{}]".format(self.csv_x["unit"].loc[0]), "Y Position [{}]".format(self.csv_z["unit"].loc[0]), "Z Position [{}]".format(self.csv_z["unit"].loc[0]), "X Velocity [{}]".format(self.csv_vx["unit"].loc[0]), "Y Velocity [{}]".format(self.csv_vy["unit"].loc[0]), "Z Velocity [{}]".format(self.csv_vz["unit"].loc[0]), "Velocity [{}]".format(self.csv_vmag["unit"].loc[0]), "Diameter [{}]".format(self.csv_diameter["unit"].loc[0]), "Temperature [{}]".format(self.csv_temperature["unit"].loc[0]), "Mass Flow Rate [{}]".format(self.csv_mfr["unit"].loc[0]), "Mass [{}]".format(self.csv_mass["unit"].loc[0]), "Time of Flight [{}]".format(self.csv_time["unit"].loc[0])]
            self.csv_statistics["Variable"] = ["X Position [{}]".format(self.csv_x["unit"].loc[0]), "Y Position [{}]".format(self.csv_z["unit"].loc[0]), "Z Position [{}]".format(self.csv_z["unit"].loc[0]), "X Velocity [{}]".format(self.csv_vx["unit"].loc[0]), "Y Velocity [{}]".format(self.csv_vy["unit"].loc[0]), "Z Velocity [{}]".format(self.csv_vz["unit"].loc[0]), "Velocity [{}]".format(self.csv_vmag["unit"].loc[0]), "Diameter [{}]".format(self.csv_diameter["unit"].loc[0]), "Temperature [{}]".format(self.csv_temperature["unit"].loc[0]), "Mass Flow Rate [{}]".format(self.csv_mfr["unit"].loc[0]), "Mass [{}]".format(self.csv_mass["unit"].loc[0]), "Time of Flight [{}]".format(self.csv_time["unit"].loc[0])]
            self.csv_summary.to_csv(workdir + "/{}_table_summary_EN.csv".format(self.name), index = False)
            self.csv_statistics.columns = ["Variable", "Minimum", "Maximum", "Ar. average", "Median"]
            self.csv_statistics.to_csv(workdir + "/{}_table_statistics_EN.csv".format(self.name), index = False)

        del self.csv_x, self.csv_y, self.csv_z, self.csv_vx, self.csv_vy, self.csv_vz, self.csv_vmag, self.csv_diameter, self.csv_temperature, self.csv_mfr, self.csv_mass, self.csv_time
        print("TABLE_SPAWN_CSV... end")

    def columns_best_fit(self, ws: worksheet.worksheet.Worksheet):
                        
            """
            Make all columns best fit
            """
            #column_letters = tuple(openpyxl.utils.get_column_letter(col_number + 1) for col_number in range(ws.max_column))
            #print("Columns_best_fit: sheet name {}".format(ws.name))
            #column_letters = tuple(gcl(col_number + 1) for col_number in range(ws.max_column))
            #for column_letter in column_letters:
                #ws.column_dimensions[column_letter].bestFit = True
            
    def auto_extractor_dict(self, dataframe, range_count, *args):
        print("****************************")
        print("BEGIN....AUTO EXTRACTOR") 
        self.tmp_dataframe_source = dataframe
        self.tmp_vals_source    = []
        self.list_extracted     = []
        for ob in dataframe:
            self.tmp_vals_source.append(ob.get(list(dataframe[0])[1]))

        self.range_count = range_count
        self.min_value = float(min(self.tmp_vals_source))
        self.max_value = float(max(self.tmp_vals_source))
        print("AUTO EXTRACTOR... total range ({}, {})".format(self.min_value, self.max_value))
        self.range_vals = list(linspace(self.min_value, self.max_value, self.range_count))

        print("AUTO EXTRACTOR... total range ({}, {})".format(self.min_value, self.max_value))

        if len(args) == 0:
            print("AUTO EXTRACTOR... BASIC VERZE")
            for i in self.range_vals:
                index_i1    = self.range_vals.index(i)
                index_i2    = self.range_vals.index(i)+1
                
                if index_i2 < len(self.range_vals):
                    ob_i1       = self.range_vals[index_i1]
                    ob_i2       = self.range_vals[index_i2]
                    self.list_extracted.append(self.extract_by_parameters_dict(self.tmp_dataframe_source, ob_i1, ob_i2))
                    self.hist_labels.append("({:.3g}, {:.3g}) [{}]".format(ob_i1, ob_i2, str(dataframe[0].get(list(dataframe[0])[2]))))
                else:
                    ob_i1       = self.range_vals[index_i1]
                    self.list_extracted.append(self.extract_by_parameters_dict(self.tmp_dataframe_source, ob_i1, ob_i1*1.2))
                    self.hist_labels.append("({:.3g}, {:.3g}) [{}]".format(ob_i1, ob_i1*1.2, str(dataframe[0].get(list(dataframe[0])[2]))))                    
                    print("auto_extractor: End of value list")
            print("AUTO EXTRACTOR HIST LABELS", self.hist_labels)

        elif len(args) == 1:
            print("AUTO EXTRACTOR... FILTER VERZE")
            for i in self.range_vals:
                index_i1    = self.range_vals.index(i)
                index_i2    = self.range_vals.index(i)+1
                
                if index_i2 < len(self.range_vals):
                    ob_i1       = self.range_vals[index_i1]
                    ob_i2       = self.range_vals[index_i2]
                    self.list_extracted.append(self.extract_by_parameters_dict(self.tmp_dataframe_source, ob_i1, ob_i2, args[0]))
                    self.hist_labels.append("({:.3g}, {:.3g}) [{}]".format(ob_i1, ob_i2, str(dataframe[0].get(list(dataframe[0])[2]))))
                else:
                    ob_i1       = self.range_vals[index_i1]
                    self.list_extracted.append(self.extract_by_parameters_dict(self.tmp_dataframe_source, ob_i1, ob_i1*1.2, args[0]))
                    self.hist_labels.append("({:.3g}, {:.3g}) [{}]".format(ob_i1, ob_i1*1.2, str(dataframe[0].get(list(dataframe[0])[2]))))  
                    print("auto_extractor: End of value list")
            print("AUTO EXTRACTOR HIST LABELS", self.hist_labels)

        print("END....AUTO EXTRACTOR")
        print("****************************\n")
        self.tmp_vals_source.clear()
        return self.list_extracted

    def extract_by_parameters_dict(self, dataframe, range_min, range_max, *args):
        print("****************************")
        print("BEGIN....EXTRACT BY PARAMETER")
        self.tmp_dataframe_tmpsource = []
        self.tmp_dataframe_tmpsource_ids = []
        self.tmp_dataframe_tmptarget = []

        if (range_max - range_min) == 0:
            range_min = range_min*0.99
            range_max = range_max*1.01
        print("EXTRACT BY PAREMETERS... dfrm {} ......key {}".format(str(dataframe[0])[0:50], list(dataframe[0])[1]))
        for ob in dataframe:
            if ob.get(list(dataframe[0])[1]) >= range_min and ob.get(list(dataframe[0])[1]) < range_max:
                self.tmp_dataframe_tmpsource.append(ob)
                self.tmp_dataframe_tmpsource_ids.append(ob.get(list(dataframe[0])[0]))

        if len(args) == 0:
            print("EXTRACT BY PARAMETERS.... EXTRAKCE ZAKLADNICH DAT")
            self.tmp_dataframe_tmpsource_ids.clear()
            print("END....EXTRACT BY PARAMETER")
            print("****************************\n")
            return self.tmp_dataframe_tmpsource

        elif len(args) == 1:
            print("EXTRACT BY PARAMETERS.... EXTRAKCE JINYCH DAT")
            if isinstance(args[0], list):
                for ob in args[0]:
                    if ob.get(list(dataframe[0])[0]) in self.tmp_dataframe_tmpsource_ids:
                        self.tmp_dataframe_tmptarget.append(ob)
            self.tmp_dataframe_tmpsource_ids.clear()
            print("END....EXTRACT BY PARAMETER")
            print("****************************\n")
            return self.tmp_dataframe_tmptarget

class gui:
    def __init__(self):
        #region Colors
        self.blue = "sky blue"
        self.grey = "dark grey"
        self.red = "red"
        self.green  = "light green"
        self.cgrey = "#e6e6e6"
        #endregion
        
        self.root = Tk()

        #region Page_number
        self.page_number = IntVar()
        self.page_number_max = IntVar()
        self.page_number.set(0)
        self.page_number_max.set(6)
        #endregion Page_number

        #region General-font,background
        self.root.option_add('*Font', 'Segoe_UI 15')
        self.root.option_add('*Button*Background', self.cgrey)
        #endregion        
        
        #region StrinVar_path
        self.path_in    = StringVar()
        self.path_out   = StringVar()
        self.path_in.set("")
        self.path_out.set("")
                
        self.dpm_location = StringVar()
        self.dpm_file = StringVar()

        self.dpm_location.set("")
        self.dpm_file.set("")
        #endregion StrinVar_path

        #region StringVar_units
        self.unit_pos = StringVar()
        self.unit_velo = StringVar()
        self.unit_dia = StringVar()
        self.unit_temp = StringVar()
        self.unit_mfr = StringVar()
        self.unit_mass = StringVar()
        self.unit_frequency = StringVar()
        self.unit_time = StringVar()
        self.unit_pos.set("m")
        self.unit_velo.set("m/s")
        self.unit_dia.set("m")
        self.unit_temp.set("K")
        self.unit_mfr.set("kg/s")
        self.unit_mass.set("kg")
        self.unit_frequency.set("1/s")
        self.unit_time.set("s")
        #endregion StringVar_units

        self.gui_language = StringVar()
        self.gui_language.set("CZ")


        self.dpm_selected= BooleanVar()
        self.dpm_generated= BooleanVar()
        self.dpm_tabulka = ""

        self.dpm_selected.set(False)
        self.dpm_generated.set(False)

        #region BooleanVar_histogram-enables
        self.histo_posx = BooleanVar()
        self.histo_posy = BooleanVar()
        self.histo_posz = BooleanVar()
        self.histo_vx = BooleanVar()
        self.histo_vy = BooleanVar()
        self.histo_vz = BooleanVar()
        self.histo_vmag = BooleanVar()
        self.histo_diameter = BooleanVar()
        self.histo_temperature = BooleanVar()
        self.histo_mfr = BooleanVar()
        self.histo_mass = BooleanVar()
        self.histo_freq = BooleanVar()
        self.histo_time = BooleanVar()
        

        self.histo_posx.set(False)
        self.histo_posy.set(False)
        self.histo_posz.set(False)
        self.histo_vx.set(False)
        self.histo_vy.set(False)
        self.histo_vz.set(False)
        self.histo_vmag.set(False)
        self.histo_diameter.set(False)
        self.histo_temperature.set(False)
        self.histo_mfr.set(False)
        self.histo_mass.set(False)
        self.histo_freq.set(False)
        self.histo_time.set(False)

        self.gui_histo_bool = BooleanVar()
        self.gui_histo_bool.set(True)

        self.plot_percent = BooleanVar()
        self.plot_percent.set(True)
        #endregion BooleanVar_histogram-enables

        self.gui_range_count = IntVar()
        self.gui_range_count.set(10)


        self.gui_multihisto_xmin = DoubleVar()
        self.gui_multihisto_xmax = DoubleVar()
        self.gui_multihisto_xmin.set(0.0)
        self.gui_multihisto_xmax.set(1000.0)
        
        #region BooleanVar_multihistogram-enables
        self.gui_multihisto_bool = BooleanVar()
        self.gui_multihisto_bool.set(False)

        self.bool_multihisto_x1 = BooleanVar()
        self.bool_multihisto_y1 = BooleanVar()
        self.bool_multihisto_z1 = BooleanVar()
        self.bool_multihisto_vx1 = BooleanVar()
        self.bool_multihisto_vy1 = BooleanVar()
        self.bool_multihisto_vz1 = BooleanVar()
        self.bool_multihisto_vmag1 = BooleanVar()
        self.bool_multihisto_diameter1 = BooleanVar()
        self.bool_multihisto_temperature1 = BooleanVar()
        self.bool_multihisto_mfr1 = BooleanVar()
        self.bool_multihisto_mass1 = BooleanVar()
        self.bool_multihisto_time1 = BooleanVar()
        self.bool_multihisto_x1.set(False)
        self.bool_multihisto_y1.set(False)
        self.bool_multihisto_z1.set(False)
        self.bool_multihisto_vx1.set(False)
        self.bool_multihisto_vy1.set(False)
        self.bool_multihisto_vz1.set(False)
        self.bool_multihisto_vmag1.set(False)
        self.bool_multihisto_diameter1.set(True)
        self.bool_multihisto_temperature1.set(False)
        self.bool_multihisto_mfr1.set(False)
        self.bool_multihisto_mass1.set(False)
        self.bool_multihisto_time1.set(False)

        self.bool_multihisto_x2 = BooleanVar()
        self.bool_multihisto_y2 = BooleanVar()
        self.bool_multihisto_z2 = BooleanVar()
        self.bool_multihisto_vx2 = BooleanVar()
        self.bool_multihisto_vy2 = BooleanVar()
        self.bool_multihisto_vz2 = BooleanVar()
        self.bool_multihisto_vmag2 = BooleanVar()
        self.bool_multihisto_diameter2 = BooleanVar()
        self.bool_multihisto_temperature2 = BooleanVar()
        self.bool_multihisto_mfr2 = BooleanVar()
        self.bool_multihisto_mass2 = BooleanVar()
        self.bool_multihisto_time2 = BooleanVar()
        self.bool_multihisto_x2.set(False)
        self.bool_multihisto_y2.set(False)
        self.bool_multihisto_z2.set(False)
        self.bool_multihisto_vx2.set(False)
        self.bool_multihisto_vy2.set(False)
        self.bool_multihisto_vz2.set(False)
        self.bool_multihisto_vmag2.set(False)
        self.bool_multihisto_diameter2.set(False)
        self.bool_multihisto_temperature2.set(False)
        self.bool_multihisto_mfr2.set(False)
        self.bool_multihisto_mass2.set(False)
        self.bool_multihisto_time2.set(True)        
        #endregion BooleanVar_multihistogram-enables
        
        #region StringVar_scatter
        self.scatter_arg_1 = StringVar()
        self.scatter_arg_2 = StringVar()
        self.scatter_arg_3 = StringVar()

        self.scatter_arg_1.set("diameter")
        self.scatter_arg_2.set("time")
        self.scatter_arg_3.set("none")
        #endregion StringVar_scatter

        #region Properties_scatter
        self.gui_scatter_xlim_bool = BooleanVar()
        self.gui_scatter_xmin = DoubleVar()
        self.gui_scatter_xmax = DoubleVar()
        self.gui_scatter_xlim_bool.set(False)        
        self.gui_scatter_xmin.set(0.0)
        self.gui_scatter_xmax.set(1000.0)

        self.gui_scatter_ylim_bool = BooleanVar()
        self.gui_scatter_ylim_min = DoubleVar()
        self.gui_scatter_ylim_max = DoubleVar()
        self.gui_scatter_ylim_bool.set(False)
        self.gui_scatter_ylim_min.set(-1000.0)
        self.gui_scatter_ylim_max.set(1000.0)
        #endregion Properties_scatter

        #region buttons
            #region PAGE_0---prev-next
        self.page_title      = Label(master = self.root, text = "")
        self.label_page0_1  = Label(master = self.root, text = "Program tvoří histogramy, filtruje data, bodové grafy a tabulky ze sample files DPM Fluentu")
        self.label_page0_2  = Label(master = self.root, text = "TechSoft Engineering, Tomáš Prejda, prejdat1993@gmail.com")
        self.button_back    = Button(master = self.root, text = "Zpět",     command = self.page_prev)
        self.button_next    = Button(master = self.root, text = "Další",    command = self.page_next)
            #endregion

            #region PAGE_1---select file
        self.label_language = Label(master = self.root, text = "Výběr jazyka výstupů")
        self.button_language_czech = Button(master = self.root, text = "Čeština", command = lambda: [self.gui_language.set("CZ"), self.get_language(), self.button_language_czech.config(state = "disabled", background = self.blue), self.button_language_english.config(state = "normal", background = self.grey) ])
        self.button_language_english = Button(master = self.root, text = "Angličtina", command = lambda: [self.gui_language.set("EN"), self.get_language(), self.button_language_czech.config(state = "normal", background = self.grey), self.button_language_english.config(state = "disabled", background = self.blue) ])

        self.button_browse_1 = Button(master = self.root, text = "Vyber zdrojový soubor", command = self.browse_1)
        self.button_browse_1_toggle = Button(master = self.root, text = "Máš vybráno?", command = self.spawn_tabulka)
        self.button_generate_source = Button(master = self.root, text = "Generovat tabulku", command = lambda:[self.generate_table(), self.button_language_czech.config(state = "disabled", background = self.blue), self.button_language_english.config(state = "normal", background = self.grey) ])

   
            #endregion

            #region PAGE_2---select units
        self.button_switch_units_global = Button(master = self.root, text = "Toto okno je dobrovolné", command = self.unit_switch_global)
                #region position_buttons
        self.label_pos      = Label(master = self.root, text = "Pozice")
        self.button_pos_mm  = Button(master = self.root, text = "[mm]",              command = lambda: [self.unit_pos.set("mm"), print("Old: {}".format(self.unit_pos.get())), self.unit_bground_pos(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_pos_cm  = Button(master = self.root, text = "[cm]",              command = lambda: [self.unit_pos.set("cm"), print("Old: {}".format(self.unit_pos.get())), self.unit_bground_pos(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_pos_dm  = Button(master = self.root, text = "[dm]",              command = lambda: [self.unit_pos.set("dm"), print("Old: {}".format(self.unit_pos.get())), self.unit_bground_pos(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_pos_m   = Button(master = self.root, text = "[m]",               command = lambda: [self.unit_pos.set("m"), print("Old: {}".format(self.unit_pos.get())), self.unit_bground_pos(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
                #endregion

                #region velocity_buttons
        self.label_velo      = Label(master = self.root, text = "Rychlost")
        self.button_velo_mms  = Button(master = self.root, text = "[mm/s]",          command = lambda: [self.unit_velo.set("mm/s"), print("Old: {}".format(self.unit_velo.get())), self.unit_bground_velo(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_velo_cms  = Button(master = self.root, text = "[cm/s]",          command = lambda: [self.unit_velo.set("cm/s"), print("Old: {}".format(self.unit_velo.get())), self.unit_bground_velo(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_velo_dms  = Button(master = self.root, text = "[dm/s]",          command = lambda: [self.unit_velo.set("dm/s"), print("Old: {}".format(self.unit_velo.get())), self.unit_bground_velo(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_velo_ms  = Button(master = self.root, text = "[m/s]",            command = lambda: [self.unit_velo.set("m/s"), print("Old: {}".format(self.unit_velo.get())), self.unit_bground_velo(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_velo_kmh  = Button(master = self.root, text = "[km/h]",          command = lambda: [self.unit_velo.set("km/h"), print("Old: {}".format(self.unit_velo.get())), self.unit_bground_velo(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
                #endregion

                #region diameter_buttons
        self.label_dia      = Label(master = self.root, text = "Průměr")
        self.button_dia_nm  = Button(master = self.root, text = "[nm]",              command = lambda: [self.unit_dia.set("nm"), print("Old: {}".format(self.unit_dia.get())), self.unit_bground_dia(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_dia_um  = Button(master = self.root, text = "[\u03BCm]",         command = lambda: [self.unit_dia.set("um"), print("Old: {}".format(self.unit_dia.get())), self.unit_bground_dia(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_dia_mm  = Button(master = self.root, text = "[mm]",              command = lambda: [self.unit_dia.set("mm"), print("Old: {}".format(self.unit_dia.get())), self.unit_bground_dia(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_dia_cm  = Button(master = self.root, text = "[cm]",              command = lambda: [self.unit_dia.set("cm"), print("Old: {}".format(self.unit_dia.get())), self.unit_bground_dia(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_dia_dm  = Button(master = self.root, text = "[dm]",              command = lambda: [self.unit_dia.set("dm"), print("Old: {}".format(self.unit_dia.get())), self.unit_bground_dia(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_dia_m   = Button(master = self.root, text = "[m]",               command = lambda: [self.unit_dia.set("m"), print("Old: {}".format(self.unit_dia.get())), self.unit_bground_dia(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
                #endregion

                #region temperature_buttons
        self.label_temp      = Label(master = self.root, text = "Teplota")
        self.button_temp_k  = Button(master = self.root, text = "[K]",              command = lambda: [self.unit_temp.set("K"), print("Old: {}".format(self.unit_temp.get())), self.unit_bground_temp(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_temp_C  = Button(master = self.root, text = "[°C]",             command = lambda: [self.unit_temp.set("°C"), print("Old: {}".format(self.unit_temp.get())), self.unit_bground_temp(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
                #endregion

                #region mfr_buttons
        self.label_mfr   = Label(master = self.root, text = "Hm. průtok")
        self.button_mfr_ugs     = Button(master = self.root, text = "[\u03BCg/s]",  command = lambda: [self.unit_mfr.set("ug/s"), print("Old: {}".format(self.unit_mfr.get())), self.unit_bground_mfr(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_mfr_mgs     = Button(master = self.root, text = "[mg/s]",       command = lambda: [self.unit_mfr.set("mg/s"), print("Old: {}".format(self.unit_mfr.get())), self.unit_bground_mfr(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_mfr_gs      = Button(master = self.root, text = "[g/s]",        command = lambda: [self.unit_mfr.set("g/s"), print("Old: {}".format(self.unit_mfr.get())), self.unit_bground_mfr(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_mfr_kgs     = Button(master = self.root, text = "[kg/s]",       command = lambda: [self.unit_mfr.set("kg/s"), print("Old: {}".format(self.unit_mfr.get())), self.unit_bground_mfr(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_mfr_kgh     = Button(master = self.root, text = "[kg/h]",       command = lambda: [self.unit_mfr.set("kg/s"), print("Old: {}".format(self.unit_mfr.get())), self.unit_bground_mfr(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])  
                #endregion
        
                #region mass_buttons
        self.label_mass         = Label(master = self.root, text = "Hmotnost")
        self.button_mass_ng     = Button(master = self.root, text = "[ng]",           command = lambda: [self.unit_mass.set("ng"), print("Old: {}".format(self.unit_mass.get())), self.unit_bground_mass(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_mass_ug     = Button(master = self.root, text = "[\u03BCg]",      command = lambda: [self.unit_mass.set("ug"), print("Old: {}".format(self.unit_mass.get())), self.unit_bground_mass(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_mass_mg     = Button(master = self.root, text = "[mg]",           command = lambda: [self.unit_mass.set("mg"), print("Old: {}".format(self.unit_mass.get())), self.unit_bground_mass(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_mass_g      = Button(master = self.root, text = "[g]",            command = lambda: [self.unit_mass.set("g"), print("Old: {}".format(self.unit_mass.get())), self.unit_bground_mass(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_mass_kg     = Button(master = self.root, text = "[kg]",           command = lambda: [self.unit_mass.set("kg"), print("Old: {}".format(self.unit_mass.get())), self.unit_bground_mass(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")]) 
                #endregion

                #region time_buttons
        self.label_time         = Label(master = self.root, text = "Čas letu")
        self.button_time_ms     = Button(master = self.root, text = "[ms]",             command = lambda: [self.unit_time.set("ms"), print("Old: {}".format(self.unit_time.get())), self.unit_bground_time(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_time_s      = Button(master = self.root, text = "[s]",              command = lambda: [self.unit_time.set("s"), print("Old: {}".format(self.unit_time.get())), self.unit_bground_time(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_time_min    = Button(master = self.root, text = "[min]",            command = lambda: [self.unit_time.set("min"), print("Old: {}".format(self.unit_time.get())), self.unit_bground_time(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_time_h      = Button(master = self.root, text = "[h]",              command = lambda: [self.unit_time.set("h"), print("Old: {}".format(self.unit_time.get())), self.unit_bground_time(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")])
        self.button_time_den     = Button(master = self.root, text = "[den]",           command = lambda: [self.unit_time.set("den"), print("Old: {}".format(self.unit_time.get())), self.unit_bground_time(), self.button_next.config(state= "disabled"), self.button_switch_units_global.config(state = "normal", background= self.red, text = "Potvrď změnu!")]) 
                #endregion

            #endregion

            #region PAGE_3---histo
        self.label_input_DPI    = Label(master  = self.root, text = "DPI")
        self.text_input_DPI     = Entry(master   = self.root)
        self.button_input_DPI   = Button(master = self.root, text = "Potvrď!", command = lambda:[self.accept_DPI(), self.page_3_enabler(), self.histo_enabler()])

        self.label_input_BINS       = Label(master  = self.root, text = "Počet sloupců histogramu")
        self.text_input_BINS        = Entry(master   = self.root)
        self.button_input_BINS      = Button(master = self.root, text = "Potvrď!", command = lambda:[self.accept_BINS(), self.page_3_enabler(), self.histo_enabler()])

        self.label_plot_percent     = Label(master  = self.root, text = "Co na osu Y?")
        self.button_plot_percent    = Button(master = self.root, text = "[%]",     command = lambda: [self.plot_percent.set(True),     self.accept_percent(), self.histo_enabler()])
        self.button_plot_nonpercent = Button(master = self.root, text = "[-]",     command = lambda: [self.plot_percent.set(False),    self.accept_percent(), self.histo_enabler()])

        self.button_reset_options   = Button(master = self.root, text = "RESET PANELU", command = self.reset_options)
        
        self.button_histo_posx = Button(master = self.root, text = "Pozice X", command = self.select_histogram_posx)
        self.button_histo_posy = Button(master = self.root, text = "Pozice Y", command = self.select_histogram_posy)
        self.button_histo_posz = Button(master = self.root, text = "Pozice Z", command = self.select_histogram_posz)

        self.button_histo_vx = Button(master = self.root, text = "Rychlost X", command = self.select_histogram_vx)
        self.button_histo_vy = Button(master = self.root, text = "Rychlost Y", command = self.select_histogram_vy)
        self.button_histo_vz = Button(master = self.root, text = "Rychlost Z", command = self.select_histogram_vz)
        self.button_histo_vmag = Button(master = self.root, text = "Rychlost (celková)", command = self.select_histogram_vmag)

        self.button_histo_diameter = Button(master = self.root, text = "Průměr", command = self.select_histogram_diameter)
        self.button_histo_temperature = Button(master = self.root, text = "Teplota", command = self.select_histogram_temperature)
        self.button_histo_mfr = Button(master = self.root, text = "Hm. průtok", command = self.select_histogram_mfr)
        self.button_histo_mass = Button(master = self.root, text = "Hmotnost", command = self.select_histogram_mass)
        self.button_histo_time = Button(master = self.root, text = "Čas", command = self.select_histogram_time)

        
        self.label_histo_xmin             = Label(master = self.root, text = "Osa X min:")
        self.text_input_histo_xmin        = Entry(master = self.root)
        self.button_histo_xmin_confirm    = Button(master = self.root, text = "Potvrď!", command = lambda: [self.accept_xmin_histo(), self.histo_enabler()])

        self.label_histo_xmax             = Label(master = self.root, text = "Osa X max:")
        self.text_input_histo_xmax        = Entry(master = self.root)
        self.button_histo_xmax_confirm    = Button(master = self.root, text = "Potvrď!",  command = lambda: [self.accept_xmax_histo(), self.histo_enabler()])
        self.button_histo_xlim_reset      = Button(master = self.root, text = "Reset!",   command = lambda: [self.button_histo_xmax_confirm.config(state = "normal", background= self.grey, text = "Potvrď"), self.text_input_histo_xmax.delete(0, 'end'), self.text_input_histo_xmax.config(state = "normal"), self.button_histo_xmin_confirm.config(state = "normal", background= self.grey, text = "Potvrď"), self.text_input_histo_xmin.delete(0, 'end'), self.text_input_histo_xmin.config(state = "normal"), self.histo_enabler()])
        self.button_histo_xlim_allow      = Button(master = self.root, text = "Manuál",   command = lambda: [self.accept_xlim_histo(), self.histo_enabler()])


        self.button_browse_2 = Button(master = self.root, text = "Vyber cílovou složku", command = lambda:[self.browse_2(), self.page_3_enabler()])
        self.button_create_histos = Button(master = self.root, text = "Generuj obrázky!", command = self.create_histo_pictures)
            #endregion 

            #region PAGE_4---multihistos!
        self.label_multihistos_col1             = Label(master  = self.root, text = "Zdroj dat")
        self.button_multihistos_x               = Button(master = self.root, text = "X"         , command = lambda:[self.turn_off_histo_col_1(), self.bool_multihisto_x1.set(True), self.button_multihistos_x.config(background =            self.blue, state = "disabled"), self.multihisto_enabler()])
        self.button_multihistos_y               = Button(master = self.root, text = "Y"         , command = lambda:[self.turn_off_histo_col_1(), self.bool_multihisto_y1.set(True), self.button_multihistos_y.config(background =            self.blue, state = "disabled"), self.multihisto_enabler()])
        self.button_multihistos_z               = Button(master = self.root, text = "Z"         , command = lambda:[self.turn_off_histo_col_1(), self.bool_multihisto_z1.set(True), self.button_multihistos_z.config(background =            self.blue, state = "disabled"), self.multihisto_enabler()])
        self.button_multihistos_vx              = Button(master = self.root, text = "Rychlost X", command = lambda:[self.turn_off_histo_col_1(), self.bool_multihisto_vx1.set(True), self.button_multihistos_vx.config(background =            self.blue, state = "disabled"), self.multihisto_enabler()])
        self.button_multihistos_vy              = Button(master = self.root, text = "Rychlost Y", command = lambda:[self.turn_off_histo_col_1(), self.bool_multihisto_vy1.set(True), self.button_multihistos_vy.config(background =            self.blue, state = "disabled"), self.multihisto_enabler()])
        self.button_multihistos_vz              = Button(master = self.root, text = "Rychlost Z", command = lambda:[self.turn_off_histo_col_1(), self.bool_multihisto_vz1.set(True), self.button_multihistos_vz.config(background =            self.blue, state = "disabled"), self.multihisto_enabler()])
        self.button_multihistos_vmag            = Button(master = self.root, text = "Rychlost"  , command = lambda:[self.turn_off_histo_col_1(), self.bool_multihisto_vmag1.set(True), self.button_multihistos_vmag.config(background =            self.blue, state = "disabled"), self.multihisto_enabler()])
        self.button_multihistos_diameter        = Button(master = self.root, text = "Průměr"    , command = lambda:[self.turn_off_histo_col_1(), self.bool_multihisto_diameter1.set(True), self.button_multihistos_diameter.config(background =            self.blue, state = "disabled"), self.multihisto_enabler()])
        self.button_multihistos_temperature     = Button(master = self.root, text = "Teplota"   , command = lambda:[self.turn_off_histo_col_1(), self.bool_multihisto_temperature1.set(True), self.button_multihistos_temperature.config(background =            self.blue, state = "disabled"), self.multihisto_enabler()])
        self.button_multihistos_mfr             = Button(master = self.root, text = "Hm. tok"   , command = lambda:[self.turn_off_histo_col_1(), self.bool_multihisto_mfr1.set(True), self.button_multihistos_mfr.config(background =            self.blue, state = "disabled"), self.multihisto_enabler()])
        self.button_multihistos_mass            = Button(master = self.root, text = "Hmotnost"  , command = lambda:[self.turn_off_histo_col_1(), self.bool_multihisto_mass1.set(True), self.button_multihistos_mass.config(background =            self.blue, state = "disabled"), self.multihisto_enabler()])
        self.button_multihistos_time            = Button(master = self.root, text = "Čas"       , command = lambda:[self.turn_off_histo_col_1(), self.bool_multihisto_time1.set(True), self.button_multihistos_time.config(background =            self.blue, state = "disabled"), self.multihisto_enabler()])
        
        self.label_temp                          = Label(master = self.root, text= "")
        self.label_temp2                         = Label(master = self.root, text= "")
        self.label_temp3                         = Label(master = self.root, text= "")
        self.label_temp4                         = Label(master = self.root, text= "")
        self.label_multihistos_col2              = Label(master  = self.root, text = "Filtr: klikni pro obrázek")
        self.button_multihistos_x2               = Button(master = self.root, text = "X"         , command = lambda:[self.turn_off_histo_col_2(), self.bool_multihisto_x2.set(True), self.multihistos_filter_x2(), self.button_multihistos_x2.config(background =            self.blue, state = "normal"), self.multihisto_enabler()])
        self.button_multihistos_y2               = Button(master = self.root, text = "Y"         , command = lambda:[self.turn_off_histo_col_2(), self.bool_multihisto_y2.set(True), self.multihistos_filter_y2(), self.button_multihistos_y2.config(background =            self.blue, state = "normal"), self.multihisto_enabler()])
        self.button_multihistos_z2               = Button(master = self.root, text = "Z"         , command = lambda:[self.turn_off_histo_col_2(), self.bool_multihisto_z2.set(True), self.multihistos_filter_z2(), self.button_multihistos_z2.config(background =            self.blue, state = "normal"), self.multihisto_enabler()])
        self.button_multihistos_vx2              = Button(master = self.root, text = "Rychlost X", command = lambda:[self.turn_off_histo_col_2(), self.bool_multihisto_vx2.set(True), self.multihistos_filter_vx2(), self.button_multihistos_vx2.config(background =            self.blue, state = "normal"), self.multihisto_enabler()])
        self.button_multihistos_vy2              = Button(master = self.root, text = "Rychlost Y", command = lambda:[self.turn_off_histo_col_2(), self.bool_multihisto_vy2.set(True), self.multihistos_filter_vy2(), self.button_multihistos_vy2.config(background =            self.blue, state = "normal"), self.multihisto_enabler()])
        self.button_multihistos_vz2              = Button(master = self.root, text = "Rychlost Z", command = lambda:[self.turn_off_histo_col_2(), self.bool_multihisto_vz2.set(True), self.multihistos_filter_vz2(), self.button_multihistos_vz2.config(background =            self.blue, state = "normal"), self.multihisto_enabler()])
        self.button_multihistos_vmag2            = Button(master = self.root, text = "Rychlost"  , command = lambda:[self.turn_off_histo_col_2(), self.bool_multihisto_vmag2.set(True), self.multihistos_filter_vmag2(), self.button_multihistos_vmag2.config(background =            self.blue, state = "normal"), self.multihisto_enabler()])
        self.button_multihistos_diameter2        = Button(master = self.root, text = "Průměr"    , command = lambda:[self.turn_off_histo_col_2(), self.bool_multihisto_diameter2.set(True), self.multihistos_filter_diameter2(), self.button_multihistos_diameter2.config(background =            self.blue, state = "normal"), self.multihisto_enabler()])
        self.button_multihistos_temperature2     = Button(master = self.root, text = "Teplota"   , command = lambda:[self.turn_off_histo_col_2(), self.bool_multihisto_temperature2.set(True), self.multihistos_filter_temperature2(), self.button_multihistos_temperature2.config(background =            self.blue, state = "normal"), self.multihisto_enabler()])
        self.button_multihistos_mfr2             = Button(master = self.root, text = "Hm. tok"   , command = lambda:[self.turn_off_histo_col_2(), self.bool_multihisto_mfr2.set(True), self.multihistos_filter_mfr2(), self.button_multihistos_mfr2.config(background =            self.blue, state = "normal"), self.multihisto_enabler()])
        self.button_multihistos_mass2            = Button(master = self.root, text = "Hmotnost"  ,  command = lambda:[self.turn_off_histo_col_2(), self.bool_multihisto_mass2.set(True), self.multihistos_filter_mass2(), self.button_multihistos_mass2.config(background =            self.blue, state = "normal"), self.multihisto_enabler()])
        self.button_multihistos_time2            = Button(master = self.root, text = "Čas"       , command = lambda:[self.turn_off_histo_col_2(), self.bool_multihisto_time2.set(True), self.multihistos_filter_time2(), self.button_multihistos_time2.config(background =            self.blue, state = "normal"), self.multihisto_enabler()])


        self.label_multihistos_ranges           = Label(master = self.root, text = "Počet rozsahů")
        self.text_input_multihistos_ranges      = Entry(master = self.root)
        self.button_multihistos_ranges_confirm  = Button(master = self.root, text = "Potvrď!",  command = lambda: [self.accept_ranges_multihistos(), self.multihisto_enabler()])
        self.button_multihistos_ranges_reset    = Button(master = self.root, text = "Reset!",   command = lambda: [self.button_multihistos_ranges_confirm.config(state = "normal", background= self.grey, text = "Potvrď"), self.text_input_multihistos_ranges.delete(0, 'end'), self.text_input_multihistos_ranges.config(state = "normal"), self.multihisto_enabler()])

        self.label_multihistos_bins           = Label(master = self.root, text = "Počet sloupců")
        self.text_input_multihistos_bins      = Entry(master = self.root)
        self.button_multihistos_bins_confirm  = Button(master = self.root, text = "Potvrď!",    command = lambda: [self.accept_BINS_multihistos(), self.multihisto_enabler()])
        self.button_multihistos_bins_reset    = Button(master = self.root, text = "Reset!",     command = lambda: [self.button_multihistos_bins_confirm.config(state = "normal", background= self.grey, text = "Potvrď"), self.text_input_multihistos_bins.delete(0, 'end'), self.text_input_multihistos_bins.config(state = "normal"), self.multihisto_enabler()])

        self.label_multihistos_xmin             = Label(master = self.root, text = "Osa X min:")
        self.text_input_multihistos_xmin        = Entry(master = self.root)
        self.button_multihistos_xmin_confirm    = Button(master = self.root, text = "Potvrď!", command = lambda: [self.accept_xmin_multihistos(), self.multihisto_enabler()])

        self.label_multihistos_xmax             = Label(master = self.root, text = "Osa X max:")
        self.text_input_multihistos_xmax        = Entry(master = self.root)
        self.button_multihistos_xmax_confirm    = Button(master = self.root, text = "Potvrď!",  command = lambda: [self.accept_xmax_multihistos(), self.multihisto_enabler()])
        self.button_multihistos_xlim_reset      = Button(master = self.root, text = "Reset!",   command = lambda: [self.button_multihistos_xmax_confirm.config(state = "normal", background= self.grey, text = "Potvrď"), self.text_input_multihistos_xmax.delete(0, 'end'), self.text_input_multihistos_xmax.config(state = "normal"), self.button_multihistos_xmin_confirm.config(state = "normal", background= self.grey, text = "Potvrď"), self.text_input_multihistos_xmin.delete(0, 'end'), self.text_input_multihistos_xmin.config(state = "normal"), self.multihisto_enabler()])
        self.button_multihistos_xlim_allow      = Button(master = self.root, text = "Manuál",   command = lambda: [self.accept_xlim_multihistos(), self.multihisto_enabler()])
        
        self.label_multihistos_tmp = Label(master = self.root, text = "Neplatný rozsah, min > max")
        self.button_browse_3 = Button(master = self.root, text = "Vyber cílovou složku", command = self.browse_3)
            #endregion

            #region PAGE_5---scatter
                #region COL1
        self.label_scatter_col1                 = Button(master  = self.root, text = "Data X")
        self.button_scatter_x1                  = Button(master = self.root, text = "Pozice X",     command = lambda: [self.scatter_col1_reset(), self.button_scatter_x1.config(state = "disabled", background = self.blue), self.scatter_arg_1.set("x")])
        self.button_scatter_y1         	        = Button(master = self.root, text = "Pozice Y",     command = lambda: [self.scatter_col1_reset(), self.button_scatter_y1.config(state = "disabled", background = self.blue), self.scatter_arg_1.set("y")])
        self.button_scatter_z1         	        = Button(master = self.root, text = "Pozice Z",     command = lambda: [self.scatter_col1_reset(), self.button_scatter_z1.config(state = "disabled", background = self.blue), self.scatter_arg_1.set("z")])
        self.button_scatter_vx1         	    = Button(master = self.root, text = "Rychl. X",     command = lambda: [self.scatter_col1_reset(), self.button_scatter_vx1.config(state = "disabled", background = self.blue), self.scatter_arg_1.set("vx")])
        self.button_scatter_vy1         	    = Button(master = self.root, text = "Rychl. Y",     command = lambda: [self.scatter_col1_reset(), self.button_scatter_vy1.config(state = "disabled", background = self.blue), self.scatter_arg_1.set("vy")])
        self.button_scatter_vz1         	    = Button(master = self.root, text = "Rychl. Z",     command = lambda: [self.scatter_col1_reset(), self.button_scatter_vz1.config(state = "disabled", background = self.blue), self.scatter_arg_1.set("vz")])
        self.button_scatter_vmag1         	    = Button(master = self.root, text = "Rychlost",     command = lambda: [self.scatter_col1_reset(), self.button_scatter_vmag1.config(state = "disabled", background = self.blue), self.scatter_arg_1.set("vmag")])
        self.button_scatter_diameter1         	= Button(master = self.root, text = "Průměr",       command = lambda: [self.scatter_col1_reset(), self.button_scatter_diameter1.config(state = "disabled", background = self.blue), self.scatter_arg_1.set("diameter")])
        self.button_scatter_temperature1        = Button(master = self.root, text = "Teplota",      command = lambda: [self.scatter_col1_reset(), self.button_scatter_temperature1.config(state = "disabled", background = self.blue), self.scatter_arg_1.set("temperature")])
        self.button_scatter_mfr1         	    = Button(master = self.root, text = "Hm. tok",      command = lambda: [self.scatter_col1_reset(), self.button_scatter_mfr1.config(state = "disabled", background = self.blue), self.scatter_arg_1.set("mfr")])
        self.button_scatter_mass1         	    = Button(master = self.root, text = "Hmotnost",     command = lambda: [self.scatter_col1_reset(), self.button_scatter_mass1.config(state = "disabled", background = self.blue), self.scatter_arg_1.set("mass")])
        self.button_scatter_time1         	    = Button(master = self.root, text = "Čas letu",     command = lambda: [self.scatter_col1_reset(), self.button_scatter_time1.config(state = "disabled", background = self.blue), self.scatter_arg_1.set("time")])
        self.label_scatter_col1_sep             = Label(master = self.root, text = "")
                #endregion COL1

                #region COL2
        self.label_scatter_col2                 = Button(master  = self.root, text = "Data Y")
        self.button_scatter_x2                  = Button(master = self.root, text = "Pozice X",     command = lambda: [self.scatter_col2_reset(), self.button_scatter_x2.config(state = "disabled", background = self.blue), self.scatter_arg_2.set("x")])
        self.button_scatter_y2         	        = Button(master = self.root, text = "Pozice Y",     command = lambda: [self.scatter_col2_reset(), self.button_scatter_y2.config(state = "disabled", background = self.blue), self.scatter_arg_2.set("y")])
        self.button_scatter_z2         	        = Button(master = self.root, text = "Pozice Z",     command = lambda: [self.scatter_col2_reset(), self.button_scatter_z2.config(state = "disabled", background = self.blue), self.scatter_arg_2.set("z")])
        self.button_scatter_vx2         	    = Button(master = self.root, text = "Rychl. X",     command = lambda: [self.scatter_col2_reset(), self.button_scatter_vx2.config(state = "disabled", background = self.blue), self.scatter_arg_2.set("vx")])
        self.button_scatter_vy2         	    = Button(master = self.root, text = "Rychl. Y",     command = lambda: [self.scatter_col2_reset(), self.button_scatter_vy2.config(state = "disabled", background = self.blue), self.scatter_arg_2.set("vy")])
        self.button_scatter_vz2         	    = Button(master = self.root, text = "Rychl. Z",     command = lambda: [self.scatter_col2_reset(), self.button_scatter_vz2.config(state = "disabled", background = self.blue), self.scatter_arg_2.set("vz")])
        self.button_scatter_vmag2         	    = Button(master = self.root, text = "Rychlost",     command = lambda: [self.scatter_col2_reset(), self.button_scatter_vmag2.config(state = "disabled", background = self.blue), self.scatter_arg_2.set("vmag")])
        self.button_scatter_diameter2         	= Button(master = self.root, text = "Průměr",       command = lambda: [self.scatter_col2_reset(), self.button_scatter_diameter2.config(state = "disabled", background = self.blue), self.scatter_arg_2.set("diameter")])
        self.button_scatter_temperature2        = Button(master = self.root, text = "Teplota",      command = lambda: [self.scatter_col2_reset(), self.button_scatter_temperature2.config(state = "disabled", background = self.blue), self.scatter_arg_2.set("temperature")])
        self.button_scatter_mfr2         	    = Button(master = self.root, text = "Hm. tok",      command = lambda: [self.scatter_col2_reset(), self.button_scatter_mfr2.config(state = "disabled", background = self.blue), self.scatter_arg_2.set("mfr")])
        self.button_scatter_mass2         	    = Button(master = self.root, text = "Hmotnost",     command = lambda: [self.scatter_col2_reset(), self.button_scatter_mass2.config(state = "disabled", background = self.blue), self.scatter_arg_2.set("mass")])
        self.button_scatter_time2         	    = Button(master = self.root, text = "Čas letu",     command = lambda: [self.scatter_col2_reset(), self.button_scatter_time2.config(state = "disabled", background = self.blue), self.scatter_arg_2.set("time")])
        self.label_scatter_col2_sep             = Label(master = self.root, text = "")
                #endregion COL2

                #region COL3
        self.label_scatter_col3                 = Button(master  = self.root, text = "Barvící funkce")
        self.button_scatter_none3               = Button(master = self.root, text = "Nic",      command = lambda: [self.scatter_col3_reset(), self.button_scatter_none3.config(state = "normal", background = self.blue), self.scatter_arg_3.set("none"), self.dpm_tabulka.dpmscatter(self.dpm_tabulka.__dict__.get(self.scatter_arg_1.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_2.get()))])
        self.button_scatter_x3                  = Button(master = self.root, text = "Pozice X", command = lambda: [self.scatter_col3_reset(), self.button_scatter_x3.config(state = "normal", background = self.blue), self.scatter_arg_3.set("x"), self.dpm_tabulka.dpmscatter(self.dpm_tabulka.__dict__.get(self.scatter_arg_1.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_2.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_3.get()))])
        self.button_scatter_y3         	        = Button(master = self.root, text = "Pozice Y", command = lambda: [self.scatter_col3_reset(), self.button_scatter_y3.config(state = "normal", background = self.blue), self.scatter_arg_3.set("y"), self.dpm_tabulka.dpmscatter(self.dpm_tabulka.__dict__.get(self.scatter_arg_1.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_2.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_3.get()))])
        self.button_scatter_z3         	        = Button(master = self.root, text = "Pozice Z", command = lambda: [self.scatter_col3_reset(), self.button_scatter_z3.config(state = "normal", background = self.blue), self.scatter_arg_3.set("z"), self.dpm_tabulka.dpmscatter(self.dpm_tabulka.__dict__.get(self.scatter_arg_1.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_2.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_3.get()))])
        self.button_scatter_vx3         	    = Button(master = self.root, text = "Rychl. X", command = lambda: [self.scatter_col3_reset(), self.button_scatter_vx3.config(state = "normal", background = self.blue), self.scatter_arg_3.set("vx"), self.dpm_tabulka.dpmscatter(self.dpm_tabulka.__dict__.get(self.scatter_arg_1.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_2.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_3.get()))])
        self.button_scatter_vy3         	    = Button(master = self.root, text = "Rychl. Y", command = lambda: [self.scatter_col3_reset(), self.button_scatter_vy3.config(state = "normal", background = self.blue), self.scatter_arg_3.set("vy"), self.dpm_tabulka.dpmscatter(self.dpm_tabulka.__dict__.get(self.scatter_arg_1.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_2.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_3.get()))])
        self.button_scatter_vz3         	    = Button(master = self.root, text = "Rychl. Z", command = lambda: [self.scatter_col3_reset(), self.button_scatter_vz3.config(state = "normal", background = self.blue), self.scatter_arg_3.set("vz"), self.dpm_tabulka.dpmscatter(self.dpm_tabulka.__dict__.get(self.scatter_arg_1.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_2.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_3.get()))])
        self.button_scatter_vmag3         	    = Button(master = self.root, text = "Rychlost", command = lambda: [self.scatter_col3_reset(), self.button_scatter_vmag3.config(state = "normal", background = self.blue), self.scatter_arg_3.set("vmag"), self.dpm_tabulka.dpmscatter(self.dpm_tabulka.__dict__.get(self.scatter_arg_1.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_2.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_3.get()))])
        self.button_scatter_diameter3         	= Button(master = self.root, text = "Průměr",   command = lambda: [self.scatter_col3_reset(), self.button_scatter_diameter3.config(state = "normal", background = self.blue), self.scatter_arg_3.set("diameter"), self.dpm_tabulka.dpmscatter(self.dpm_tabulka.__dict__.get(self.scatter_arg_1.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_2.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_3.get()))])
        self.button_scatter_temperature3        = Button(master = self.root, text = "Teplota",  command = lambda: [self.scatter_col3_reset(), self.button_scatter_temperature3.config(state = "normal", background = self.blue), self.scatter_arg_3.set("temperature"), self.dpm_tabulka.dpmscatter(self.dpm_tabulka.__dict__.get(self.scatter_arg_1.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_2.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_3.get()))])
        self.button_scatter_mfr3         	    = Button(master = self.root, text = "Hm. tok",  command = lambda: [self.scatter_col3_reset(), self.button_scatter_mfr3.config(state = "normal", background = self.blue), self.scatter_arg_3.set("mfr"), self.dpm_tabulka.dpmscatter(self.dpm_tabulka.__dict__.get(self.scatter_arg_1.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_2.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_3.get()))])
        self.button_scatter_mass3         	    = Button(master = self.root, text = "Hmotnost", command = lambda: [self.scatter_col3_reset(), self.button_scatter_mass3.config(state = "normal", background = self.blue), self.scatter_arg_3.set("mass"), self.dpm_tabulka.dpmscatter(self.dpm_tabulka.__dict__.get(self.scatter_arg_1.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_2.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_3.get()))])
        self.button_scatter_time3         	    = Button(master = self.root, text = "Čas letu", command = lambda: [self.scatter_col3_reset(), self.button_scatter_time3.config(state = "normal", background = self.blue), self.scatter_arg_3.set("time"), self.dpm_tabulka.dpmscatter(self.dpm_tabulka.__dict__.get(self.scatter_arg_1.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_2.get()), self.dpm_tabulka.__dict__.get(self.scatter_arg_3.get()))])        
                #endregion COL3

        self.button_browse_4                    = Button(master     = self.root, text = "Vyber cílovou složku", command = self.browse_4)

        self.label_scatter_ylim_min             = Label(master      = self.root, text = "Barva min:")
        self.text_input_scatter_ylim_min        = Entry(master      = self.root)
        self.button_scatter_ylim_min_confirm    = Button(master     = self.root, text = "Potvrď!", command = lambda: [self.accept_ymin_scatter(), self.scatter_enabler()])


        self.label_scatter_ylim_max             = Label(master  = self.root, text = "Barva max:")
        self.text_input_scatter_ylim_max        = Entry(master  = self.root)
        self.button_scatter_ylim_max_confirm    = Button(master = self.root, text = "Potvrď!",  command = lambda: [self.accept_ymax_scatter(), self.scatter_enabler()])
        self.button_scatter_ylim_lim_reset      = Button(master = self.root, text = "Reset!",   command = lambda: [self.button_scatter_ylim_max_confirm.config(state = "normal", background= self.grey, text = "Potvrď"), self.text_input_scatter_ylim_max.delete(0, 'end'), self.text_input_scatter_ylim_max.config(state = "normal"), self.button_scatter_ylim_min_confirm.config(state = "normal", background= self.grey, text = "Potvrď"), self.text_input_scatter_ylim_min.delete(0, 'end'), self.text_input_scatter_ylim_min.config(state = "normal"), self.scatter_enabler()])
        self.button_scatter_ylim_lim_allow      = Button(master = self.root, text = "Manuál",   command = lambda: [self.accept_ylim_scatter(), self.scatter_enabler()])

        self.label_scatter_xmin             = Label(master      = self.root, text = "Osa X min:")
        self.text_input_scatter_xmin        = Entry(master      = self.root)
        self.button_scatter_xmin_confirm    = Button(master     = self.root, text = "Potvrď!", command = lambda: [self.accept_xmin_scatter(), self.scatter_enabler()])


        self.label_scatter_xmax             = Label(master  = self.root, text = "Osa X max:")
        self.text_input_scatter_xmax        = Entry(master  = self.root)
        self.button_scatter_xmax_confirm    = Button(master = self.root, text = "Potvrď!",  command = lambda: [self.accept_xmax_scatter(), self.scatter_enabler()])
        self.button_scatter_xlim_reset      = Button(master = self.root, text = "Reset!",   command = lambda: [self.button_scatter_xmax_confirm.config(state = "normal", background= self.grey, text = "Potvrď"), self.text_input_scatter_xmax.delete(0, 'end'), self.text_input_scatter_xmax.config(state = "normal"), self.button_scatter_xmin_confirm.config(state = "normal", background= self.grey, text = "Potvrď"), self.text_input_scatter_xmin.delete(0, 'end'), self.text_input_scatter_xmin.config(state = "normal"), self.scatter_enabler()])
        self.button_scatter_xlim_allow      = Button(master = self.root, text = "Manuál",   command = lambda: [self.accept_xlim_scatter(), self.scatter_enabler()])
            #endregion

            #region PAGE_6---table
        self.button_browse_5 = Button(master = self.root, text = "Vyber adresář", command = lambda: [self.browse_5(), self.button_browse_5.config(background= self.green)])
        self.button_table_spawn_table = Button(master = self.root, text = "Vytvoř tabulky", command = lambda: [self.dpm_tabulka.table_spawn_csv(), self.button_table_spawn_table.config(background = self.green)])
            #endregion PAGE_6---table

            #region init_button_colors
        self.page_0_browser()
        self.unit_bground_pos()
        self.unit_bground_velo()
        self.unit_bground_dia()
        self.unit_bground_temp()
        self.unit_bground_mfr()
        self.unit_bground_mass()
        self.unit_bground_time()
        self.spawn_tabulka_toggle()
        self.button_reset_options.config(background=self.red)
        self.accept_percent()


        self.button_language_czech.config(state = "disabled", background = self.blue)
        self.button_language_english.config(state = "disabled", background = self.grey)
           
                #region histo_button_init
        self.button_histo_posx.config(background=self.grey)
        self.button_histo_posy.config(background=self.grey)
        self.button_histo_posz.config(background=self.grey)
        self.button_histo_vx.config(background=self.grey)
        self.button_histo_vy.config(background=self.grey)
        self.button_histo_vz.config(background=self.grey)
        self.button_histo_vmag.config(background=self.grey)
        self.button_histo_diameter.config(background=self.grey)
        self.button_histo_temperature.config(background=self.grey)
        self.button_histo_mfr.config(background=self.grey)
        self.button_histo_mass.config(background=self.grey)
        self.button_histo_time.config(background=self.grey)

        
        self.button_histo_xmin_confirm.config(background=self.grey, state= "disabled")
        self.button_histo_xmax_confirm.config(background=self.grey, state = "disabled")
        self.text_input_histo_xmin.config(state = "disabled")
        self.text_input_histo_xmax.config(state = "disabled")
        self.button_histo_xlim_reset.config(state= "disabled")
        self.button_histo_xlim_reset.config(background=self.red)
        self.button_histo_xlim_allow.config(background=self.green, text = "Automat")
                #endregion histo_button_init

                #region multihisto_button_init
        self.button_multihistos_x.config(state = "normal", background = self.grey)
        self.button_multihistos_y.config(state = "normal", background = self.grey)
        self.button_multihistos_z.config(state = "normal", background = self.grey)
        self.button_multihistos_vx.config(state = "normal", background = self.grey)
        self.button_multihistos_vy.config(state = "normal", background = self.grey)
        self.button_multihistos_vz.config(state = "normal", background = self.grey)
        self.button_multihistos_vmag.config(state = "normal", background = self.grey)
        self.button_multihistos_diameter.config(state = "disabled", background = self.blue)
        self.button_multihistos_temperature.config(state = "normal", background = self.grey)
        self.button_multihistos_mfr.config(state = "normal", background = self.grey)
        self.button_multihistos_mass.config(state = "normal", background = self.grey)
        self.button_multihistos_time.config(state = "normal", background = self.grey)

        self.button_multihistos_x2.config(state = "normal", background = self.grey)
        self.button_multihistos_y2.config(state = "normal", background = self.grey)
        self.button_multihistos_z2.config(state = "normal", background = self.grey)
        self.button_multihistos_vx2.config(state = "normal", background = self.grey)
        self.button_multihistos_vy2.config(state = "normal", background = self.grey)
        self.button_multihistos_vz2.config(state = "normal", background = self.grey)
        self.button_multihistos_vmag2.config(state = "normal", background = self.grey)
        self.button_multihistos_diameter2.config(state = "normal", background = self.grey)
        self.button_multihistos_temperature2.config(state = "normal", background = self.grey)
        self.button_multihistos_mfr2.config(state = "normal", background = self.grey)
        self.button_multihistos_mass2.config(state = "normal", background = self.grey)
        self.button_multihistos_time2.config(state = "normal", background = self.blue)

        self.button_multihistos_bins_confirm.config(state = "normal", background = self.grey)
        self.button_multihistos_bins_reset.config(state = "normal", background = self.red)
        self.button_multihistos_ranges_confirm.config(state = "normal", background = self.grey)
        self.button_multihistos_ranges_reset.config(state = "normal", background = self.red)

        self.button_multihistos_xmin_confirm.config(state = "normal", background = self.grey)
        self.button_multihistos_xmax_confirm.config(state = "normal", background = self.grey)
        self.button_multihistos_xlim_reset.config(state = "normal", background = self.red)
        self.button_multihistos_xlim_allow.config(state = "normal", background = self.grey)

        self.gui_multihisto_bool.set(True)
        self.text_input_multihistos_xmin.config(state = "disabled")
        self.text_input_multihistos_xmax.config(state = "disabled")
        self.button_multihistos_xlim_reset.config(state = "disabled")
        self.button_multihistos_xmin_confirm.config(state = "disabled")
        self.button_multihistos_xmax_confirm.config(state = "disabled")
        self.button_multihistos_xlim_allow.config(state = "normal", background= self.green, text = "Automat")
                #endregion multihisto_button_init
            
                #region scatter_button_init
        self.label_scatter_col1.config(state= "disabled", background = "yellow")
        self.button_scatter_x1.config(state = "normal", background = self.grey)
        self.button_scatter_y1.config(state = "normal", background = self.grey)
        self.button_scatter_z1.config(state = "normal", background = self.grey)
        self.button_scatter_vx1.config(state = "normal", background = self.grey)
        self.button_scatter_vy1.config(state = "normal", background = self.grey)
        self.button_scatter_vz1.config(state = "normal", background = self.grey)
        self.button_scatter_vmag1.config(state = "normal", background = self.grey)
        self.button_scatter_diameter1.config(state = "disabled", background = self.blue)
        self.button_scatter_temperature1.config(state = "normal", background = self.grey)
        self.button_scatter_mfr1.config(state = "normal", background = self.grey)
        self.button_scatter_mass1.config(state = "normal", background = self.grey)
        self.button_scatter_time1.config(state = "normal", background = self.grey)

        self.label_scatter_col2.config(state= "disabled", background = "yellow")
        self.button_scatter_x2.config(state = "normal", background = self.grey)
        self.button_scatter_y2.config(state = "normal", background = self.grey)
        self.button_scatter_z2.config(state = "normal", background = self.grey)
        self.button_scatter_vx2.config(state = "normal", background = self.grey)
        self.button_scatter_vy2.config(state = "normal", background = self.grey)
        self.button_scatter_vz2.config(state = "normal", background = self.grey)
        self.button_scatter_vmag2.config(state = "normal", background = self.grey)
        self.button_scatter_diameter2.config(state = "normal", background = self.grey)
        self.button_scatter_temperature2.config(state = "normal", background = self.grey)
        self.button_scatter_mfr2.config(state = "normal", background = self.grey)
        self.button_scatter_mass2.config(state = "normal", background = self.grey)
        self.button_scatter_time2.config(state = "disabled", background = self.blue)

        self.label_scatter_col3.config(state= "disabled", background = "yellow")
        self.button_scatter_none3.config(state = "normal", background = self.blue)
        self.button_scatter_x3.config(state = "normal", background = self.grey)
        self.button_scatter_y3.config(state = "normal", background = self.grey)
        self.button_scatter_z3.config(state = "normal", background = self.grey)
        self.button_scatter_vx3.config(state = "normal", background = self.grey)
        self.button_scatter_vy3.config(state = "normal", background = self.grey)
        self.button_scatter_vz3.config(state = "normal", background = self.grey)
        self.button_scatter_vmag3.config(state = "normal", background = self.grey)
        self.button_scatter_diameter3.config(state = "normal", background = self.grey)
        self.button_scatter_temperature3.config(state = "normal", background = self.grey)
        self.button_scatter_mfr3.config(state = "normal", background = self.grey)
        self.button_scatter_mass3.config(state = "normal", background = self.grey)
        self.button_scatter_time3.config(state = "normal", background = self.grey)        

        self.button_scatter_ylim_lim_allow.config(background = self.grey)
        self.button_scatter_xlim_allow.config(background = self.grey)
        self.button_scatter_ylim_lim_reset.config(state = "normal", background = self.red)
        self.button_scatter_xlim_reset.config(state = "normal", background = self.red)  

        self.gui_scatter_ylim_bool.set(True)
        self.text_input_scatter_ylim_min.config(state = "disabled")
        self.text_input_scatter_ylim_max.config(state = "disabled")
        self.button_scatter_ylim_lim_reset.config(state = "disabled")
        self.button_scatter_ylim_min_confirm.config(state = "disabled")
        self.button_scatter_ylim_max_confirm.config(state = "disabled")
        self.button_scatter_ylim_lim_allow.config(state = "normal", background= self.green, text = "Automat")

        self.gui_scatter_xlim_bool.set(True)
        self.text_input_scatter_xmin.config(state = "disabled")
        self.text_input_scatter_xmax.config(state = "disabled")
        self.button_scatter_xlim_reset.config(state = "disabled")
        self.button_scatter_xmin_confirm.config(state = "disabled")
        self.button_scatter_xmax_confirm.config(state = "disabled")
        self.button_scatter_xlim_allow.config(state = "normal", background= self.green, text = "Automat")
                #endregion scatter_button_init

                #region table_button_init
        self.button_table_spawn_table.config(background= self.grey)
        self.button_browse_5.config(background= self.grey)
                #endregion
            #endregion 
        #endregion


        #region printout
        self.printout_bools()
        #endregion

    def page_next(self):
        self.page_cleanup()
        if self.page_number.get() == 0:
            self.page_1_browser()
            self.page_number.set(self.page_number.get()+1)
        elif self.page_number.get() == 1:
            self.page_2_browser()
            self.page_number.set(self.page_number.get()+1)

        elif self.page_number.get() == 2:
            self.page_3_browser()
            self.page_number.set(self.page_number.get()+1)
        

        elif self.page_number.get() == 3:
            self.page_4_browser()
            self.page_number.set(self.page_number.get()+1)

        elif self.page_number.get() == 4:
            self.page_5_browser()
            self.page_number.set(self.page_number.get()+1)

        elif self.page_number.get() == 5:
            self.page_6_browser()
            self.page_number.set(self.page_number.get()+1)

        else:
            pass
        print("Next, stranka {}".format(self.page_number.get()))
    
    def page_prev(self):
        self.page_cleanup()
        if self.page_number.get() == 0:
            pass
        elif self.page_number.get() == 1:
            self.page_0_browser()
            self.page_number.set(self.page_number.get()-1)

        elif self.page_number.get() == 2:
            self.page_1_browser()
            self.page_number.set(self.page_number.get()-1)    

        elif self.page_number.get() == 3:
            self.page_2_browser()
            self.page_number.set(self.page_number.get()-1)
        
        elif self.page_number.get() == 4:
            self.page_3_browser()
            self.page_number.set(self.page_number.get()-1)

        elif self.page_number.get() == 5:
            self.page_4_browser()
            self.page_number.set(self.page_number.get()-1)

        elif self.page_number.get() == 6:
            self.page_5_browser()
            self.page_number.set(self.page_number.get()-1)    

        elif self.page_number.get() == 7:
            self.page_5_browser()
            self.page_number.set(self.page_number.get()-1)   
        print("Back, stranka {}".format(self.page_number.get()))

 

    #region browse_functions
    def browse_1(self):
        filename = filedialog.askopenfilename(filetypes=[('DPM summary', '.dpm')])
        self.path_in.set(filename)
        print("Zdroj: {}".format(self.path_in.get()))        
        return self.path_in.get()

    def browse_2(self):
        foldername = filedialog.askdirectory()
        self.path_out.set(foldername)
        print("Cíl: {}".format(self.path_out.get()))

        if self.path_out.get() == "":
            self.button_next.config(state = "disabled")
        else:
            self.button_next.config(state = "normal")
            self.button_browse_2.config(state = "normal", text = self.path_out.get(), background=self.green)
            self.button_create_histos.config(state = "normal", text = "Generuj obrázky", background = self.grey)     
        return str(self.path_out.get())

    def browse_3(self):
        foldername = filedialog.askdirectory()
        self.path_out.set(foldername)
        print("Cíl: {}".format(self.path_out.get()))
        self.button_browse_3.config(state = "normal", text = self.path_out.get(), background=self.green)
        self.dpm_tabulka.workdir = str(self.path_out.get())
        return str(self.path_out.get())

    def browse_4(self):
        foldername = filedialog.askdirectory()
        self.path_out.set(foldername)
        print("Cíl: {}".format(self.path_out.get()))
        self.button_browse_4.config(state = "normal", text = self.path_out.get(), background=self.green)
        self.dpm_tabulka.workdir = str(self.path_out.get())
        return str(self.path_out.get())

    def browse_5(self):
        foldername = filedialog.askdirectory()
        self.path_out.set(foldername)
        print("Cíl: {}".format(self.path_out.get()))
        self.button_browse_5.config(state = "normal", text = self.path_out.get(), background=self.green)
        self.dpm_tabulka.workdir = str(self.path_out.get())
        return str(self.path_out.get())
    #endregion browse_functions

    #region pages
    def page_cleanup(self):
        self.button_back.grid_forget()
        self.button_next.grid_forget()
        self.button_browse_1.grid_forget()
        self.button_browse_1_toggle.grid_forget()
        self.button_generate_source.grid_forget()
        self.button_browse_2.grid_forget()
        self.page_title.config(text = "!!!!!!!!!!!!!!!!!!!!!!!!!!!")
        self.button_switch_units_global.grid_forget()

        self.label_page0_1.grid_forget()
        self.label_page0_2.grid_forget()

        self.label_language.grid_forget()
        self.button_language_czech.grid_forget()
        self.button_language_english.grid_forget()

        self.label_pos.grid_forget()
        self.button_pos_mm.grid_forget()
        self.button_pos_cm.grid_forget()
        self.button_pos_dm.grid_forget()
        self.button_pos_m.grid_forget()   

        self.label_velo.grid_forget()
        self.button_velo_mms.grid_forget()
        self.button_velo_cms.grid_forget()
        self.button_velo_dms.grid_forget()
        self.button_velo_ms.grid_forget()
        self.button_velo_kmh.grid_forget()
        
        self.label_dia.grid_forget()
        self.button_dia_nm.grid_forget()
        self.button_dia_um.grid_forget()
        self.button_dia_mm.grid_forget()
        self.button_dia_cm.grid_forget()
        self.button_dia_dm.grid_forget()
        self.button_dia_m.grid_forget()  

        self.label_temp.grid_forget()
        self.button_temp_k.grid_forget()
        self.button_temp_C.grid_forget()

        self.label_mfr.grid_forget()
        self.button_mfr_ugs.grid_forget()
        self.button_mfr_mgs.grid_forget()
        self.button_mfr_gs.grid_forget()
        self.button_mfr_kgs.grid_forget()
        self.button_mfr_kgh.grid_forget()

        self.label_mass.grid_forget()
        self.button_mass_ng.grid_forget()
        self.button_mass_ug.grid_forget()
        self.button_mass_mg.grid_forget()
        self.button_mass_g.grid_forget()
        self.button_mass_kg.grid_forget()

        self.label_time.grid_forget()
        self.button_time_ms.grid_forget()
        self.button_time_s.grid_forget()
        self.button_time_min.grid_forget()
        self.button_time_h.grid_forget()
        self.button_time_den.grid_forget()


        self.label_input_DPI.grid_forget()
        self.text_input_DPI.grid_forget()
        self.button_input_DPI.grid_forget()
        self.label_input_BINS.grid_forget()
        self.text_input_BINS.grid_forget()
        self.button_input_BINS.grid_forget()
        self.button_reset_options.grid_forget()
        self.label_plot_percent.grid_forget()
        self.button_plot_percent.grid_forget()
        self.button_plot_nonpercent.grid_forget()    

        self.button_histo_posx.grid_forget()
        self.button_histo_posy.grid_forget()
        self.button_histo_posz.grid_forget()
        self.button_histo_vx.grid_forget()
        self.button_histo_vy.grid_forget()
        self.button_histo_vz.grid_forget()
        self.button_histo_vmag.grid_forget()
        self.button_histo_diameter.grid_forget()
        self.button_histo_temperature.grid_forget()
        self.button_histo_mfr.grid_forget()
        self.button_histo_mass.grid_forget()
        self.button_histo_time.grid_forget()


        self.label_histo_xmin.grid_forget()
        self.text_input_histo_xmin.grid_forget()
        self.button_histo_xmin_confirm.grid_forget()

        self.label_histo_xmax.grid_forget()
        self.text_input_histo_xmax.grid_forget()
        self.button_histo_xmax_confirm.grid_forget()
        self.button_histo_xlim_reset.grid_forget()
        self.button_histo_xlim_allow.grid_forget()

        self.button_create_histos.grid_forget()


        self.button_browse_3.grid_forget()
        self.label_multihistos_col1.grid_forget()
        self.button_multihistos_x.grid_forget()
        self.button_multihistos_y.grid_forget()
        self.button_multihistos_z.grid_forget()
        self.button_multihistos_vx.grid_forget()
        self.button_multihistos_vy.grid_forget()
        self.button_multihistos_vz.grid_forget()
        self.button_multihistos_vmag.grid_forget()
        self.button_multihistos_diameter.grid_forget()
        self.button_multihistos_temperature.grid_forget()
        self.button_multihistos_mfr.grid_forget()
        self.button_multihistos_mass.grid_forget()
        self.button_multihistos_time.grid_forget()
        
        self.label_temp.grid_forget()
        self.label_temp2.grid_forget()
        self.label_temp3.grid_forget()
        self.label_temp4.grid_forget()
        
        self.label_multihistos_col2.grid_forget()
        self.button_multihistos_x2.grid_forget()
        self.button_multihistos_y2.grid_forget()
        self.button_multihistos_z2.grid_forget()
        self.button_multihistos_vx2.grid_forget()
        self.button_multihistos_vy2.grid_forget()
        self.button_multihistos_vz2.grid_forget()
        self.button_multihistos_vmag2.grid_forget()
        self.button_multihistos_diameter2.grid_forget()
        self.button_multihistos_temperature2.grid_forget()
        self.button_multihistos_mfr2.grid_forget()
        self.button_multihistos_mass2.grid_forget()
        self.button_multihistos_time2.grid_forget()

        self.label_multihistos_bins.grid_forget()
        self.text_input_multihistos_bins.grid_forget()
        self.button_multihistos_bins_confirm.grid_forget()
        self.button_multihistos_bins_reset.grid_forget()
        self.label_multihistos_ranges.grid_forget()
        self.text_input_multihistos_ranges.grid_forget()
        self.button_multihistos_ranges_confirm.grid_forget()
        self.button_multihistos_ranges_reset.grid_forget()

        self.label_multihistos_xmin.grid_forget()
        self.text_input_multihistos_xmin.grid_forget()
        self.button_multihistos_xmin_confirm.grid_forget()
        self.label_multihistos_xmax.grid_forget()
        self.text_input_multihistos_xmax.grid_forget()
        self.button_multihistos_xmax_confirm.grid_forget()
        self.button_multihistos_xlim_reset.grid_forget()
        self.button_multihistos_xlim_allow.grid_forget()
        self.label_multihistos_tmp.grid_forget()


        self.label_scatter_col1.grid_forget()
        self.button_scatter_x1.grid_forget()
        self.button_scatter_y1.grid_forget()
        self.button_scatter_z1.grid_forget()
        self.button_scatter_vx1.grid_forget()
        self.button_scatter_vy1.grid_forget()
        self.button_scatter_vz1.grid_forget()
        self.button_scatter_vmag1.grid_forget()
        self.button_scatter_diameter1.grid_forget()
        self.button_scatter_temperature1.grid_forget()
        self.button_scatter_mfr1.grid_forget()
        self.button_scatter_mass1.grid_forget()
        self.button_scatter_time1.grid_forget()

        self.label_scatter_col2.grid_forget()
        self.button_scatter_x2.grid_forget()
        self.button_scatter_y2.grid_forget()
        self.button_scatter_z2.grid_forget()
        self.button_scatter_vx2.grid_forget()
        self.button_scatter_vy2.grid_forget()
        self.button_scatter_vz2.grid_forget()
        self.button_scatter_vmag2.grid_forget()
        self.button_scatter_diameter2.grid_forget()
        self.button_scatter_temperature2.grid_forget()
        self.button_scatter_mfr2.grid_forget()
        self.button_scatter_mass2.grid_forget()
        self.button_scatter_time2.grid_forget()
        
        self.label_scatter_col3.grid_forget()
        self.button_scatter_none3.grid_forget()
        self.button_scatter_x3.grid_forget()
        self.button_scatter_y3.grid_forget()
        self.button_scatter_z3.grid_forget()
        self.button_scatter_vx3.grid_forget()
        self.button_scatter_vy3.grid_forget()
        self.button_scatter_vz3.grid_forget()
        self.button_scatter_vmag3.grid_forget()
        self.button_scatter_diameter3.grid_forget()
        self.button_scatter_temperature3.grid_forget()
        self.button_scatter_mfr3.grid_forget()
        self.button_scatter_mass3.grid_forget()
        self.button_scatter_time3.grid_forget()  

        self.label_scatter_col1_sep.grid_forget()  
        self.label_scatter_col2_sep.grid_forget() 

        self.button_browse_4.grid_forget() 

        
        self.label_scatter_ylim_min.grid_forget() 
        self.text_input_scatter_ylim_min.grid_forget() 
        self.button_scatter_ylim_min_confirm.grid_forget() 

        self.label_scatter_ylim_max.grid_forget() 
        self.text_input_scatter_ylim_max.grid_forget() 
        self.button_scatter_ylim_max_confirm.grid_forget() 
        self.button_scatter_ylim_lim_reset.grid_forget() 
        self.button_scatter_ylim_lim_allow.grid_forget() 

        self.label_scatter_xmin.grid_forget() 
        self.text_input_scatter_xmin.grid_forget() 
        self.button_scatter_xmin_confirm.grid_forget() 

        self.label_scatter_xmax.grid_forget() 
        self.text_input_scatter_xmax.grid_forget() 
        self.button_scatter_xmax_confirm.grid_forget() 
        self.button_scatter_xlim_reset.grid_forget() 
        self.button_scatter_xlim_allow.grid_forget() 

        self.button_browse_5.grid_forget()
        self.button_table_spawn_table.grid_forget()
 
    def page_0_browser(self):
        a = 0
        b = a + 1
        c = b + 1
        d = c + 1
        e = d + 1 
        f = e + 1
        g = f + 1
        h = g + 1
        j = h + 1

        self.page_title.config(text= "Automatické vyhodnocování DPM")
        self.page_title.grid(row = 0, column = 0, columnspan = 3, sticky = N+S+E+W, pady = 0)
        self.label_page0_1.grid(row = 1, column = 0, sticky = N+S+E+W, pady = 0)
        self.label_page0_2.grid(row = 2, column = 0, sticky = N+S+E+W, pady = 0)
        #self.button_browse_1.grid(row = 1, column = 1, sticky = N+S+E+W, pady = 0)
        self.button_back.grid_forget()
        self.button_next.grid(row = 3, column = c, sticky = N+S+E+W, pady = 0)
        
    def page_1_browser(self):
        a = 0
        b = a + 1
        c = b + 1
        d = c + 1
        e = d + 1 
        f = e + 1
        g = f + 1
        h = g + 1
        j = h + 1

        self.page_title.config(text= "Výběr zdrojového souboru")
        self.page_title.grid(row = a, column = a, columnspan = 3, sticky = N+S+E+W, pady = 0)
        self.button_browse_1.grid(row = b, column = b, sticky = N+S+E+W, pady = 0)
        self.button_browse_1_toggle.grid(row = b, column = c, sticky = N+S+E+W, pady = 0)
        self.button_generate_source.grid(row = b, column = d, sticky = N+S+E+W, pady = 0)

        self.label_temp.grid(row = c, column = a, sticky = N+S+E+W, pady = 0)
        self.label_language.grid(row = d, column = a, sticky = N+S+E+W, pady = 0)
        self.button_language_czech.grid(row = d, column = b, sticky = N+S+E+W, pady = 0)
        self.button_language_english.grid(row = d, column = c, sticky = N+S+E+W, pady = 0)

        self.label_temp2.grid(row = e, column = a, sticky = N+S+E+W, pady = 0)
        self.button_back.grid(row = f, column = b, sticky = N+S+E+W, pady = 0)
        self.button_next.grid(row = f, column = c, sticky = N+S+E+W, pady = 0)


        if self.dpm_selected.get() == False:
            self.button_generate_source.config(state="disabled")
        else:
            self.button_generate_source.config(state="normal")


        if self.dpm_generated.get() == False:
            self.button_next.config(state="disabled")
        else:
            self.button_next.config(state="normal")
            self.button_generate_source.config(state = "disabled")

    def page_2_browser(self):
        a = 0
        b = a + 1
        c = b + 1
        d = c + 1
        e = d + 1 
        f = e + 1
        g = f + 1
        h = g + 1
        j = h + 1


        self.page_title.config(text= "Výběr jednotek")
        self.page_title.grid(row = a, column = 0, columnspan = 3, sticky = N+S+E+W, pady = 0)
        
        self.label_pos.grid(row = b, column = a, sticky = N+S+E+W, pady = 0)
        self.button_pos_mm.grid(row = b, column = b, sticky = N+S+E+W, pady = 0)
        self.button_pos_cm.grid(row = b, column = c, sticky = N+S+E+W, pady = 0)
        self.button_pos_dm.grid(row = b, column = d, sticky = N+S+E+W, pady = 0)
        self.button_pos_m.grid(row = b, column = e, columnspan=3 ,sticky = N+S+E+W, pady = 0)

        self.label_velo.grid(row = c, column = a, sticky = N+S+E+W, pady = 0)
        self.button_velo_mms.grid(row = c, column = b, sticky = N+S+E+W, pady = 0)
        self.button_velo_cms.grid(row = c, column = c, sticky = N+S+E+W, pady = 0)
        self.button_velo_dms.grid(row = c, column = d, sticky = N+S+E+W, pady = 0)
        self.button_velo_ms.grid(row = c, column = e, sticky = N+S+E+W, pady = 0)
        self.button_velo_kmh.grid(row = c, column = f, columnspan=2, sticky = N+S+E+W, pady = 0)

        self.label_dia.grid(row = d, column = a, sticky = N+S+E+W, pady = 0)
        self.button_dia_nm.grid(row = d, column = b, sticky = N+S+E+W, pady = 0)
        self.button_dia_um.grid(row = d, column = c, sticky = N+S+E+W, pady = 0)
        self.button_dia_mm.grid(row = d, column = d, sticky = N+S+E+W, pady = 0)
        self.button_dia_cm.grid(row = d, column = e, sticky = N+S+E+W, pady = 0)
        self.button_dia_dm.grid(row = d, column = f, sticky = N+S+E+W, pady = 0)
        self.button_dia_m.grid(row = d, column = g, sticky = N+S+E+W, pady = 0)

        self.label_temp.grid(row = e, column = a, sticky = N+S+E+W, pady = 0)
        self.button_temp_k.grid(row = e, column = b, columnspan = 3, sticky = N+S+E+W, pady = 0)
        self.button_temp_C.grid(row = e, column = e, columnspan = 3, sticky = N+S+E+W, pady = 0)

        self.label_mfr.grid(row = f, column = a, sticky = N+S+E+W, pady = 0)
        self.button_mfr_ugs.grid(row = f, column = b, sticky = N+S+E+W, pady = 0)
        self.button_mfr_mgs.grid(row = f, column = c, sticky = N+S+E+W, pady = 0)
        self.button_mfr_gs.grid(row = f, column = d, sticky = N+S+E+W, pady = 0)
        self.button_mfr_kgs.grid(row = f, column = e, sticky = N+S+E+W, pady = 0)
        self.button_mfr_kgh.grid(row = f, column = f,columnspan=2, sticky = N+S+E+W, pady = 0)

        self.label_mass.grid(row = g, column = a, sticky = N+S+E+W, pady = 0)
        self.button_mass_ng.grid(row = g, column = b, sticky = N+S+E+W, pady = 0)
        self.button_mass_ug.grid(row = g, column = c, sticky = N+S+E+W, pady = 0)
        self.button_mass_mg.grid(row = g, column = d, sticky = N+S+E+W, pady = 0)
        self.button_mass_g.grid(row = g, column = e, sticky = N+S+E+W, pady = 0)
        self.button_mass_kg.grid(row = g, column = f, columnspan=2, sticky = N+S+E+W, pady = 0)

        self.label_time.grid(row = h, column = a, sticky = N+S+E+W, pady = 0)
        self.button_time_ms.grid(row = h, column = b, sticky = N+S+E+W, pady = 0)
        self.button_time_s.grid(row = h, column = c, sticky = N+S+E+W, pady = 0)
        self.button_time_min.grid(row = h, column = d, sticky = N+S+E+W, pady = 0)
        self.button_time_h.grid(row = h, column = e, sticky = N+S+E+W, pady = 0)
        self.button_time_den.grid(row = h, column = f, columnspan=2,sticky = N+S+E+W, pady = 0)

        self.button_switch_units_global.grid(row = j, column = b, columnspan = 6, sticky = N+S+E+W, pady = 0)
        self.button_switch_units_global.config(text = "Toto okno je dobrovolné", background = self.blue, state = "disabled")

        self.button_back.grid(row = 10, column = b, sticky = N+S+E+W, pady = 0)
        self.button_next.grid(row = 10, column = c, sticky = N+S+E+W, pady = 0)

    def page_3_browser(self):
        a = 0
        b = a + 1
        c = b + 1
        d = c + 1
        e = d + 1 
        f = e + 1
        g = f + 1
        h = g + 1
        i = h + 1
        j = i + 1
        k = j + 1
        l = k + 1
        m = l + 1
        n = m + 1
        o = n + 1
        p = o + 1
        q = p + 1
        self.page_title.config(text= "Kam chceš jaké histogramy?")
        self.page_title.grid(row = a, column = a, columnspan = 3, sticky = N+S+E+W, pady = 0)

        self.button_histo_posx.grid(row = b, column = a, sticky = N+S+E+W, pady = 0)
        self.button_histo_posy.grid(row = b, column = b, sticky = N+S+E+W, pady = 0)
        self.button_histo_posz.grid(row = b, column = c, columnspan=2, sticky = N+S+E+W, pady = 0)
        self.button_histo_vx.grid(row = c, column = a, sticky = N+S+E+W, pady = 0)
        self.button_histo_vy.grid(row = c, column = b, sticky = N+S+E+W, pady = 0)
        self.button_histo_vz.grid(row = c, column = c, sticky = N+S+E+W, pady = 0)
        self.button_histo_vmag.grid(row = c, column = d, sticky = N+S+E+W, pady = 0)
        self.button_histo_diameter.grid(row = d, column = a, columnspan=2, sticky = N+S+E+W, pady = 0)
        self.button_histo_temperature.grid(row = d, column = c, columnspan=2,sticky = N+S+E+W, pady = 0)
        self.button_histo_mfr.grid(row = e, column = a, columnspan=2,sticky = N+S+E+W, pady = 0)
        self.button_histo_mass.grid(row = e, column = c, columnspan=2,sticky = N+S+E+W, pady = 0)
        self.button_histo_time.grid(row = f, column = a, columnspan = 4, sticky = N+S+E+W, pady = 0)
        
        self.label_temp.grid(row = g, column = a, sticky = N+S+E+W, pady = 0)

        self.label_input_DPI.grid(row = h, column = a, sticky = N+S+E+W, pady = 0)
        self.text_input_DPI.grid(row = h, column = b, columnspan = 2, sticky = N+S+E+W, pady = 0)
        self.button_input_DPI.grid(row = h, column = d, columnspan = 2, sticky = N+S+E+W, pady = 0)

        self.label_input_BINS.grid(row = i, column = a, sticky = N+S+E+W, pady = 0)
        self.text_input_BINS.grid(row = i, column = b,  columnspan = 2, sticky = N+S+E+W, pady = 0)
        self.button_input_BINS.grid(row = i, column = d,  columnspan = 2, sticky = N+S+E+W, pady = 0)

        self.label_plot_percent.grid(row = j, column = a, sticky = N+S+E+W, pady = 0)
        self.button_plot_nonpercent.grid(row = j, column = b, columnspan = 2, sticky = N+S+E+W, pady = 0)
        self.button_plot_percent.grid(row = j, column = d, columnspan = 2, sticky = N+S+E+W, pady = 0)

        self.label_temp2.grid(row = k, column = a, sticky = N+S+E+W, pady = 0)


        self.label_histo_xmin.grid(row = l, column = a, columnspan = 1,  sticky = N+S+E+W, pady = 0)
        self.text_input_histo_xmin.grid(row = l, column = b, columnspan = 1,  sticky = N+S+E+W, pady = 0)
        self.button_histo_xmin_confirm.grid(row = l, column = c, columnspan = 1,  sticky = N+S+E+W, pady = 0)

        self.label_histo_xmax.grid(row = m, column = a, columnspan = 1,  sticky = N+S+E+W, pady = 0)
        self.text_input_histo_xmax.grid(row = m, column = b, columnspan = 1,  sticky = N+S+E+W, pady = 0)
        self.button_histo_xmax_confirm.grid(row = m, column = c, columnspan = 1,  sticky = N+S+E+W, pady = 0)
        self.button_histo_xlim_reset.grid(row = l, rowspan = 2,  column = d, columnspan = 1,  sticky = N+S+E+W, pady = 0)
        self.button_histo_xlim_allow.grid(row = l, rowspan = 2,  column = e, columnspan = 1,  sticky = N+S+E+W, pady = 0)
        
        self.label_temp3.grid(row = n, column = a, sticky = N+S+E+W, pady = 0)

        self.button_browse_2.grid(row = o, column = 1, sticky = N+S+E+W, pady = 0)
        self.button_create_histos.grid(row = o, column = 2, columnspan = 3, sticky = N+S+E+W, pady = 0)
        self.label_temp4.grid(row = p, column = a, sticky = N+S+E+W, pady = 0)
        self.button_back.grid(row = q, column = b, sticky = N+S+E+W, pady = 0)
        self.button_next.grid(row = q, column = c, sticky = N+S+E+W, pady = 0)
        self.button_reset_options.grid(row = q, column = d, columnspan = 3, sticky = N+S+E+W, pady = 0)
        
    
        
        self.text_input_histo_xmin.delete(0, "end")
        self.text_input_histo_xmin.insert(0, str(self.gui_multihisto_xmin.get()))      
        self.text_input_histo_xmax.delete(0, "end")
        self.text_input_histo_xmax.insert(0, str(self.gui_multihisto_xmax.get()))


        if self.text_input_DPI.get() == "" or self.text_input_BINS.get() == "":
            self.button_create_histos.config(state = "disabled")
            self.text_input_DPI.delete(0, "end")
            self.text_input_DPI.insert(0, 500)     
            self.text_input_BINS.delete(0, "end")
            self.text_input_BINS.insert(0, 20)

        else:
            self.button_create_histos.config(state = "normal")

        if self.path_out.get() == "":
            self.button_create_histos.config(state = "disabled", background= self.red, text = "Nemáš ještě cílový adresář!")
        else:
            self.button_create_histos.config(state = "normal", background= self.grey, text =  "Generuj histogramy!")
        self.page_3_enabler()
        self.histo_enabler()
        print("Page 3 browser: stranka {}".format(self.page_number.get()))
        

    def page_4_browser(self):
        a = 0
        b = a + 1
        c = b + 1
        d = c + 1
        e = d + 1 
        f = e + 1
        g = f + 1
        h = g + 1
        j = h + 1
        k = j + 1
        l = k + 1
        m = l + 1

        self.page_title.config(text= "Generace multi-histogramů")
        self.page_title.grid(row = a, column = a, columnspan = 3, sticky = N+S+E+W, pady = 0)

        self.label_multihistos_col1.grid(row = b, column = a, sticky = N+S+E+W, pady = 0)
        self.button_multihistos_x.grid(row = b+1, column = a, sticky = N+S+E+W, pady = 0)
        self.button_multihistos_y.grid(row = b+1, column = b, sticky = N+S+E+W, pady = 0)
        self.button_multihistos_z.grid(row = b+1, column = c, columnspan = 2, sticky = N+S+E+W, pady = 0)
        self.button_multihistos_vx.grid(row = c+1, column = a, sticky = N+S+E+W, pady = 0)
        self.button_multihistos_vy.grid(row = c+1, column = b, sticky = N+S+E+W, pady = 0)
        self.button_multihistos_vz.grid(row = c+1, column = c, sticky = N+S+E+W, pady = 0)
        self.button_multihistos_vmag.grid(row = c+1, column = d, sticky = N+S+E+W, pady = 0)
        self.button_multihistos_diameter.grid(row = d+1, column = a, columnspan = 2,  sticky = N+S+E+W, pady = 0)
        self.button_multihistos_temperature.grid(row = d+1, column = c,columnspan = 2,  sticky = N+S+E+W, pady = 0)
        self.button_multihistos_mfr.grid(row = e+1, column = a, columnspan = 2,  sticky = N+S+E+W, pady = 0)
        self.button_multihistos_mass.grid(row = e+1, column = c, columnspan = 2, sticky = N+S+E+W, pady = 0)
        self.button_multihistos_time.grid(row = f+1, column = a, columnspan = 4, sticky = N+S+E+W, pady = 0)

        self.label_temp.grid(row = b, rowspan = 5,  column = e, sticky = N+S+E+W, pady = 0)
        self.label_multihistos_col2.grid(row = b, column = a+5, sticky = N+S+E+W, pady = 0)
        self.button_multihistos_x2.grid(row = b+1, column = a+5, sticky = N+S+E+W, pady = 0)
        self.button_multihistos_y2.grid(row = b+1, column = b+5, sticky = N+S+E+W, pady = 0)
        self.button_multihistos_z2.grid(row = b+1, column = c+5, columnspan = 2, sticky = N+S+E+W, pady = 0)
        self.button_multihistos_vx2.grid(row = c+1, column = a+5, sticky = N+S+E+W, pady = 0)
        self.button_multihistos_vy2.grid(row = c+1, column = b+5, sticky = N+S+E+W, pady = 0)
        self.button_multihistos_vz2.grid(row = c+1, column = c+5, sticky = N+S+E+W, pady = 0)
        self.button_multihistos_vmag2.grid(row = c+1, column = d+5, sticky = N+S+E+W, pady = 0)
        self.button_multihistos_diameter2.grid(row = d+1, column = a+5, columnspan = 2,  sticky = N+S+E+W, pady = 0)
        self.button_multihistos_temperature2.grid(row = d+1, column = c+5,columnspan = 2,  sticky = N+S+E+W, pady = 0)
        self.button_multihistos_mfr2.grid(row = e+1, column = a+5, columnspan = 2,  sticky = N+S+E+W, pady = 0)
        self.button_multihistos_mass2.grid(row = e+1, column = c+5, columnspan = 2, sticky = N+S+E+W, pady = 0)
        self.button_multihistos_time2.grid(row = f+1, column = a+5, columnspan = 4, sticky = N+S+E+W, pady = 0)


        self.label_multihistos_bins.grid(row = g+1, column = a, columnspan = 1,  sticky = N+S+E+W, pady = 0)
        self.text_input_multihistos_bins.grid(row = g+1, column = b, columnspan = 1,  sticky = N+S+E+W, pady = 0)
        self.button_multihistos_bins_confirm.grid(row = g+1, column = c, columnspan = 1,  sticky = N+S+E+W, pady = 0)
        self.button_multihistos_bins_reset.grid(row = g+1, column = d, columnspan = 1,  sticky = N+S+E+W, pady = 0)

        self.label_multihistos_ranges.grid(row = h+1, column = a, columnspan = 1,  sticky = N+S+E+W, pady = 0)
        self.text_input_multihistos_ranges.grid(row = h+1, column = b, columnspan = 1,  sticky = N+S+E+W, pady = 0)
        self.button_multihistos_ranges_confirm.grid(row = h+1, column = c, columnspan = 1,  sticky = N+S+E+W, pady = 0)
        self.button_multihistos_ranges_reset.grid(row = h+1, column = d, columnspan = 1,  sticky = N+S+E+W, pady = 0)

        self.label_multihistos_xmin.grid(row = g+1, column = a+5, columnspan = 1,  sticky = N+S+E+W, pady = 0)
        self.text_input_multihistos_xmin.grid(row = g+1, column = b+5, columnspan = 1,  sticky = N+S+E+W, pady = 0)
        self.button_multihistos_xmin_confirm.grid(row = g+1, column = c+5, columnspan = 1,  sticky = N+S+E+W, pady = 0)

        self.label_multihistos_xmax.grid(row = h+1, column = a+5, columnspan = 1,  sticky = N+S+E+W, pady = 0)
        self.text_input_multihistos_xmax.grid(row = h+1, column = b+5, columnspan = 1,  sticky = N+S+E+W, pady = 0)
        self.button_multihistos_xmax_confirm.grid(row = h+1, column = c+5, columnspan = 1,  sticky = N+S+E+W, pady = 0)
        self.button_multihistos_xlim_reset.grid(row = g+1, rowspan = 2,  column = d+5, columnspan = 1,  sticky = N+S+E+W, pady = 0)
        self.button_multihistos_xlim_allow.grid(row = g+1, rowspan = 2,  column = e+5, columnspan = 1,  sticky = N+S+E+W, pady = 0)
        
        self.button_browse_3.grid(row = l+1, column = a, columnspan = 4,  sticky = N+S+E+W, pady = 0)
        self.button_browse_3.config(state = "normal", text = self.button_browse_2["text"])
        self.button_back.grid(row = m+1, column = b, sticky = N+S+E+W, pady = 0)
        self.button_next.grid(row = m+1, column = c, sticky = N+S+E+W, pady = 0)

        if self.text_input_multihistos_bins.get() == "":
            self.text_input_multihistos_bins.insert(0, "5")
        elif self.text_input_multihistos_bins.get() == "5":
            pass
        self.text_input_multihistos_ranges.delete(0, "end")
        self.text_input_multihistos_ranges.insert(0, str(self.gui_range_count.get()))
        
        self.text_input_multihistos_xmin.delete(0, "end")
        self.text_input_multihistos_xmin.insert(0, str(self.gui_multihisto_xmin.get()))      
        self.text_input_multihistos_xmax.delete(0, "end")
        self.text_input_multihistos_xmax.insert(0, str(self.gui_multihisto_xmax.get()))

        if self.gui_multihisto_bool.get() == True:
            self.text_input_multihistos_xmin.config(state = "disabled")
            self.text_input_multihistos_xmax.config(state = "disabled")
            self.button_multihistos_xlim_reset.config(state = "disabled")
            self.button_multihistos_xmin_confirm.config(state = "disabled")
            self.button_multihistos_xmax_confirm.config(state = "disabled")
            self.button_multihistos_xlim_allow.config(state = "normal", background= self.green, text = "Automat")
            self.dpm_tabulka.xlim_bool = False
        self.multihisto_enabler()
        
    def page_5_browser(self):
        a = 1
        b = a + 1
        c = b + 1
        d = c + 1
        e = d + 1 
        f = e + 1
        g = f + 1
        h = g + 1
        i = h + 1
        j = i + 1
        k = j + 1
        l = k + 1
        m = l + 1
        n = m + 1
        o = n + 1
        p = o + 1
        q = p + 1
        r = q + 1
        s = r + 1
        t = s + 1 
        u = s + 1
        v = u + 1
        w = v + 1


        self.page_title.config(text= "Generace bodových grafů")
        self.page_title.grid(row = 0, column = 0, columnspan = 3, sticky = N+S+E+W, pady = 0)

        self.label_scatter_col1.grid(row = a, column = a, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_x1.grid(row = b, column = a, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_y1.grid(row = c, column = a, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_z1.grid(row = d, column = a, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_vx1.grid(row = e, column = a, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_vy1.grid(row = f, column = a, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_vz1.grid(row = g, column = a, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_vmag1.grid(row = h, column = a, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_diameter1.grid(row = i, column = a, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_temperature1.grid(row = j, column = a, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_mfr1.grid(row = k, column = a, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_mass1.grid(row = l, column = a, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_time1.grid(row = m, column = a, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.label_scatter_col1_sep.grid(row = a, column = b,  rowspan = 13, columnspan = 1, sticky = N+S+E+W, pady = 0)

        self.label_scatter_col2.grid(row = a, column = c, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_x2.grid(row = b, column = c, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_y2.grid(row = c, column = c, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_z2.grid(row = d, column = c, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_vx2.grid(row = e, column = c, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_vy2.grid(row = f, column = c, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_vz2.grid(row = g, column = c, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_vmag2.grid(row = h, column = c, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_diameter2.grid(row = i, column = c, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_temperature2.grid(row = j, column = c, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_mfr2.grid(row = k, column = c, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_mass2.grid(row = l, column = c, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_time2.grid(row = m, column = c, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.label_scatter_col2_sep.grid(row = a, column = d, rowspan = 13, columnspan = 1, sticky = N+S+E+W, pady = 0)
        
        self.label_scatter_col3.grid(row = a, column = e, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_x3.grid(row = b, column = e, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_y3.grid(row = c, column = e, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_z3.grid(row = d, column = e, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_vx3.grid(row = e, column = e, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_vy3.grid(row = f, column = e, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_vz3.grid(row = g, column = e, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_vmag3.grid(row = h, column = e, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_diameter3.grid(row = i, column = e, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_temperature3.grid(row = j, column = e, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_mfr3.grid(row = k, column = e, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_mass3.grid(row = l, column = e, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_time3.grid(row = m, column = e, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_none3.grid(row = n, column = e, columnspan = 1, sticky = N+S+E+W, pady = 0)

        self.button_browse_4.grid(row = o, column = a, columnspan = 5, sticky = N+S+E+W, pady = 0)

        if self.path_out.get() != "":
            self.button_browse_4.config(text= self.path_out.get())
        else:
            self.button_browse_4.config(text = "Vyber adresář")

        self.label_scatter_ylim_min.grid(row = p, column = a, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.text_input_scatter_ylim_min.grid(row = p, column = b, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_ylim_min_confirm.grid(row = p, column = c, columnspan = 1, sticky = N+S+E+W, pady = 0)

        self.label_scatter_ylim_max.grid(row = q, column = a, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.text_input_scatter_ylim_max.grid(row = q, column = b, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_ylim_max_confirm.grid(row = q, column = c, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_ylim_lim_reset.grid(row = p, column = d, rowspan = 2,  columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_ylim_lim_allow.grid(row = p, column = e, rowspan = 2, columnspan = 1, sticky = N+S+E+W, pady = 0)


        self.label_scatter_xmin.grid(row = r, column = a, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.text_input_scatter_xmin.grid(row = r, column = b, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_xmin_confirm.grid(row = r, column = c, columnspan = 1, sticky = N+S+E+W, pady = 0)

        self.label_scatter_xmax.grid(row = s, column = a, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.text_input_scatter_xmax.grid(row = s, column = b, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_xmax_confirm.grid(row = s, column = c, columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_xlim_reset.grid(row = r, column = d, rowspan = 2,  columnspan = 1, sticky = N+S+E+W, pady = 0)
        self.button_scatter_xlim_allow.grid(row = r, column = e, rowspan = 2, columnspan = 1, sticky = N+S+E+W, pady = 0)

        self.button_back.grid(row = u, column = b, sticky = N+S+E+W, pady = 0)
        self.button_next.grid(row = u, column = c, sticky = N+S+E+W, pady = 0)


        if self.text_input_scatter_ylim_min.get() == "":
            self.text_input_scatter_ylim_min.delete(0, "end")
            self.text_input_scatter_ylim_min.insert(0, self.gui_scatter_ylim_min.get())     

        if self.text_input_scatter_ylim_max.get() == "":
            self.text_input_scatter_ylim_max.delete(0, "end")
            self.text_input_scatter_ylim_max.insert(0, self.gui_scatter_ylim_max.get())    

        if self.text_input_scatter_xmin.get() == "":
            self.text_input_scatter_xmin.delete(0, "end")
            self.text_input_scatter_xmin.insert(0, self.gui_scatter_xmin.get())

        if self.text_input_scatter_xmax.get() == "":
            self.text_input_scatter_xmax.delete(0, "end")
            self.text_input_scatter_xmax.insert(0, self.gui_scatter_xmax.get())
        self.scatter_enabler()

    def page_6_browser(self):
        a = 0
        b = a + 1
        c = b + 1
        d = c + 1
        e = d + 1 
        f = e + 1
        g = f + 1
        h = g + 1
        j = h + 1

        self.page_title.config(text= "Generace tabulek (csv, Excel)")
        self.page_title.grid(row = a, column = a, columnspan = 3, sticky = N+S+E+W, pady = 0)
        self.button_browse_5.grid(row = b, column = b, sticky = N+S+E+W, pady = 0)
        self.button_table_spawn_table.grid(row = b, column = c, sticky = N+S+E+W, pady = 0)
        self.button_back.grid(row = c, column = b, sticky = N+S+E+W, pady = 0)
        #self.button_next.grid(row = c, column = c, sticky = N+S+E+W, pady = 0)

        if self.path_out.get() != "":
            self.button_browse_5.config(text= self.path_out.get())
        else:
            self.button_browse_5.config(text = "Vyber adresář")
    #endregion



    #region UNIT_COLOURS
    def unit_bground_pos(self):
        if (self.unit_pos.get() == "mm"):
            self.button_pos_mm.config(background=self.blue)
            self.button_pos_cm.config(background=self.grey)
            self.button_pos_dm.config(background=self.grey)
            self.button_pos_m.config(background=self.grey)
        elif (self.unit_pos.get() == "cm"):
            self.button_pos_mm.config(background=self.grey)
            self.button_pos_cm.config(background=self.blue)
            self.button_pos_dm.config(background=self.grey)
            self.button_pos_m.config(background=self.grey)            
        elif (self.unit_pos.get() == "dm"):
            self.button_pos_mm.config(background=self.grey)
            self.button_pos_cm.config(background=self.grey)
            self.button_pos_dm.config(background=self.blue)
            self.button_pos_m.config(background=self.grey)               
        elif (self.unit_pos.get() == "m"):
            self.button_pos_mm.config(background=self.grey)
            self.button_pos_cm.config(background=self.grey)
            self.button_pos_dm.config(background=self.grey)
            self.button_pos_m.config(background=self.blue)   
        else:
            pass            
 
    def unit_bground_velo(self):
        if (self.unit_velo.get() == "mm/s"):
            self.button_velo_mms.config(background=self.blue)
            self.button_velo_cms.config(background=self.grey)
            self.button_velo_dms.config(background=self.grey)
            self.button_velo_ms.config(background=self.grey)
            self.button_velo_kmh.config(background=self.grey)
        elif (self.unit_velo.get() == "cm/s"):
            self.button_velo_mms.config(background=self.grey)
            self.button_velo_cms.config(background=self.blue)
            self.button_velo_dms.config(background=self.grey)
            self.button_velo_ms.config(background=self.grey)
            self.button_velo_kmh.config(background=self.grey)   
        elif (self.unit_velo.get() == "dm/s"):
            self.button_velo_mms.config(background=self.grey)
            self.button_velo_cms.config(background=self.grey)
            self.button_velo_dms.config(background=self.blue)
            self.button_velo_ms.config(background=self.grey)
            self.button_velo_kmh.config(background=self.grey) 
        elif (self.unit_velo.get() == "m/s"):
            self.button_velo_mms.config(background=self.grey)
            self.button_velo_cms.config(background=self.grey)
            self.button_velo_dms.config(background=self.grey)
            self.button_velo_ms.config(background=self.blue)
            self.button_velo_kmh.config(background=self.grey)  
        elif (self.unit_velo.get() == "km/h"):
            self.button_velo_mms.config(background=self.grey)
            self.button_velo_cms.config(background=self.grey)
            self.button_velo_dms.config(background=self.grey)
            self.button_velo_ms.config(background=self.grey)
            self.button_velo_kmh.config(background=self.blue)              
        else:
            pass    
    
    def unit_bground_dia(self):
        if (self.unit_dia.get() == "nm"):
            self.button_dia_nm.config(background=self.blue)
            self.button_dia_um.config(background=self.grey)
            self.button_dia_mm.config(background=self.grey)
            self.button_dia_cm.config(background=self.grey)
            self.button_dia_dm.config(background=self.grey)
            self.button_dia_m.config(background=self.grey)
        elif (self.unit_dia.get() == "um"):
            self.button_dia_nm.config(background=self.grey)
            self.button_dia_um.config(background=self.blue)
            self.button_dia_mm.config(background=self.grey)
            self.button_dia_cm.config(background=self.grey)
            self.button_dia_dm.config(background=self.grey)
            self.button_dia_m.config(background=self.grey)           
        elif (self.unit_dia.get() == "mm"):
            self.button_dia_nm.config(background=self.grey)
            self.button_dia_um.config(background=self.grey)
            self.button_dia_mm.config(background=self.blue)
            self.button_dia_cm.config(background=self.grey)
            self.button_dia_dm.config(background=self.grey)
            self.button_dia_m.config(background=self.grey)              
        elif (self.unit_dia.get() == "cm"):
            self.button_dia_nm.config(background=self.grey)
            self.button_dia_um.config(background=self.grey)
            self.button_dia_mm.config(background=self.grey)
            self.button_dia_cm.config(background=self.blue)
            self.button_dia_dm.config(background=self.grey)
            self.button_dia_m.config(background=self.grey) 
        elif (self.unit_dia.get() == "dm"):
            self.button_dia_nm.config(background=self.grey)
            self.button_dia_um.config(background=self.grey)
            self.button_dia_mm.config(background=self.grey)
            self.button_dia_cm.config(background=self.grey)
            self.button_dia_dm.config(background=self.blue)
            self.button_dia_m.config(background=self.grey) 
        elif (self.unit_dia.get() == "m"):
            self.button_dia_nm.config(background=self.grey)
            self.button_dia_um.config(background=self.grey)
            self.button_dia_mm.config(background=self.grey)
            self.button_dia_cm.config(background=self.grey)
            self.button_dia_dm.config(background=self.grey)
            self.button_dia_m.config(background=self.blue) 
        else:
            pass             

    def unit_bground_temp(self):
        if (self.unit_temp.get() == "K"):
            self.button_temp_k.config(background=self.blue)
            self.button_temp_C.config(background=self.grey)
        elif (self.unit_temp.get() == "°C"):
            self.button_temp_k.config(background=self.grey)
            self.button_temp_C.config(background=self.blue)             
        else:
            pass            
  
    def unit_bground_mfr(self):
        if (self.unit_mfr.get() == "ug/s"):
            self.button_mfr_ugs.config(background=self.blue)
            self.button_mfr_mgs.config(background=self.grey)
            self.button_mfr_gs.config(background=self.grey)
            self.button_mfr_kgs.config(background=self.grey)
            self.button_mfr_kgh.config(background=self.grey)
        elif (self.unit_mfr.get() == "mg/s"):
            self.button_mfr_ugs.config(background=self.grey)
            self.button_mfr_mgs.config(background=self.blue)
            self.button_mfr_gs.config(background=self.grey)
            self.button_mfr_kgs.config(background=self.grey)
            self.button_mfr_kgh.config(background=self.grey)  
        elif (self.unit_mfr.get() == "g/s"):
            self.button_mfr_ugs.config(background=self.grey)
            self.button_mfr_mgs.config(background=self.grey)
            self.button_mfr_gs.config(background=self.blue)
            self.button_mfr_kgs.config(background=self.grey)
            self.button_mfr_kgh.config(background=self.grey)  
        elif (self.unit_mfr.get() == "kg/s"):
            self.button_mfr_ugs.config(background=self.grey)
            self.button_mfr_mgs.config(background=self.grey)
            self.button_mfr_gs.config(background=self.grey)
            self.button_mfr_kgs.config(background=self.blue)
            self.button_mfr_kgh.config(background=self.grey)  
        elif (self.unit_mfr.get() == "kg/h"):
            self.button_mfr_ugs.config(background=self.grey)
            self.button_mfr_mgs.config(background=self.grey)
            self.button_mfr_gs.config(background=self.grey)
            self.button_mfr_kgs.config(background=self.grey)
            self.button_mfr_kgh.config(background=self.blue)               
        else:
            pass       
    
    def unit_bground_mass(self):
        if (self.unit_mass.get() == "ng"):
            self.button_mass_ng.config(background=self.blue)
            self.button_mass_ug.config(background=self.grey)
            self.button_mass_mg.config(background=self.grey)
            self.button_mass_g.config(background=self.grey)
            self.button_mass_kg.config(background=self.grey)
        elif (self.unit_mass.get() == "ug"):
            self.button_mass_ng.config(background=self.grey)
            self.button_mass_ug.config(background=self.blue)
            self.button_mass_mg.config(background=self.grey)
            self.button_mass_g.config(background=self.grey)
            self.button_mass_kg.config(background=self.grey)
        elif (self.unit_mass.get() == "mg"):
            self.button_mass_ng.config(background=self.grey)
            self.button_mass_ug.config(background=self.grey)
            self.button_mass_mg.config(background=self.blue)
            self.button_mass_g.config(background=self.grey)
            self.button_mass_kg.config(background=self.grey)
        elif (self.unit_mass.get() == "g"):
            self.button_mass_ng.config(background=self.grey)
            self.button_mass_ug.config(background=self.grey)
            self.button_mass_mg.config(background=self.grey)
            self.button_mass_g.config(background=self.blue)
            self.button_mass_kg.config(background=self.grey)
        elif (self.unit_mass.get() == "kg"):
            self.button_mass_ng.config(background=self.grey)
            self.button_mass_ug.config(background=self.grey)
            self.button_mass_mg.config(background=self.grey)
            self.button_mass_g.config(background=self.grey)
            self.button_mass_kg.config(background=self.blue)
        else:
            pass

    def unit_bground_time(self):
        if (self.unit_time.get() == "ms"):
            self.button_time_ms.config(background=self.blue)
            self.button_time_s.config(background=self.grey)
            self.button_time_min.config(background=self.grey)
            self.button_time_h.config(background=self.grey)
            self.button_time_den.config(background=self.grey)
            
        elif (self.unit_time.get() == "s"):
            self.button_time_ms.config(background=self.grey)
            self.button_time_s.config(background=self.blue)
            self.button_time_min.config(background=self.grey)
            self.button_time_h.config(background=self.grey)
            self.button_time_den.config(background=self.grey)

        elif (self.unit_time.get() == "min"):
            self.button_time_ms.config(background=self.grey)
            self.button_time_s.config(background=self.grey)
            self.button_time_min.config(background=self.blue)
            self.button_time_h.config(background=self.grey)
            self.button_time_den.config(background=self.grey)

        elif (self.unit_time.get() == "h"):
            self.button_time_ms.config(background=self.grey)
            self.button_time_s.config(background=self.grey)
            self.button_time_min.config(background=self.grey)
            self.button_time_h.config(background=self.blue)
            self.button_time_den.config(background=self.grey)

        elif (self.unit_time.get() == "den"):
            self.button_time_ms.config(background=self.grey)
            self.button_time_s.config(background=self.grey)
            self.button_time_min.config(background=self.grey)
            self.button_time_h.config(background=self.grey)
            self.button_time_den.config(background=self.blue)

        else:
            pass
    #endregion





    def display(self):
        self.root.mainloop()
    
    def get_language(self):
        self.dpm_tabulka.language = self.gui_language.get()

    def spawn_tabulka(self):
        temp_spawn_tabulka_path = self.path_in.get().split("/")
        pozice = len(temp_spawn_tabulka_path)
        temp_spawn_tabulka_location = ""

        for i in range (0, len(temp_spawn_tabulka_path)-1):
            if i > 0:
                temp_spawn_tabulka_location += "/"
                temp_spawn_tabulka_location += temp_spawn_tabulka_path[i]
            else:
                temp_spawn_tabulka_location += temp_spawn_tabulka_path[i]

        self.dpm_file.set(temp_spawn_tabulka_path[-1])
        self.dpm_location.set(temp_spawn_tabulka_location)
        
        
        
        del(temp_spawn_tabulka_location)
        del(temp_spawn_tabulka_path)
        del(pozice)

        print(self.dpm_file.get())
        print(self.dpm_location.get())

        self.spawn_tabulka_toggle()

    def spawn_tabulka_toggle(self):
        if self.path_in.get() == "":
            self.button_browse_1_toggle.config(background=self.red, text= "Až vybereš, klikni.", state="normal")
        else:
            self.dpm_selected.set(True)
            if self.dpm_selected.get() == True:
                self.button_generate_source.config(state = "normal")
                self.button_browse_1_toggle.config(background=self.green, text= self.dpm_file.get(), state= "disabled")

    def generate_table(self):
        print(self.path_in.get())
        source = open(self.path_in.get(), "r")
        output = open(self.path_in.get() + "ptcl","w")
        for char in source.readlines()[1:]:
            output.write('\t'.join(
                                    char
                                    .replace("(", "")
                                    .replace(")","")
                                    .replace("\n","=")
                                    .split()
                                    )
                                    .replace("=", "\n"))
        source.close()
        output.close()

        if self.dpm_generated == False:
            self.button_generate_source.config(background=self.red, text= "Vygeneruj tabulku", state="normal")

        else:
            self.dpm_generated.set(True)
            if self.dpm_generated.get() == True:
                self.button_next.config(state = "normal")
                self.button_generate_source.config(background=self.green, text= "Tabulka připravena", state= "disabled")

                if (".ptcldpm" not in self.dpm_file.get()):
                    self.dpm_file.set(self.dpm_file.get() + "ptcl")
                else:
                    self.dpm_file.set(self.dpm_file.get())
                
                self.dpm_tabulka = tabulka(self.dpm_file.get()[:-8], self.dpm_location.get(), self.dpm_file.get(), self.unit_pos, self.unit_velo, self.unit_dia, self.unit_temp, self.unit_mfr, self.unit_mass, "1/s", self.unit_time)




    def unit_switch_global(self):
        self.button_next.config(state = "normal")
        self.button_switch_units_global.config(state = "disabled", background = self.green, text = "Jednotky změněny")
        
        print("\n\n\nNové jednotky budou:")
        print("Pozice: {}".format(self.unit_pos.get()))
        print("Rychlost: {}".format(self.unit_velo.get()))
        print("Průměr: {}".format(self.unit_dia.get()))
        print("MFR: {}".format(self.unit_mfr.get()))
        print("Teplota: {}".format(self.unit_temp.get()))
        print("Hmotnost: {}".format(self.unit_mass.get()))
        print("Frekvence: {}".format(self.unit_frequency.get()))
        print("Čas letu: {}".format(self.unit_time.get()))

        self.dpm_tabulka.unitswitch_pos(str(self.unit_pos.get()))
        self.dpm_tabulka.unitswitch_velo(str(self.unit_velo.get()))
        self.dpm_tabulka.unitswitch_diameter(str(self.unit_dia.get()))
        self.dpm_tabulka.unitswitch_mfr(str(self.unit_mfr.get()))
        self.dpm_tabulka.unitswitch_temperature(str(self.unit_temp.get()))
        self.dpm_tabulka.unitswitch_mass(str(self.unit_mass.get()))
        self.dpm_tabulka.unitswitch_frequency(str(self.unit_frequency.get()))
        self.dpm_tabulka.unitswitch_time(str(self.unit_time.get()))
        self.dpm_tabulka.update_units()





    def accept_DPI(self):
        numbers = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]       
        for char in self.text_input_DPI.get():
            if char in numbers:
                self.button_input_DPI.config(background=self.green, text = "Schváleno", state = "disabled")
                self.text_input_DPI.config(state = "disabled")
                self.dpm_tabulka.val_dpi = self.text_input_DPI.get()
            else:
                print("Toto není přirozené číslo")
                self.button_input_DPI.config(background=self.red, text = "Znovu! Jen přirozené číslo")
                self.text_input_DPI.config(state = "normal")
                self.text_input_DPI.delete(0, 'end')
                self.text_input_DPI.get()
                self.button_input_DPI.config(state = "normal")
                break

        if self.text_input_DPI.get() == "" or self.text_input_BINS.get() == "":
            self.button_next.config(state = "disabled")
        else:
            self.button_next.config(state = "normal")

    def accept_BINS(self):
        numbers = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]       
        for char in self.text_input_BINS.get():
            if char in numbers:
                self.button_input_BINS.config(background=self.green, text = "Schváleno", state = "disabled")
                self.text_input_BINS.config(state = "disabled")
                self.dpm_tabulka.val_bin_single = int(self.text_input_BINS.get())
            else:
                print("Toto není přirozené číslo")
                self.button_input_BINS.config(background=self.red, text = "Znovu! Jen přirozené číslo")
                self.text_input_BINS.config(state = "normal")
                self.text_input_BINS.delete(0, 'end')
                self.text_input_BINS.get()
                self.button_input_BINS.config(state = "normal")
                break

    def accept_xmin_histo(self):
        numbers = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "e", ".", "-"]       
        for char in self.text_input_histo_xmin.get(): 
            if char in numbers:
                self.button_histo_xmin_confirm.config(background=self.green, text = "Schváleno", state = "disabled")
                self.button_histo_xlim_reset.config(state = "normal")
                self.text_input_histo_xmin.config(state = "disabled")
            else:
                print("Toto není číslo")
                self.button_histo_xmin_confirm.config(background=self.red, text = "Znovu! Jen přirozené číslo")
                self.text_input_histo_xmin.config(state = "normal")
                self.button_histo_xlim_reset.config(state = "disabled")
                self.text_input_histo_xmin.delete(0, 'end')
                self.text_input_histo_xmin.get()
                self.button_histo_xmin_confirm.config(state = "normal")
                break

        if self.text_input_histo_xmin.get() == "":
            self.button_next.config(state = "disabled")
        else:
            self.button_next.config(state = "normal")
            self.gui_multihisto_xmin.set(float((self.text_input_histo_xmin.get())))
            self.dpm_tabulka.xmin_histo                = float((self.text_input_histo_xmin.get()))

        if self.button_histo_xmin_confirm["background"] == self.green and self.button_histo_xmax_confirm["background"] == self.green:
            self.dpm_tabulka.xlim_histo_bool = True

    def accept_xmax_histo(self):
        numbers = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "e", ".", "-"]       
        for char in self.text_input_histo_xmax.get(): 
            if char in numbers:
                self.button_histo_xmax_confirm.config(background=self.green, text = "Schváleno", state = "disabled")
                self.button_histo_xlim_reset.config(state = "normal")
                self.text_input_histo_xmax.config(state = "disabled")
            else:
                print("Toto není číslo")
                self.button_histo_xmax_confirm.config(background=self.red, text = "Znovu! Jen přirozené číslo")
                self.text_input_histo_xmax.config(state = "normal")
                self.button_histo_xlim_reset.config(state = "disabled")
                self.text_input_histo_xmax.delete(0, 'end')
                self.text_input_histo_xmax.get()
                self.button_histo_xmax_confirm.config(state = "normal")
                break

        if self.text_input_histo_xmax.get() == "":
            self.button_next.config(state = "disabled")
        else:
            self.button_next.config(state = "normal")
            self.gui_multihisto_xmax.set(float((self.text_input_histo_xmax.get())))
            self.dpm_tabulka.xmax_histo                = float((self.text_input_histo_xmax.get()))

        if self.button_histo_xmin_confirm["background"] == self.green and self.button_histo_xmax_confirm["background"] == self.green:
            self.dpm_tabulka.xlim_histo_bool = True

    def accept_xlim_histo(self):
        print("ACCEPT HISTO XLIMITS: ... pre-var: {}".format(self.dpm_tabulka.histo_xlim_bool))
        if self.gui_histo_bool.get() == False:
            self.gui_histo_bool.set(True)
            self.text_input_histo_xmin.config(state = "disabled")
            self.text_input_histo_xmax.config(state = "disabled")
            self.button_histo_xlim_reset.config(state = "disabled")
            self.button_histo_xmin_confirm.config(state = "disabled")
            self.button_histo_xmax_confirm.config(state = "disabled")
            self.button_histo_xlim_allow.config(state = "normal", background= self.green, text = "Automat")
            self.dpm_tabulka.histo_xlim_bool = True
        else:
            self.gui_histo_bool.set(False)
            if (self.button_histo_xmin_confirm["background"] != self.green and self.button_histo_xmax_confirm["background"] != self.green):
                self.text_input_histo_xmin.config(state = "normal")
                self.button_histo_xmin_confirm.config(state = "normal")
                self.text_input_histo_xmax.config(state = "normal")
                self.button_histo_xmax_confirm.config(state = "normal")

            self.button_histo_xlim_reset.config(state = "normal")
            self.button_histo_xlim_allow.config(state = "normal", background= self.grey, text = "Manuál")
        
        self.dpm_tabulka.histo_xlim_bool =not( self.gui_histo_bool.get() )
        print("ACCEPT HISTO XLIMITS: ... post-var: {}".format(self.dpm_tabulka.xlim_histo_bool))
        print("Histo xlim bool? {}".format(self.gui_histo_bool.get()))

    def histo_enabler(self):
        condition_manual = (self.button_histo_xlim_allow["background"] == self.grey) 
        condition_automat = (self.button_histo_xlim_allow["background"] == self.green)
        condition_buttons_ranbins = (self.button_input_BINS["background"] == self.green)
        condition_buttons_xlims = (self.button_histo_xmin_confirm["background"] == self.green) and (self.button_histo_xmax_confirm["background"] == self.green)

        if (condition_automat and condition_buttons_ranbins):
            self.button_histo_posx.config(state = "normal")
            self.button_histo_posy.config(state = "normal")
            self.button_histo_posz.config(state = "normal")
            self.button_histo_vx.config(state = "normal")
            self.button_histo_vy.config(state = "normal")
            self.button_histo_vz.config(state = "normal")
            self.button_histo_vmag.config(state = "normal")
            self.button_histo_diameter.config(state = "normal")
            self.button_histo_temperature.config(state = "normal")
            self.button_histo_mfr.config(state = "normal")
            self.button_histo_mass.config(state = "normal")
            self.button_histo_time.config(state = "normal")

        elif (condition_manual and condition_buttons_ranbins and condition_buttons_xlims):
            self.button_histo_posx.config(state = "normal")
            self.button_histo_posy.config(state = "normal")
            self.button_histo_posz.config(state = "normal")
            self.button_histo_vx.config(state = "normal")
            self.button_histo_vy.config(state = "normal")
            self.button_histo_vz.config(state = "normal")
            self.button_histo_vmag.config(state = "normal")
            self.button_histo_diameter.config(state = "normal")
            self.button_histo_temperature.config(state = "normal")
            self.button_histo_mfr.config(state = "normal")
            self.button_histo_mass.config(state = "normal")
            self.button_histo_time.config(state = "normal")

        else:
            self.button_histo_posx.config(state = "disabled")
            self.button_histo_posy.config(state = "disabled")
            self.button_histo_posz.config(state = "disabled")
            self.button_histo_vx.config(state = "disabled")
            self.button_histo_vy.config(state = "disabled")
            self.button_histo_vz.config(state = "disabled")
            self.button_histo_vmag.config(state = "disabled")
            self.button_histo_diameter.config(state = "disabled")
            self.button_histo_temperature.config(state = "disabled")
            self.button_histo_mfr.config(state = "disabled")
            self.button_histo_mass.config(state = "disabled")
            self.button_histo_time.config(state = "disabled")
            self.button_create_histos.config(state = "disabled")



    def accept_BINS_multihistos(self):
        numbers = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]       
        for char in self.text_input_multihistos_bins.get(): 
            if char in numbers:
                self.button_multihistos_bins_confirm.config(background=self.green, text = "Schváleno", state = "disabled")
                self.button_multihistos_bins_reset.config(state = "normal")
                self.text_input_multihistos_bins.config(state = "disabled")
            else:
                print("Toto není přirozené číslo")
                self.button_multihistos_bins_confirm.config(background=self.red, text = "Znovu! Jen přirozené číslo")
                self.text_input_multihistos_bins.config(state = "normal")
                self.button_multihistos_bins_reset.config(state = "disabled")
                self.text_input_multihistos_bins.delete(0, 'end')
                self.text_input_multihistos_bins.get()
                self.button_multihistos_bins_confirm.config(state = "normal")
                break

        if self.text_input_multihistos_bins.get() == "":
            self.button_next.config(state = "disabled")
        else:
            self.button_next.config(state = "normal")
            self.dpm_tabulka.val_bin_multi              = int((self.text_input_multihistos_bins.get()))

    def accept_ranges_multihistos(self):
        numbers = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]       
        for char in self.text_input_multihistos_ranges.get(): 
            if char in numbers:
                self.button_multihistos_ranges_confirm.config(background=self.green, text = "Schváleno", state = "disabled")
                self.button_multihistos_ranges_reset.config(state = "normal")
                self.text_input_multihistos_ranges.config(state = "disabled")
            else:
                print("Toto není přirozené číslo")
                self.button_multihistos_ranges_confirm.config(background=self.red, text = "Znovu! Jen přirozené číslo")
                self.text_input_multihistos_ranges.config(state = "normal")
                self.button_multihistos_ranges_reset.config(state = "disabled")
                self.text_input_multihistos_ranges.delete(0, 'end')
                self.text_input_multihistos_ranges.get()
                self.button_multihistos_ranges_confirm.config(state = "normal")
                break

        if self.text_input_multihistos_ranges.get() == "":
            self.button_next.config(state = "disabled")
        else:
            self.button_next.config(state = "normal")
            self.gui_range_count.set(int((self.text_input_multihistos_ranges.get())))
            self.dpm_tabulka.range_count                = int((self.text_input_multihistos_ranges.get()))

    def accept_xmin_multihistos(self):
        numbers = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "e", ".", "-"]       
        for char in self.text_input_multihistos_xmin.get(): 
            if char in numbers:
                self.button_multihistos_xmin_confirm.config(background=self.green, text = "Schváleno", state = "disabled")
                self.button_multihistos_xlim_reset.config(state = "normal")
                self.text_input_multihistos_xmin.config(state = "disabled")
            else:
                print("Toto není číslo")
                self.button_multihistos_xmin_confirm.config(background=self.red, text = "Znovu! Jen přirozené číslo")
                self.text_input_multihistos_xmin.config(state = "normal")
                self.button_multihistos_xlim_reset.config(state = "disabled")
                self.text_input_multihistos_xmin.delete(0, 'end')
                self.text_input_multihistos_xmin.get()
                self.button_multihistos_xmin_confirm.config(state = "normal")
                break

        if self.text_input_multihistos_xmin.get() == "":
            self.button_next.config(state = "disabled")
        else:
            self.button_next.config(state = "normal")
            self.gui_multihisto_xmin.set(float((self.text_input_multihistos_xmin.get())))
            self.dpm_tabulka.xmin_multihisto                = float((self.text_input_multihistos_xmin.get()))

        if self.button_multihistos_xmin_confirm["background"] == self.green and self.button_multihistos_xmax_confirm["background"] == self.green:
            self.dpm_tabulka.xlim_bool = True

    def accept_xmax_multihistos(self):
        numbers = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "e", ".", "-"]       
        for char in self.text_input_multihistos_xmax.get(): 
            if char in numbers:
                self.button_multihistos_xmax_confirm.config(background=self.green, text = "Schváleno", state = "disabled")
                self.button_multihistos_xlim_reset.config(state = "normal")
                self.text_input_multihistos_xmax.config(state = "disabled")
            else:
                print("Toto není číslo")
                self.button_multihistos_xmax_confirm.config(background=self.red, text = "Znovu! Jen přirozené číslo")
                self.text_input_multihistos_xmax.config(state = "normal")
                self.button_multihistos_xlimreset.config(state = "disabled")
                self.text_input_multihistos_xmax.delete(0, 'end')
                self.text_input_multihistos_xmax.get()
                self.button_multihistos_xmax_confirm.config(state = "normal")
                break

        if self.text_input_multihistos_xmax.get() == "":
            self.button_next.config(state = "disabled")
        else:
            self.button_next.config(state = "normal")
            self.gui_multihisto_xmax.set(float((self.text_input_multihistos_xmax.get())))
            self.dpm_tabulka.xmax_multihisto                = float((self.text_input_multihistos_xmax.get()))

        if self.button_multihistos_xmin_confirm["background"] == self.green and self.button_multihistos_xmax_confirm["background"] == self.green:
            self.dpm_tabulka.xlim_bool = True

    def accept_xlim_multihistos(self):
        if self.gui_multihisto_bool.get() == False:
            self.gui_multihisto_bool.set(True)
            self.text_input_multihistos_xmin.config(state = "disabled")
            self.text_input_multihistos_xmax.config(state = "disabled")
            self.button_multihistos_xlim_reset.config(state = "disabled")
            self.button_multihistos_xmin_confirm.config(state = "disabled")
            self.button_multihistos_xmax_confirm.config(state = "disabled")
            self.button_multihistos_xlim_allow.config(state = "normal", background= self.green, text = "Automat")
            self.dpm_tabulka.xlim_bool = False
        else:
            self.gui_multihisto_bool.set(False)
            if (self.button_multihistos_xmin_confirm["background"] != self.green and self.button_multihistos_xmax_confirm["background"] != self.green):
                self.text_input_multihistos_xmin.config(state = "normal")
                self.button_multihistos_xmin_confirm.config(state = "normal")
                self.text_input_multihistos_xmax.config(state = "normal")
                self.button_multihistos_xmax_confirm.config(state = "normal")

            self.button_multihistos_xlim_reset.config(state = "normal")
            self.button_multihistos_xlim_allow.config(state = "normal", background= self.grey, text = "Manuál")
            self.dpm_tabulka.xlim_bool = True
        print("Multihisto xlim bool? {}".format(self.gui_multihisto_bool.get()))
        print("Multihisto xlim zapnuto? {}".format(self.dpm_tabulka.xlim_bool))

    def accept_percent(self):
        if self.plot_percent.get() == True:
            self.button_plot_percent.config(state = "disabled", text = "[%]", background = self.blue)
            self.button_plot_nonpercent.config(state = "normal", text = "[-]", background = self.grey)

        else:
            self.button_plot_percent.config(state = "normal", text = "[%]", background = self.grey)
            self.button_plot_nonpercent.config(state = "disabled", text = "[-]", background = self.blue)

        print("Accept percent: {}".format(self.plot_percent.get()))
    
    def reset_options(self):
        self.text_input_DPI.config(state = "normal")
        self.button_input_BINS.config(background=self.grey, text = "Potvrď", state = "normal")
        self.text_input_BINS.config(state = "normal")
        self.button_input_DPI.config(background=self.grey, text = "Potvrď", state = "normal")

        self.plot_percent.set(True)
        self.button_plot_percent.config(state = "disabled", text = "[%]", background = self.blue)
        self.button_plot_nonpercent.config(state = "normal", text = "[-]", background = self.grey)

        if self.text_input_DPI.get() == "" or self.text_input_BINS.get() == "":
            self.button_next.config(state = "disabled")
        else:
            self.button_next.config(state = "normal")
        self.text_input_BINS.delete(0, 'end')
        self.text_input_DPI.delete(0, 'end')

        self.button_histo_posx.config(state = "normal", background = self.grey)
        self.histo_posx.set(False)
        self.button_histo_posy.config(state = "normal", background = self.grey)
        self.histo_posy.set(False)
        self.button_histo_posz.config(state = "normal", background = self.grey)
        self.histo_posz.set(False)

        self.button_histo_vx.config(state = "normal", background = self.grey)
        self.histo_vx.set(False)
        self.button_histo_vy.config(state = "normal", background = self.grey)
        self.histo_vy.set(False)
        self.button_histo_vz.config(state = "normal", background = self.grey)
        self.histo_vz.set(False)
        self.button_histo_vmag.config(state = "normal", background = self.grey)
        self.histo_vmag.set(False)

        self.button_histo_diameter.config(state = "normal", background = self.grey)
        self.histo_diameter.set(False)
        self.button_histo_temperature.config(state = "normal", background = self.grey)
        self.histo_temperature.set(False)
        self.button_histo_mfr.config(state = "normal", background = self.grey)
        self.histo_mfr.set(False)
        self.button_histo_mass.config(state = "normal", background = self.grey)
        self.histo_mass.set(False)
        self.button_histo_time.config(state = "normal", background = self.grey)
        self.histo_time.set(False)
        self.button_browse_2.config(text = "Vyber složku", state = "normal", background = self.grey)
        self.path_out.set("")

        self.text_input_histo_xmin.delete(0, "end")
        self.text_input_histo_xmax.delete(0, "end")
        self.button_histo_xmin_confirm.config(state = "disabled", background = self.grey, text = "potvrď")
        self.button_histo_xmax_confirm.config(state = "disabled", background = self.grey, text = "potvrď")
        self.button_histo_xlim_reset.config(state = "disabled", background = self.red)
        self.button_histo_xlim_allow.config(state = "normal", background = self.green, text = "Automat")
        self.gui_histo_bool.set(True)
        self.dpm_tabulka.histo_xlim_bool= False


        self.button_create_histos.config(state = "disabled", background= self.grey, text = "Generuj obrázky")
        self.page_3_enabler()
        self.histo_enabler()



    #region histogram_selection
    def page_3_enabler(self):
        self.printout_bools()
        condition_pos   = self.histo_posx.get()     == False and self.histo_posy.get()  == False and self.histo_posz.get()  == False
        condition_velo  = self.histo_vx.get()       == False and self.histo_vy.get()    == False and self.histo_vz.get()    == False and self.histo_vmag.get()  == False     
        condition_prop  = self.histo_diameter.get()      == False and self.histo_temperature.get()  == False and self.histo_mfr.get()   == False and self.histo_mass.get()  == False and self.histo_time.get() == False   
        condition_singlehisto_ticks =   condition_pos and condition_velo and condition_prop

        condition_vals  = self.button_input_DPI["background"] == self.green and self.button_input_BINS["background"] == self.green

        if condition_vals:
            self.button_browse_2.config(state = "normal")
            if condition_singlehisto_ticks:
                pass
            else:
                self.button_create_histos.config(state = "normal")

        else:
            self.button_browse_2.config(state = "disabled")
            self.button_create_histos.config(state = "disabled")
        
    def select_histogram_posx(self):
        if (self.histo_posx.get() == False):
            self.button_histo_posx.config(state = "normal", background = self.blue)
            self.histo_posx.set(True)
        else:
            self.button_histo_posx.config(state = "normal", background = self.grey)
            self.histo_posx.set(False)
        self.page_3_enabler()           

    def select_histogram_posy(self):
        if (self.histo_posy.get() == False):
            self.button_histo_posy.config(state = "normal", background = self.blue)
            self.histo_posy.set(True)
        else:
            self.button_histo_posy.config(state = "normal", background = self.grey)
            self.histo_posy.set(False) 
        self.page_3_enabler()           

    def select_histogram_posz(self):
        if (self.histo_posz.get() == False):
            self.button_histo_posz.config(state = "normal", background = self.blue)
            self.histo_posz.set(True)
        else:
            self.button_histo_posz.config(state = "normal", background = self.grey)
            self.histo_posz.set(False) 
        self.page_3_enabler() 

    def select_histogram_vx(self):
        if (self.histo_vx.get() == False):
            self.button_histo_vx.config(state = "normal", background = self.blue)
            self.histo_vx.set(True)
        else:
            self.button_histo_vx.config(state = "normal", background = self.grey)
            self.histo_vx.set(False)           
        self.page_3_enabler() 

    def select_histogram_vy(self):
        if (self.histo_vy.get() == False):
            self.button_histo_vy.config(state = "normal", background = self.blue)
            self.histo_vy.set(True)
        else:
            self.button_histo_vy.config(state = "normal", background = self.grey)
            self.histo_vy.set(False) 
        self.page_3_enabler() 

    def select_histogram_vz(self):
        if (self.histo_vz.get() == False):
            self.button_histo_vz.config(state = "normal", background = self.blue)
            self.histo_vz.set(True)
        else:
            self.button_histo_vz.config(state = "normal", background = self.grey)
            self.histo_vz.set(False) 
        self.page_3_enabler() 

    def select_histogram_vmag(self):
        if (self.histo_vmag.get() == False):
            self.button_histo_vmag.config(state = "normal", background = self.blue)
            self.histo_vmag.set(True)
        else:
            self.button_histo_vmag.config(state = "normal", background = self.grey)
            self.histo_vmag.set(False)
        self.page_3_enabler() 

    def select_histogram_diameter(self):
        if (self.histo_diameter.get() == False):
            self.button_histo_diameter.config(state = "normal", background = self.blue)
            self.histo_diameter.set(True)
        else:
            self.button_histo_diameter.config(state = "normal", background = self.grey)
            self.histo_diameter.set(False)
        self.page_3_enabler() 

    def select_histogram_temperature(self):
        if (self.histo_temperature.get() == False):
            self.button_histo_temperature.config(state = "normal", background = self.blue)
            self.histo_temperature.set(True)
        else:
            self.button_histo_temperature.config(state = "normal", background = self.grey)
            self.histo_temperature.set(False)
        self.page_3_enabler() 

    def select_histogram_mfr(self):
        if (self.histo_mfr.get() == False):
            self.button_histo_mfr.config(state = "normal", background = self.blue)
            self.histo_mfr.set(True)
        else:
            self.button_histo_mfr.config(state = "normal", background = self.grey)
            self.histo_mfr.set(False)
        self.page_3_enabler()

    def select_histogram_mass(self):
        if (self.histo_mass.get() == False):
            self.button_histo_mass.config(state = "normal", background = self.blue)
            self.histo_mass.set(True)
        else:
            self.button_histo_mass.config(state = "normal", background = self.grey)
            self.histo_mass.set(False)
        self.page_3_enabler() 

    def select_histogram_time(self):
        if (self.histo_time.get() == False):
            self.button_histo_time.config(state = "normal", background = self.blue)
            self.histo_time.set(True)
        else:
            self.button_histo_time.config(state = "normal", background = self.grey)
            self.histo_time.set(False)
        self.page_3_enabler() 
    #endregion

    def printout_bools(self):
        print(2*"\n" + "Page {} bools".format(self.page_number.get()))
        print("Histo PosX: {}".format(self.histo_posx.get()))
        print("Histo PosY: {}".format(self.histo_posy.get()))
        print("Histo PosZ: {}".format(self.histo_posz.get()))
        print("Histo VeloX: {}".format(self.histo_vx.get()))
        print("Histo VeloY: {}".format(self.histo_vy.get()))
        print("Histo VeloZ: {}".format(self.histo_vz.get()))
        print("Histo VeloMAG: {}".format(self.histo_vmag.get()))
        print("Histo Dia: {}".format(self.histo_diameter.get()))
        print("Histo Temp: {}".format(self.histo_temperature.get()))
        print("Histo MFR: {}".format(self.histo_mfr.get()))
        print("Histo Mass: {}".format(self.histo_mass.get()))
        print("Histo Time: {}".format(self.histo_time.get()))




    def create_histo_pictures(self):
        self.dpm_tabulka.val_dpi = int((self.text_input_DPI.get()))
        self.dpm_tabulka.val_bin = int((self.text_input_BINS.get()))
        self.dpm_tabulka.workdir = str(self.path_out.get())
        self.dpm_tabulka.bool_distribution_unit = bool(self.plot_percent.get())
        print("Create_histo_pictures procento? {}".format(self.plot_percent.get()))

        if (self.histo_posx.get() == True):
            self.dpm_tabulka.dpmhisto_single(self.dpm_tabulka.x)
        
        if (self.histo_posy.get() == True):
            self.dpm_tabulka.dpmhisto_single(self.dpm_tabulka.y)

        if (self.histo_posz.get() == True):
            self.dpm_tabulka.dpmhisto_single(self.dpm_tabulka.z)

        if (self.histo_vx.get() == True):
            self.dpm_tabulka.dpmhisto_single(self.dpm_tabulka.vx)

        if (self.histo_vy.get() == True):
            self.dpm_tabulka.dpmhisto_single(self.dpm_tabulka.vy)

        if (self.histo_vz.get() == True):
            self.dpm_tabulka.dpmhisto_single(self.dpm_tabulka.vz)

        if (self.histo_vmag.get() == True):
            self.dpm_tabulka.dpmhisto_single(self.dpm_tabulka.vmag)

        if (self.histo_diameter.get() == True):
            self.dpm_tabulka.dpmhisto_single(self.dpm_tabulka.diameter) 

        if (self.histo_temperature.get() == True):
            self.dpm_tabulka.dpmhisto_single(self.dpm_tabulka.temperature)
                                             
        if (self.histo_mfr.get() == True):
            self.dpm_tabulka.dpmhisto_single(self.dpm_tabulka.mfr)

        if (self.histo_mass.get() == True):
            self.dpm_tabulka.dpmhisto_single(self.dpm_tabulka.mass)

        if (self.histo_time.get() == True):
            self.dpm_tabulka.dpmhisto_single(self.dpm_tabulka.time)

        self.button_create_histos.config(state = "normal", background= self.green, text = "VYGENEROVÁNO")

    def turn_off_histo_col_1(self):
        self.multihistos_colormanager1()
        self.bool_multihisto_x1.set(False)
        self.bool_multihisto_y1.set(False)
        self.bool_multihisto_z1.set(False)
        self.bool_multihisto_vx1.set(False)
        self.bool_multihisto_vy1.set(False)
        self.bool_multihisto_vz1.set(False)
        self.bool_multihisto_vmag1.set(False)
        self.bool_multihisto_diameter1.set(False)
        self.bool_multihisto_temperature1.set(False)
        self.bool_multihisto_mfr1.set(False)
        self.bool_multihisto_mass1.set(False)
        self.bool_multihisto_time1.set(False)

    def turn_off_histo_col_2(self):
        self.multihistos_colormanager2()
        self.bool_multihisto_x2.set(False)
        self.bool_multihisto_y2.set(False)
        self.bool_multihisto_z2.set(False)
        self.bool_multihisto_vx2.set(False)
        self.bool_multihisto_vy2.set(False)
        self.bool_multihisto_vz2.set(False)
        self.bool_multihisto_vmag2.set(False)
        self.bool_multihisto_diameter2.set(False)
        self.bool_multihisto_temperature2.set(False)
        self.bool_multihisto_mfr2.set(False)
        self.bool_multihisto_mass2.set(False)
        self.bool_multihisto_time2.set(False)
 


    def multihistos_colormanager1(self):
        self.button_multihistos_x.config(state = "normal", background = self.grey)
        self.button_multihistos_y.config(state = "normal", background = self.grey)
        self.button_multihistos_z.config(state = "normal", background = self.grey)
        self.button_multihistos_vx.config(state = "normal", background = self.grey)
        self.button_multihistos_vy.config(state = "normal", background = self.grey)
        self.button_multihistos_vz.config(state = "normal", background = self.grey)
        self.button_multihistos_vmag.config(state = "normal", background = self.grey)
        self.button_multihistos_diameter.config(state = "normal", background = self.grey)
        self.button_multihistos_temperature.config(state = "normal", background = self.grey)
        self.button_multihistos_mfr.config(state = "normal", background = self.grey)
        self.button_multihistos_mass.config(state = "normal", background = self.grey)
        self.button_multihistos_time.config(state = "normal", background = self.grey)

    def multihistos_colormanager2(self):
        self.button_multihistos_x2.config(state = "normal", background = self.grey)
        self.button_multihistos_y2.config(state = "normal", background = self.grey)
        self.button_multihistos_z2.config(state = "normal", background = self.grey)
        self.button_multihistos_vx2.config(state = "normal", background = self.grey)
        self.button_multihistos_vy2.config(state = "normal", background = self.grey)
        self.button_multihistos_vz2.config(state = "normal", background = self.grey)
        self.button_multihistos_vmag2.config(state = "normal", background = self.grey)
        self.button_multihistos_diameter2.config(state = "normal", background = self.grey)
        self.button_multihistos_temperature2.config(state = "normal", background = self.grey)
        self.button_multihistos_mfr2.config(state = "normal", background = self.grey)
        self.button_multihistos_mass2.config(state = "normal", background = self.grey)
        self.button_multihistos_time2.config(state = "normal", background = self.grey)

    def multihisto_enabler(self):
        condition_manual = (self.button_multihistos_xlim_allow["background"] == self.grey) 
        condition_automat = (self.button_multihistos_xlim_allow["background"] == self.green)
        condition_buttons_ranbins = (self.button_multihistos_bins_confirm["background"] == self.green) and (self.button_multihistos_ranges_confirm["background"] == self.green)
        condition_buttons_xlims = (self.button_multihistos_xmin_confirm["background"] == self.green) and (self.button_multihistos_xmax_confirm["background"] == self.green)

        if (condition_automat and condition_buttons_ranbins):
            self.button_multihistos_x2.config(state = "normal")
            self.button_multihistos_y2.config(state = "normal")
            self.button_multihistos_z2.config(state = "normal")
            self.button_multihistos_vx2.config(state = "normal")
            self.button_multihistos_vy2.config(state = "normal")
            self.button_multihistos_vz2.config(state = "normal")
            self.button_multihistos_vmag2.config(state = "normal")
            self.button_multihistos_diameter2.config(state = "normal")
            self.button_multihistos_temperature2.config(state = "normal")
            self.button_multihistos_mfr2.config(state = "normal")
            self.button_multihistos_mass2.config(state = "normal")
            self.button_multihistos_time2.config(state = "normal")
            self.label_multihistos_tmp.grid_forget()

        elif (condition_manual and condition_buttons_ranbins and condition_buttons_xlims):
            self.button_multihistos_x2.config(state = "normal")
            self.button_multihistos_y2.config(state = "normal")
            self.button_multihistos_z2.config(state = "normal")
            self.button_multihistos_vx2.config(state = "normal")
            self.button_multihistos_vy2.config(state = "normal")
            self.button_multihistos_vz2.config(state = "normal")
            self.button_multihistos_vmag2.config(state = "normal")
            self.button_multihistos_diameter2.config(state = "normal")
            self.button_multihistos_temperature2.config(state = "normal")
            self.button_multihistos_mfr2.config(state = "normal")
            self.button_multihistos_mass2.config(state = "normal")
            self.button_multihistos_time2.config(state = "normal")
            self.label_multihistos_tmp.grid_forget()

        else:
            self.button_multihistos_x2.config(state = "disabled")
            self.button_multihistos_y2.config(state = "disabled")
            self.button_multihistos_z2.config(state = "disabled")
            self.button_multihistos_vx2.config(state = "disabled")
            self.button_multihistos_vy2.config(state = "disabled")
            self.button_multihistos_vz2.config(state = "disabled")
            self.button_multihistos_vmag2.config(state = "disabled")
            self.button_multihistos_diameter2.config(state = "disabled")
            self.button_multihistos_temperature2.config(state = "disabled")
            self.button_multihistos_mfr2.config(state = "disabled")
            self.button_multihistos_mass2.config(state = "disabled")
            self.button_multihistos_time2.config(state = "disabled")



    def scatter_col1_reset(self):
        self.button_scatter_x1.config(state = "normal", background = self.grey)
        self.button_scatter_y1.config(state = "normal", background = self.grey)
        self.button_scatter_z1.config(state = "normal", background = self.grey)
        self.button_scatter_vx1.config(state = "normal", background = self.grey)
        self.button_scatter_vy1.config(state = "normal", background = self.grey)
        self.button_scatter_vz1.config(state = "normal", background = self.grey)
        self.button_scatter_vmag1.config(state = "normal", background = self.grey)
        self.button_scatter_diameter1.config(state = "normal", background = self.grey)
        self.button_scatter_temperature1.config(state = "normal", background = self.grey)
        self.button_scatter_mfr1.config(state = "normal", background = self.grey)
        self.button_scatter_mass1.config(state = "normal", background = self.grey)
        self.button_scatter_time1.config(state = "normal", background = self.grey)
    
    def scatter_col2_reset(self):
        self.button_scatter_x2.config(state = "normal", background = self.grey)
        self.button_scatter_y2.config(state = "normal", background = self.grey)
        self.button_scatter_z2.config(state = "normal", background = self.grey)
        self.button_scatter_vx2.config(state = "normal", background = self.grey)
        self.button_scatter_vy2.config(state = "normal", background = self.grey)
        self.button_scatter_vz2.config(state = "normal", background = self.grey)
        self.button_scatter_vmag2.config(state = "normal", background = self.grey)
        self.button_scatter_diameter2.config(state = "normal", background = self.grey)
        self.button_scatter_temperature2.config(state = "normal", background = self.grey)
        self.button_scatter_mfr2.config(state = "normal", background = self.grey)
        self.button_scatter_mass2.config(state = "normal", background = self.grey)
        self.button_scatter_time2.config(state = "normal", background = self.grey)    
    
    def scatter_col3_reset(self):
        self.button_scatter_none3.config(state = "normal", background = self.grey)
        self.button_scatter_x3.config(state = "normal", background = self.grey)
        self.button_scatter_y3.config(state = "normal", background = self.grey)
        self.button_scatter_z3.config(state = "normal", background = self.grey)
        self.button_scatter_vx3.config(state = "normal", background = self.grey)
        self.button_scatter_vy3.config(state = "normal", background = self.grey)
        self.button_scatter_vz3.config(state = "normal", background = self.grey)
        self.button_scatter_vmag3.config(state = "normal", background = self.grey)
        self.button_scatter_diameter3.config(state = "normal", background = self.grey)
        self.button_scatter_temperature3.config(state = "normal", background = self.grey)
        self.button_scatter_mfr3.config(state = "normal", background = self.grey)
        self.button_scatter_mass3.config(state = "normal", background = self.grey)
        self.button_scatter_time3.config(state = "normal", background = self.grey)    
    
    #region multihistos_plotters
    def multihistos_filter_x2(self):
        if self.bool_multihisto_x2.get() == True:
            if self.bool_multihisto_x1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.x, self.gui_range_count.get(), self.dpm_tabulka.x))
            elif self.bool_multihisto_y1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.y, self.gui_range_count.get(), self.dpm_tabulka.x))
            elif self.bool_multihisto_z1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.z, self.gui_range_count.get(), self.dpm_tabulka.x))
            elif self.bool_multihisto_vx1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vx, self.gui_range_count.get(), self.dpm_tabulka.x))
            elif self.bool_multihisto_vy1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vy, self.gui_range_count.get(), self.dpm_tabulka.x))
            elif self.bool_multihisto_vz1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vz, self.gui_range_count.get(), self.dpm_tabulka.x))
            elif self.bool_multihisto_vmag1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vmag, self.gui_range_count.get(), self.dpm_tabulka.x))
            elif self.bool_multihisto_diameter1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.diameter, self.gui_range_count.get(), self.dpm_tabulka.x))
            elif self.bool_multihisto_temperature1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.temperature, self.gui_range_count.get(), self.dpm_tabulka.x))
            elif self.bool_multihisto_mfr1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mfr, self.gui_range_count.get(), self.dpm_tabulka.x))
            elif self.bool_multihisto_mass1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mass, self.gui_range_count.get(), self.dpm_tabulka.x))            
            elif self.bool_multihisto_time1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.time, self.gui_range_count.get(), self.dpm_tabulka.x))      
            else:
                pass

    def multihistos_filter_y2(self):
        if self.bool_multihisto_y2.get() == True:
            if self.bool_multihisto_x1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.x, self.gui_range_count.get(), self.dpm_tabulka.y))
            elif self.bool_multihisto_y1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.y, self.gui_range_count.get(), self.dpm_tabulka.y))
            elif self.bool_multihisto_z1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.z, self.gui_range_count.get(), self.dpm_tabulka.y))
            elif self.bool_multihisto_vx1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vx, self.gui_range_count.get(), self.dpm_tabulka.y))
            elif self.bool_multihisto_vy1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vy, self.gui_range_count.get(), self.dpm_tabulka.y))
            elif self.bool_multihisto_vz1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vz, self.gui_range_count.get(), self.dpm_tabulka.y))
            elif self.bool_multihisto_vmag1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vmag, self.gui_range_count.get(), self.dpm_tabulka.y))
            elif self.bool_multihisto_diameter1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.diameter, self.gui_range_count.get(), self.dpm_tabulka.y))
            elif self.bool_multihisto_temperature1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.temperature, self.gui_range_count.get(), self.dpm_tabulka.y))
            elif self.bool_multihisto_mfr1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mfr, self.gui_range_count.get(), self.dpm_tabulka.y))
            elif self.bool_multihisto_mass1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mass, self.gui_range_count.get(), self.dpm_tabulka.y))            
            elif self.bool_multihisto_time1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.time, self.gui_range_count.get(), self.dpm_tabulka.y))      
            else:
                pass

    def multihistos_filter_z2(self):
        if self.bool_multihisto_z2.get() == True:
            if self.bool_multihisto_x1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.x, self.gui_range_count.get(), self.dpm_tabulka.z))
            elif self.bool_multihisto_y1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.y, self.gui_range_count.get(), self.dpm_tabulka.z))
            elif self.bool_multihisto_z1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.z, self.gui_range_count.get(), self.dpm_tabulka.z))
            elif self.bool_multihisto_vx1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vx, self.gui_range_count.get(), self.dpm_tabulka.z))
            elif self.bool_multihisto_vy1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vy, self.gui_range_count.get(), self.dpm_tabulka.z))
            elif self.bool_multihisto_vz1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vz, self.gui_range_count.get(), self.dpm_tabulka.z))
            elif self.bool_multihisto_vmag1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vmag, self.gui_range_count.get(), self.dpm_tabulka.z))
            elif self.bool_multihisto_diameter1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.diameter, self.gui_range_count.get(), self.dpm_tabulka.z))
            elif self.bool_multihisto_temperature1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.temperature, self.gui_range_count.get(), self.dpm_tabulka.z))
            elif self.bool_multihisto_mfr1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mfr, self.gui_range_count.get(), self.dpm_tabulka.z))
            elif self.bool_multihisto_mass1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mass, self.gui_range_count.get(), self.dpm_tabulka.z))            
            elif self.bool_multihisto_time1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.time, self.gui_range_count.get(), self.dpm_tabulka.z))      
            else:
                pass

    def multihistos_filter_vx2(self):
        if self.bool_multihisto_vx2.get() == True:
            if self.bool_multihisto_x1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.x, self.gui_range_count.get(), self.dpm_tabulka.vx))
            elif self.bool_multihisto_y1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.y, self.gui_range_count.get(), self.dpm_tabulka.vx))
            elif self.bool_multihisto_z1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.z, self.gui_range_count.get(), self.dpm_tabulka.vx))
            elif self.bool_multihisto_vx1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vx, self.gui_range_count.get(), self.dpm_tabulka.vx))
            elif self.bool_multihisto_vy1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vy, self.gui_range_count.get(), self.dpm_tabulka.vx))
            elif self.bool_multihisto_vz1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vz, self.gui_range_count.get(), self.dpm_tabulka.vx))
            elif self.bool_multihisto_vmag1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vmag, self.gui_range_count.get(), self.dpm_tabulka.vx))
            elif self.bool_multihisto_diameter1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.diameter, self.gui_range_count.get(), self.dpm_tabulka.vx))
            elif self.bool_multihisto_temperature1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.temperature, self.gui_range_count.get(), self.dpm_tabulka.vx))
            elif self.bool_multihisto_mfr1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mfr, self.gui_range_count.get(), self.dpm_tabulka.vx))
            elif self.bool_multihisto_mass1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mass, self.gui_range_count.get(), self.dpm_tabulka.vx))            
            elif self.bool_multihisto_time1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.time, self.gui_range_count.get(), self.dpm_tabulka.vx))      
            else:
                pass

    def multihistos_filter_vy2(self):
        if self.bool_multihisto_vy2.get() == True:
            if self.bool_multihisto_x1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.x, self.gui_range_count.get(), self.dpm_tabulka.vy))
            elif self.bool_multihisto_y1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.y, self.gui_range_count.get(), self.dpm_tabulka.vy))
            elif self.bool_multihisto_z1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.z, self.gui_range_count.get(), self.dpm_tabulka.vy))
            elif self.bool_multihisto_vx1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vx, self.gui_range_count.get(), self.dpm_tabulka.vy))
            elif self.bool_multihisto_vy1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vy, self.gui_range_count.get(), self.dpm_tabulka.vy))
            elif self.bool_multihisto_vz1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vz, self.gui_range_count.get(), self.dpm_tabulka.vy))
            elif self.bool_multihisto_vmag1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vmag, self.gui_range_count.get(), self.dpm_tabulka.vy))
            elif self.bool_multihisto_diameter1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.diameter, self.gui_range_count.get(), self.dpm_tabulka.vy))
            elif self.bool_multihisto_temperature1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.temperature, self.gui_range_count.get(), self.dpm_tabulka.vy))
            elif self.bool_multihisto_mfr1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mfr, self.gui_range_count.get(), self.dpm_tabulka.vy))
            elif self.bool_multihisto_mass1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mass, self.gui_range_count.get(), self.dpm_tabulka.vy))            
            elif self.bool_multihisto_time1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.time, self.gui_range_count.get(), self.dpm_tabulka.vy))      
            else:
                pass

    def multihistos_filter_vz2(self):
        if self.bool_multihisto_vz2.get() == True:
            if self.bool_multihisto_x1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.x, self.gui_range_count.get(), self.dpm_tabulka.vz))
            elif self.bool_multihisto_y1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.y, self.gui_range_count.get(), self.dpm_tabulka.vz))
            elif self.bool_multihisto_z1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.z, self.gui_range_count.get(), self.dpm_tabulka.vz))
            elif self.bool_multihisto_vx1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vx, self.gui_range_count.get(), self.dpm_tabulka.vz))
            elif self.bool_multihisto_vy1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vy, self.gui_range_count.get(), self.dpm_tabulka.vz))
            elif self.bool_multihisto_vz1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vz, self.gui_range_count.get(), self.dpm_tabulka.vz))
            elif self.bool_multihisto_vmag1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vmag, self.gui_range_count.get(), self.dpm_tabulka.vz))
            elif self.bool_multihisto_diameter1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.diameter, self.gui_range_count.get(), self.dpm_tabulka.vz))
            elif self.bool_multihisto_temperature1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.temperature, self.gui_range_count.get(), self.dpm_tabulka.vz))
            elif self.bool_multihisto_mfr1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mfr, self.gui_range_count.get(), self.dpm_tabulka.vz))
            elif self.bool_multihisto_mass1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mass, self.gui_range_count.get(), self.dpm_tabulka.vz))            
            elif self.bool_multihisto_time1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.time, self.gui_range_count.get(), self.dpm_tabulka.vz))      
            else:
                pass

    def multihistos_filter_vmag2(self):
        if self.bool_multihisto_vmag2.get() == True:
            if self.bool_multihisto_x1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.x, self.gui_range_count.get(), self.dpm_tabulka.vmag))
            elif self.bool_multihisto_y1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.y, self.gui_range_count.get(), self.dpm_tabulka.vmag))
            elif self.bool_multihisto_z1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.z, self.gui_range_count.get(), self.dpm_tabulka.vmag))
            elif self.bool_multihisto_vx1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vx, self.gui_range_count.get(), self.dpm_tabulka.vmag))
            elif self.bool_multihisto_vy1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vy, self.gui_range_count.get(), self.dpm_tabulka.vmag))
            elif self.bool_multihisto_vz1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vz, self.gui_range_count.get(), self.dpm_tabulka.vmag))
            elif self.bool_multihisto_vmag1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vmag, self.gui_range_count.get(), self.dpm_tabulka.vmag))
            elif self.bool_multihisto_diameter1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.diameter, self.gui_range_count.get(), self.dpm_tabulka.vmag))
            elif self.bool_multihisto_temperature1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.temperature, self.gui_range_count.get(), self.dpm_tabulka.vmag))
            elif self.bool_multihisto_mfr1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mfr, self.gui_range_count.get(), self.dpm_tabulka.vmag))
            elif self.bool_multihisto_mass1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mass, self.gui_range_count.get(), self.dpm_tabulka.vmag))            
            elif self.bool_multihisto_time1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.time, self.gui_range_count.get(), self.dpm_tabulka.vmag))      
            else:
                pass

    def multihistos_filter_diameter2(self):
        if self.bool_multihisto_diameter2.get() == True:
            if self.bool_multihisto_x1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.x, self.gui_range_count.get(), self.dpm_tabulka.diameter))
            elif self.bool_multihisto_y1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.y, self.gui_range_count.get(), self.dpm_tabulka.diameter))
            elif self.bool_multihisto_z1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.z, self.gui_range_count.get(), self.dpm_tabulka.diameter))
            elif self.bool_multihisto_vx1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vx, self.gui_range_count.get(), self.dpm_tabulka.diameter))
            elif self.bool_multihisto_vy1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vy, self.gui_range_count.get(), self.dpm_tabulka.diameter))
            elif self.bool_multihisto_vz1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vz, self.gui_range_count.get(), self.dpm_tabulka.diameter))
            elif self.bool_multihisto_vmag1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vmag, self.gui_range_count.get(), self.dpm_tabulka.diameter))
            elif self.bool_multihisto_diameter1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.diameter, self.gui_range_count.get(), self.dpm_tabulka.diameter))
            elif self.bool_multihisto_temperature1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.temperature, self.gui_range_count.get(), self.dpm_tabulka.diameter))
            elif self.bool_multihisto_mfr1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mfr, self.gui_range_count.get(), self.dpm_tabulka.diameter))
            elif self.bool_multihisto_mass1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mass, self.gui_range_count.get(), self.dpm_tabulka.diameter))            
            elif self.bool_multihisto_time1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.time, self.gui_range_count.get(), self.dpm_tabulka.diameter))      
            else:
                pass

    def multihistos_filter_temperature2(self):
        if self.bool_multihisto_temperature2.get() == True:
            if self.bool_multihisto_x1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.x, self.gui_range_count.get(), self.dpm_tabulka.time))
            elif self.bool_multihisto_y1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.y, self.gui_range_count.get(), self.dpm_tabulka.time))
            elif self.bool_multihisto_z1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.z, self.gui_range_count.get(), self.dpm_tabulka.time))
            elif self.bool_multihisto_vx1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vx, self.gui_range_count.get(), self.dpm_tabulka.time))
            elif self.bool_multihisto_vy1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vy, self.gui_range_count.get(), self.dpm_tabulka.time))
            elif self.bool_multihisto_vz1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vz, self.gui_range_count.get(), self.dpm_tabulka.time))
            elif self.bool_multihisto_vmag1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vmag, self.gui_range_count.get(), self.dpm_tabulka.time))
            elif self.bool_multihisto_diameter1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.diameter, self.gui_range_count.get(), self.dpm_tabulka.time))
            elif self.bool_multihisto_temperature1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.temperature, self.gui_range_count.get(), self.dpm_tabulka.time))
            elif self.bool_multihisto_mfr1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mfr, self.gui_range_count.get(), self.dpm_tabulka.time))
            elif self.bool_multihisto_mass1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mass, self.gui_range_count.get(), self.dpm_tabulka.time))            
            elif self.bool_multihisto_time1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.time, self.gui_range_count.get(), self.dpm_tabulka.time))      
            else:
                pass

    def multihistos_filter_mfr2(self):
        if self.bool_multihisto_mfr2.get() == True:
            if self.bool_multihisto_x1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.x, self.gui_range_count.get(), self.dpm_tabulka.mfr))
            elif self.bool_multihisto_y1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.y, self.gui_range_count.get(), self.dpm_tabulka.mfr))
            elif self.bool_multihisto_z1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.z, self.gui_range_count.get(), self.dpm_tabulka.mfr))
            elif self.bool_multihisto_vx1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vx, self.gui_range_count.get(), self.dpm_tabulka.mfr))
            elif self.bool_multihisto_vy1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vy, self.gui_range_count.get(), self.dpm_tabulka.mfr))
            elif self.bool_multihisto_vz1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vz, self.gui_range_count.get(), self.dpm_tabulka.mfr))
            elif self.bool_multihisto_vmag1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vmag, self.gui_range_count.get(), self.dpm_tabulka.mfr))
            elif self.bool_multihisto_diameter1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.diameter, self.gui_range_count.get(), self.dpm_tabulka.mfr))
            elif self.bool_multihisto_temperature1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.temperature, self.gui_range_count.get(), self.dpm_tabulka.mfr))
            elif self.bool_multihisto_mfr1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mfr, self.gui_range_count.get(), self.dpm_tabulka.mfr))
            elif self.bool_multihisto_mass1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mass, self.gui_range_count.get(), self.dpm_tabulka.mfr))            
            elif self.bool_multihisto_time1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.time, self.gui_range_count.get(), self.dpm_tabulka.mfr))      
            else:
                pass

    def multihistos_filter_mass2(self):
        if self.bool_multihisto_mass2.get() == True:
            if self.bool_multihisto_x1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.x, self.gui_range_count.get(), self.dpm_tabulka.mass))
            elif self.bool_multihisto_y1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.y, self.gui_range_count.get(), self.dpm_tabulka.mass))
            elif self.bool_multihisto_z1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.z, self.gui_range_count.get(), self.dpm_tabulka.mass))
            elif self.bool_multihisto_vx1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vx, self.gui_range_count.get(), self.dpm_tabulka.mass))
            elif self.bool_multihisto_vy1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vy, self.gui_range_count.get(), self.dpm_tabulka.mass))
            elif self.bool_multihisto_vz1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vz, self.gui_range_count.get(), self.dpm_tabulka.mass))
            elif self.bool_multihisto_vmag1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vmag, self.gui_range_count.get(), self.dpm_tabulka.mass))
            elif self.bool_multihisto_diameter1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.diameter, self.gui_range_count.get(), self.dpm_tabulka.mass))
            elif self.bool_multihisto_temperature1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.temperature, self.gui_range_count.get(), self.dpm_tabulka.mass))
            elif self.bool_multihisto_mfr1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mfr, self.gui_range_count.get(), self.dpm_tabulka.mass))
            elif self.bool_multihisto_mass1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mass, self.gui_range_count.get(), self.dpm_tabulka.mass))            
            elif self.bool_multihisto_time1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.time, self.gui_range_count.get(), self.dpm_tabulka.mass))      
            else:
                pass

    def multihistos_filter_time2(self):
        if self.bool_multihisto_time2.get() == True:
            if self.bool_multihisto_x1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.x, self.gui_range_count.get(), self.dpm_tabulka.time))
            elif self.bool_multihisto_y1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.y, self.gui_range_count.get(), self.dpm_tabulka.time))
            elif self.bool_multihisto_z1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.z, self.gui_range_count.get(), self.dpm_tabulka.time))
            elif self.bool_multihisto_vx1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vx, self.gui_range_count.get(), self.dpm_tabulka.time))
            elif self.bool_multihisto_vy1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vy, self.gui_range_count.get(), self.dpm_tabulka.time))
            elif self.bool_multihisto_vz1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vz, self.gui_range_count.get(), self.dpm_tabulka.time))
            elif self.bool_multihisto_vmag1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.vmag, self.gui_range_count.get(), self.dpm_tabulka.time))
            elif self.bool_multihisto_diameter1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.diameter, self.gui_range_count.get(), self.dpm_tabulka.time))
            elif self.bool_multihisto_temperature1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.temperature, self.gui_range_count.get(), self.dpm_tabulka.time))
            elif self.bool_multihisto_mfr1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mfr, self.gui_range_count.get(), self.dpm_tabulka.time))
            elif self.bool_multihisto_mass1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.mass, self.gui_range_count.get(), self.dpm_tabulka.time))            
            elif self.bool_multihisto_time1.get() == True:
                self.dpm_tabulka.dpmhisto_multiple(self.dpm_tabulka.auto_extractor_dict(self.dpm_tabulka.time, self.gui_range_count.get(), self.dpm_tabulka.time))      
            else:
                pass
    #endregion multihistos_plotters  



    def accept_ymin_scatter(self):
        numbers = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "e", ".", "-"]       
        for char in self.text_input_scatter_ylim_min.get(): 
            if char in numbers:
                self.button_scatter_ylim_min_confirm.config(background=self.green, text = "Schváleno", state = "disabled")
                self.button_scatter_ylim_lim_reset.config(state = "normal")
                self.text_input_scatter_ylim_min.config(state = "disabled")
            else:
                print("Toto není číslo")
                self.button_scatter_ylim_min_confirm.config(background=self.red, text = "Znovu! To není číslo"
                )
                self.text_input_scatter_ylim_min.config(state = "normal")
                self.button_scatter_ylim_lim_reset.config(state = "disabled")
                self.text_input_scatter_ylim_min.delete(0, 'end')
                self.text_input_scatter_ylim_min.get()
                self.button_scatter_ylim_min_confirm.config(state = "normal")
                break

        if self.text_input_scatter_ylim_min.get() != "":
            self.gui_scatter_ylim_min.set(float((self.text_input_scatter_ylim_min.get())))
            self.dpm_tabulka.ymin_scatter               = float((self.text_input_scatter_ylim_min.get()))
        
        if self.button_scatter_ylim_min_confirm["background"] == self.green and self.button_scatter_ylim_max_confirm["background"] == self.green:
            self.dpm_tabulka.scatter_ylim_bool = True
            self.dpm_tabulka.ymin_scatter               = float((self.text_input_scatter_ylim_min.get()))

    def accept_ymax_scatter(self):
        numbers = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "e", ".", "-"]       
        for char in self.text_input_scatter_ylim_max.get(): 
            if char in numbers:
                self.button_scatter_ylim_max_confirm.config(background=self.green, text = "Schváleno", state = "disabled")
                self.button_scatter_ylim_lim_reset.config(state = "normal")
                self.text_input_scatter_ylim_max.config(state = "disabled")
            else:
                print("Toto není číslo")
                self.button_scatter_ylim_max_confirm.config(background=self.red, text = "Znovu! To není číslo")
                self.text_input_scatter_ylim_max.config(state = "normal")
                self.button_scatter_ylim_lim_reset.config(state = "disabled")
                self.text_input_scatter_ylim_max.delete(0, 'end')
                self.text_input_scatter_ylim_max.get()
                self.button_scatter_ylim_max_confirm.config(state = "normal")
                break

        if self.text_input_scatter_ylim_max.get() == "":
            self.button_next.config(state = "disabled")
        else:
            self.button_next.config(state = "normal")
            self.gui_scatter_ylim_max.set(float((self.text_input_scatter_ylim_max.get())))
            self.dpm_tabulka.ymax_scatter               = float((self.text_input_scatter_ylim_max.get()))

        if self.button_scatter_ylim_min_confirm["background"] == self.green and self.button_scatter_ylim_max_confirm["background"] == self.green:
            self.dpm_tabulka.scatter_ylim_bool = True
            self.dpm_tabulka.ymax_scatter               = float((self.text_input_scatter_ylim_max.get()))

    def accept_ylim_scatter(self):
        if self.gui_scatter_ylim_bool.get() == False:
            self.gui_scatter_ylim_bool.set(True)
            self.text_input_scatter_ylim_min.config(state = "disabled")
            self.text_input_scatter_ylim_max.config(state = "disabled")
            self.button_scatter_ylim_lim_reset.config(state = "disabled")
            self.button_scatter_ylim_min_confirm.config(state = "disabled")
            self.button_scatter_ylim_max_confirm.config(state = "disabled")
            self.button_scatter_ylim_lim_allow.config(state = "normal", background= self.green, text = "Automat")
            self.dpm_tabulka.scatter_ylim_bool = False
        else:
            self.gui_scatter_ylim_bool.set(False)
            if self.button_scatter_ylim_min_confirm["background"] != self.green or self.button_scatter_ylim_max_confirm["background"] != self.green:
                self.text_input_scatter_ylim_min.config(state = "normal")
                self.text_input_scatter_ylim_max.config(state = "normal")
                self.button_scatter_ylim_min_confirm.config(state = "normal")
                self.button_scatter_ylim_max_confirm.config(state = "normal")
            self.button_scatter_ylim_lim_reset.config(state = "normal")
            self.button_scatter_ylim_lim_allow.config(state = "normal", background= self.grey, text = "Manuál")
            self.dpm_tabulka.scatter_ylim_bool = True
        print("Multihisto xlim bool? {}".format(self.gui_scatter_ylim_bool.get()))
        print("Multihisto xlim zapnuto? {}".format(self.dpm_tabulka.scatter_ylim_bool))



    def accept_xmin_scatter(self):
        numbers = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "e", ".", "-"]       
        for char in self.text_input_scatter_xmin.get(): 
            if char in numbers:
                self.button_scatter_xmin_confirm.config(background=self.green, text = "Schváleno", state = "disabled")
                self.button_scatter_xlim_reset.config(state = "normal")
                self.text_input_scatter_xmin.config(state = "disabled")
            else:
                print("Toto není číslo")
                self.button_scatter_xmin_confirm.config(background=self.red, text = "Znovu! To není číslo")
                self.text_input_scatter_xmin.config(state = "normal")
                self.button_scatter_xlim_reset.config(state = "disabled")
                self.text_input_scatter_xmin.delete(0, 'end')
                self.text_input_scatter_xmin.get()
                self.button_scatter_xmin_confirm.config(state = "normal")
                break

        if self.text_input_scatter_xmin.get() != "":
            self.gui_scatter_xmin.set(float((self.text_input_scatter_xmin.get())))
            self.dpm_tabulka.xmin_scatter               = float((self.text_input_scatter_xmin.get()))

        if self.button_scatter_xmin_confirm["background"] == self.green and self.button_scatter_xmax_confirm["background"] == self.green:
            self.dpm_tabulka.scatter_xlim_bool = True
        self.dpm_tabulka.xmin_scatter               = float((self.text_input_scatter_xmin.get()))

    def accept_xmax_scatter(self):
        numbers = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "e", ".", "-"]       
        for char in self.text_input_scatter_xmax.get(): 
            if char in numbers:
                self.button_scatter_xmax_confirm.config(background=self.green, text = "Schváleno", state = "disabled")
                self.button_scatter_xlim_reset.config(state = "normal")
                self.text_input_scatter_xmax.config(state = "disabled")
            else:
                print("Toto není číslo")
                self.button_scatter_xmax_confirm.config(background=self.red, text = "Znovu! To není číslo")
                self.text_input_scatter_xmax.config(state = "normal")
                self.button_scatter_xlim_reset.config(state = "disabled")
                self.text_input_scatter_xmax.delete(0, 'end')
                self.text_input_scatter_xmax.get()
                self.button_scatter_xmax_confirm.config(state = "normal")
                break

        if self.text_input_scatter_xmax.get() != "":
            self.gui_scatter_xmax.set(float((self.text_input_scatter_xmax.get())))
            self.dpm_tabulka.xmax_scatter               = float((self.text_input_scatter_xmax.get()))

        if self.button_scatter_xmin_confirm["background"] == self.green and self.button_scatter_xmax_confirm["background"] == self.green:
            self.dpm_tabulka.scatter_xlim_bool = True
        self.dpm_tabulka.xmax_scatter               = float((self.text_input_scatter_xmax.get()))

    def accept_xlim_scatter(self):
        if self.gui_scatter_xlim_bool.get() == False:
            self.gui_scatter_xlim_bool.set(True)
            self.text_input_scatter_xmin.config(state = "disabled")
            self.text_input_scatter_xmax.config(state = "disabled")
            self.button_scatter_xlim_reset.config(state = "disabled")
            self.button_scatter_xmin_confirm.config(state = "disabled")
            self.button_scatter_xmax_confirm.config(state = "disabled")
            self.button_scatter_xlim_allow.config(state = "normal", background= self.green, text = "Automat")
            self.dpm_tabulka.scatter_xlim_bool = False
        else:
            self.gui_scatter_xlim_bool.set(False)
            if self.button_scatter_xmin_confirm["background"] != self.green or self.button_scatter_xmax_confirm["background"] != self.green:
                self.text_input_scatter_xmin.config(state = "normal")
                self.text_input_scatter_xmax.config(state = "normal")
                self.button_scatter_xmin_confirm.config(state = "normal")
                self.button_scatter_xmax_confirm.config(state = "normal")
            self.button_scatter_xlim_reset.config(state = "normal")
            self.button_scatter_xlim_allow.config(state = "normal", background= self.grey, text = "Manuál")
            self.dpm_tabulka.scatter_xlim_bool = True
        print("Multihisto xlim bool? {}".format(self.gui_scatter_xlim_bool.get()))
        print("Multihisto xlim zapnuto? {}".format(self.dpm_tabulka.scatter_xlim_bool))

    def scatter_enabler(self):
        condition_range_auto    = self.button_scatter_ylim_lim_allow["text"]   == "Automat"
        condition_xlim_auto     = self.button_scatter_xlim_allow["text"]        == "Automat"
        condition_range_manual  = self.button_scatter_ylim_lim_allow["text"]   == "Manuál"
        condition_xlim_manual  = self.button_scatter_xlim_allow["text"]        == "Manuál"
        condition_buttons_xlim  = self.button_scatter_xmin_confirm["background"] == self.green and self.button_scatter_xmax_confirm["background"] == self.green
        condition_buttons_range = self.button_scatter_ylim_min_confirm["background"] == self.green and self.button_scatter_ylim_max_confirm["background"] == self.green

        if condition_range_auto and condition_xlim_auto:
            self.button_scatter_x3.config(state= "normal")
            self.button_scatter_y3.config(state= "normal")
            self.button_scatter_z3.config(state= "normal")
            self.button_scatter_vx3.config(state= "normal")
            self.button_scatter_vy3.config(state= "normal")
            self.button_scatter_vz3.config(state= "normal")
            self.button_scatter_vmag3.config(state= "normal")
            self.button_scatter_diameter3.config(state= "normal")
            self.button_scatter_temperature3.config(state= "normal")
            self.button_scatter_mfr3.config(state= "normal")
            self.button_scatter_mass3.config(state= "normal")
            self.button_scatter_time3.config(state= "normal")
            self.button_scatter_none3.config(state= "normal")

        elif condition_range_auto and condition_xlim_manual and condition_buttons_xlim:
            self.button_scatter_x3.config(state= "normal")
            self.button_scatter_y3.config(state= "normal")
            self.button_scatter_z3.config(state= "normal")
            self.button_scatter_vx3.config(state= "normal")
            self.button_scatter_vy3.config(state= "normal")
            self.button_scatter_vz3.config(state= "normal")
            self.button_scatter_vmag3.config(state= "normal")
            self.button_scatter_diameter3.config(state= "normal")
            self.button_scatter_temperature3.config(state= "normal")
            self.button_scatter_mfr3.config(state= "normal")
            self.button_scatter_mass3.config(state= "normal")
            self.button_scatter_time3.config(state= "normal")
            self.button_scatter_none3.config(state= "normal")

        elif condition_range_manual and condition_xlim_auto and condition_buttons_range:
            self.button_scatter_x3.config(state= "normal")
            self.button_scatter_y3.config(state= "normal")
            self.button_scatter_z3.config(state= "normal")
            self.button_scatter_vx3.config(state= "normal")
            self.button_scatter_vy3.config(state= "normal")
            self.button_scatter_vz3.config(state= "normal")
            self.button_scatter_vmag3.config(state= "normal")
            self.button_scatter_diameter3.config(state= "normal")
            self.button_scatter_temperature3.config(state= "normal")
            self.button_scatter_mfr3.config(state= "normal")
            self.button_scatter_mass3.config(state= "normal")
            self.button_scatter_time3.config(state= "normal")
            self.button_scatter_none3.config(state= "normal")                  

        elif condition_range_manual and condition_xlim_manual and condition_buttons_xlim and condition_buttons_range:
            self.button_scatter_x3.config(state= "normal")
            self.button_scatter_y3.config(state= "normal")
            self.button_scatter_z3.config(state= "normal")
            self.button_scatter_vx3.config(state= "normal")
            self.button_scatter_vy3.config(state= "normal")
            self.button_scatter_vz3.config(state= "normal")
            self.button_scatter_vmag3.config(state= "normal")
            self.button_scatter_diameter3.config(state= "normal")
            self.button_scatter_temperature3.config(state= "normal")
            self.button_scatter_mfr3.config(state= "normal")
            self.button_scatter_mass3.config(state= "normal")
            self.button_scatter_time3.config(state= "normal")
            self.button_scatter_none3.config(state= "normal")

        else:
            self.button_scatter_x3.config(state= "disabled")
            self.button_scatter_y3.config(state= "disabled")
            self.button_scatter_z3.config(state= "disabled")
            self.button_scatter_vx3.config(state= "disabled")
            self.button_scatter_vy3.config(state= "disabled")
            self.button_scatter_vz3.config(state= "disabled")
            self.button_scatter_vmag3.config(state= "disabled")
            self.button_scatter_diameter3.config(state= "disabled")
            self.button_scatter_temperature3.config(state= "disabled")
            self.button_scatter_mfr3.config(state= "disabled")
            self.button_scatter_mass3.config(state= "disabled")
            self.button_scatter_time3.config(state= "disabled")
            self.button_scatter_none3.config(state= "disabled")          





okno = gui()
okno.display()
remove(okno.dpm_file.get())


#(self, name, location, file, positionunit, velocityunit, diameterunit, temperatureunit, mfrunit, massunit, frequencyunit, timeunit):
#pepa = tabulka ("Tabulka", "D:/DPM", "150_10_35_40.dpmptcl", "m", "m/s", "m", "k", "kg/s", "kg", "1/s", "s")
#pepa.table_spawn_csv()
